import sqlite3
import csv
from datetime import date, timedelta, datetime
import os # For the demo part

class InventoryManager:
    def __init__(self, db_filepath="inventory.db"):
        self.db_filepath = db_filepath
        self.conn = None  # For persistent in-memory connection
        if self.db_filepath == ":memory:":
            self.conn = sqlite3.connect(":memory:")
            self.conn.row_factory = sqlite3.Row
        self._initialize_db()
        # Garden produce is not part of this class for now
        # self.my_garden_produce = []

    def _get_db_connection(self):
        """Establishes and returns a database connection."""
        if self.conn and self.db_filepath == ":memory:":
            return self.conn
        # For file-based databases, create a new connection each time
        conn = sqlite3.connect(self.db_filepath)
        conn.row_factory = sqlite3.Row # Access columns by name
        return conn

    def close_connection(self):
        """Closes the persistent connection if it exists (mainly for in-memory DBs)."""
        if self.conn and self.db_filepath == ":memory:":
            self.conn.close()
            self.conn = None

    def _initialize_db(self):
        """Creates database tables if they don't already exist."""
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()

                # Products Table
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS products (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT NOT NULL UNIQUE,
                        unit_of_measure TEXT NOT NULL,
                        default_expiry_days INTEGER NOT NULL,
                        par_level REAL DEFAULT 0,
                        max_holding_amount REAL DEFAULT 0,
                        purchase_location TEXT,
                        consumption_override_rate REAL DEFAULT NULL,
                        category_id INTEGER,
                        subcategory_id INTEGER,
                        FOREIGN KEY (category_id) REFERENCES categories (id),
                        FOREIGN KEY (subcategory_id) REFERENCES subcategories (id)
                    )
                ''')

                # Current Inventory Table
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS inventory_items (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        product_id INTEGER,
                        name TEXT NOT NULL,
                        quantity TEXT NOT NULL, 
                        purchase_date TEXT NOT NULL,
                        expiry_date TEXT NOT NULL,
                        original_quantity_string TEXT,
                        FOREIGN KEY (product_id) REFERENCES products (id)
                    )
                ''')
                # Historical (Consumed) Items Table
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS historical_items (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        product_id INTEGER,
                        name TEXT NOT NULL,
                        quantity_consumed_this_time REAL NOT NULL,
                        original_quantity_string TEXT,
                        purchase_date TEXT, 
                        expiry_date TEXT,
                        consumed_date TEXT NOT NULL,
                        cost_of_goods_used REAL,
                        FOREIGN KEY (product_id) REFERENCES products (id)
                    )
                ''')

                # PurchaseLog Table
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS PurchaseLog (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        product_id INTEGER,
                        purchase_date TEXT NOT NULL,
                        quantity_purchased REAL NOT NULL,
                        cost_per_unit REAL NOT NULL,
                        vendor TEXT,
                        FOREIGN KEY (product_id) REFERENCES products (id)
                    )
                ''')

                # Production Items Table
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS production_items (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT NOT NULL,
                        associated_product_id INTEGER,
                        plant_date TEXT,
                        time_to_harvest_days INTEGER,
                        expected_harvest_period_days INTEGER,
                        expected_yield_total REAL,
                        status TEXT CHECK(status IN ('Growing', 'Harvesting', 'Finished')),
                        FOREIGN KEY (associated_product_id) REFERENCES products (id)
                    )
                ''')

                # Categories Table
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS categories (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT NOT NULL UNIQUE
                    )
                ''')

                # Subcategories Table
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS subcategories (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT NOT NULL,
                        category_id INTEGER NOT NULL,
                        FOREIGN KEY (category_id) REFERENCES categories (id),
                        UNIQUE (name, category_id)
                    )
                ''')
                conn.commit()
        except sqlite3.Error as e:
            raise sqlite3.Error(f"Database initialization error: {e}")
            # Depending on app design, might want to raise this or handle more gracefully

    def migrate_text_categories_to_ids(self):
        """
        Simulates migration of products with old text-based category/subcategory
        to new ID-based category_id/subcategory_id.
        This method is intended for demonstration and setup.
        """
        print("Starting migration of text categories to IDs...")
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()

                # --- a. Simulate Pre-existing Schema and Data ---
                try:
                    # Add old columns if they don't exist (for simulation)
                    # Note: In a real migration, these columns would already exist.
                    # This is to make the script runnable for demonstration even if DB is new.
                    cursor.execute("SELECT old_category_text FROM products LIMIT 1")
                except sqlite3.OperationalError: # Column doesn't exist
                    print("Simulating old schema: Adding old_category_text and old_subcategory_text columns...")
                    conn.executescript('''
                        ALTER TABLE products ADD COLUMN old_category_text TEXT;
                        ALTER TABLE products ADD COLUMN old_subcategory_text TEXT;
                    ''')
                    # Insert sample data only if we just added the columns (i.e., fresh simulation)
                    print("Inserting sample products with text categories for migration...")
                    sample_products = [
                        ("Apple", "pcs", 14, "Produce", "Fruit"),
                        ("Milk", "liter", 7, "Dairy", None),
                        ("Carrot", "kg", 21, "Produce", "Vegetable"),
                        ("Yogurt", "pcs", 10, "Dairy", "Cultured"),
                        ("Orange Juice", "liter", 7, "Beverages", None),
                        ("Banana", "pcs", 5, "Produce", "Fruit") # Another Produce/Fruit
                    ]
                    for p_name, uom, exp_days, cat_text, subcat_text in sample_products:
                        # Check if product already exists by name to avoid IntegrityError during simulation
                        cursor.execute("SELECT id FROM products WHERE name = ?", (p_name,))
                        if not cursor.fetchone():
                            cursor.execute("""
                                INSERT INTO products (name, unit_of_measure, default_expiry_days, old_category_text, old_subcategory_text, category_id, subcategory_id)
                                VALUES (?, ?, ?, ?, ?, NULL, NULL)
                            """, (p_name, uom, exp_days, cat_text, subcat_text))
                        else:
                            # If product exists, update its old text fields for migration simulation
                            cursor.execute("""
                                UPDATE products SET old_category_text = ?, old_subcategory_text = ?
                                WHERE name = ? AND category_id IS NULL
                            """, (cat_text, subcat_text, p_name))
                    conn.commit()
                    print("Sample data inserted/updated for old text categories.")

                # --- b. Extract Distinct Categories and Populate categories Table ---
                print("Populating 'categories' table from old_category_text...")
                cursor.execute("SELECT DISTINCT old_category_text FROM products WHERE old_category_text IS NOT NULL")
                distinct_categories = [row['old_category_text'] for row in cursor.fetchall()]

                for cat_name in distinct_categories:
                    add_cat_result = self.add_category(cat_name) # Uses existing method, handles duplicates
                    if add_cat_result['success']:
                        print(f"Category '{cat_name}' (ID: {add_cat_result.get('category_id')}) processed/ensured in categories table.")
                    elif "already exists" in add_cat_result.get('message',''):
                         print(f"Category '{cat_name}' already exists, skipping addition.")
                    else:
                        print(f"Warning: Could not add category '{cat_name}': {add_cat_result.get('message')}")
                conn.commit()

                # --- c. Extract Distinct Subcategories and Populate subcategories Table ---
                print("Populating 'subcategories' table...")
                cursor.execute("""
                    SELECT DISTINCT old_category_text, old_subcategory_text
                    FROM products
                    WHERE old_category_text IS NOT NULL AND old_subcategory_text IS NOT NULL
                """)
                distinct_subcategories = cursor.fetchall()

                for row in distinct_subcategories:
                    old_cat_text = row['old_category_text']
                    old_subcat_text = row['old_subcategory_text']

                    category_obj = self.get_category_by_name(old_cat_text)
                    if category_obj:
                        cat_id = category_obj['id']
                        add_subcat_result = self.add_subcategory(old_subcat_text, cat_id) # Handles duplicates
                        if add_subcat_result['success']:
                            print(f"Subcategory '{old_subcat_text}' (ID: {add_subcat_result.get('subcategory_id')}) under Category '{old_cat_text}' processed.")
                        elif "already exists" in add_subcat_result.get('message',''):
                            print(f"Subcategory '{old_subcat_text}' under '{old_cat_text}' already exists.")
                        else:
                             print(f"Warning: Could not add subcategory '{old_subcat_text}' for category '{old_cat_text}': {add_subcat_result.get('message')}")
                    else:
                        print(f"Warning: Category '{old_cat_text}' not found for subcategory '{old_subcat_text}'. Skipping subcategory.")
                conn.commit()

                # --- d. Update products Table with category_id and subcategory_id ---
                print("Updating 'products' table with category_id and subcategory_id...")
                cursor.execute("SELECT id, old_category_text, old_subcategory_text FROM products WHERE category_id IS NULL")
                products_to_update = cursor.fetchall()

                for product_row in products_to_update:
                    prod_id = product_row['id']
                    cat_id_to_set = None
                    sub_cat_id_to_set = None

                    if product_row['old_category_text']:
                        category_data = self.get_category_by_name(product_row['old_category_text'])
                        if category_data:
                            cat_id_to_set = category_data['id']
                            if product_row['old_subcategory_text']:
                                subcategory_data = self.get_subcategory_by_name_and_category_id(product_row['old_subcategory_text'], cat_id_to_set)
                                if subcategory_data:
                                    sub_cat_id_to_set = subcategory_data['id']
                                else:
                                    print(f"Warning: Subcategory '{product_row['old_subcategory_text']}' not found in DB for product ID {prod_id}. Subcategory ID will be NULL.")
                        else:
                            print(f"Warning: Category '{product_row['old_category_text']}' not found in DB for product ID {prod_id}. Category ID will be NULL.")

                    if cat_id_to_set is not None:
                        cursor.execute("""
                            UPDATE products SET category_id = ?, subcategory_id = ?
                            WHERE id = ?
                        """, (cat_id_to_set, sub_cat_id_to_set, prod_id))
                        print(f"Product ID {prod_id} updated with CategoryID: {cat_id_to_set}, SubcategoryID: {sub_cat_id_to_set}")
                conn.commit()

                # --- e. Simulate Dropping Old Columns ---
                # In a real scenario, backup before dropping.
                # SQLite's DROP COLUMN is only supported in version 3.35.0+
                # This might fail in older versions.
                print("Attempting to drop old text category columns (simulation)...")
                try:
                    conn.executescript('''
                        ALTER TABLE products DROP COLUMN old_category_text;
                        ALTER TABLE products DROP COLUMN old_subcategory_text;
                    ''')
                    print("Successfully dropped old_category_text and old_subcategory_text columns.")
                except sqlite3.OperationalError as e_drop:
                    print(f"Could not drop old columns (this is expected if SQLite version < 3.35.0 or columns already dropped): {e_drop}")

                conn.commit()
                return {"success": True, "message": "Migration of text categories to IDs completed."}

        except sqlite3.Error as e:
            return {"success": False, "message": f"Migration failed: {e}"}


    # --- Category and Subcategory Management Methods ---
    def add_category(self, name):
        """Adds a new category to the categories table."""
        if not name or not isinstance(name, str) or not name.strip():
            return {"success": False, "message": "Category name must be a non-empty string."}
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("INSERT INTO categories (name) VALUES (?)", (name.strip(),))
                conn.commit()
                category_id = cursor.lastrowid
                return {"success": True, "message": f"Category '{name.strip()}' added successfully.", "category_id": category_id}
        except sqlite3.IntegrityError:
            return {"success": False, "message": f"Category name '{name.strip()}' already exists."}
        except sqlite3.Error as e:
            return {"success": False, "message": f"Database error adding category: {e}"}

    def add_subcategory(self, name, category_id):
        """Adds a new subcategory to the subcategories table."""
        if not name or not isinstance(name, str) or not name.strip():
            return {"success": False, "message": "Subcategory name must be a non-empty string."}
        if not isinstance(category_id, int):
             return {"success": False, "message": "Category ID must be an integer."}

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                # Check if category_id exists
                cursor.execute("SELECT id FROM categories WHERE id = ?", (category_id,))
                if not cursor.fetchone():
                    return {"success": False, "message": f"Parent category with ID {category_id} not found."}

                cursor.execute("INSERT INTO subcategories (name, category_id) VALUES (?, ?)", (name.strip(), category_id))
                conn.commit()
                subcategory_id = cursor.lastrowid
                return {"success": True, "message": f"Subcategory '{name.strip()}' added successfully under category ID {category_id}.", "subcategory_id": subcategory_id}
        except sqlite3.IntegrityError:
            # This can be due to UNIQUE constraint on (name, category_id) or FOREIGN KEY constraint (though checked above)
            return {"success": False, "message": f"Subcategory '{name.strip()}' already exists for category ID {category_id}, or foreign key constraint failed."}
        except sqlite3.Error as e:
            return {"success": False, "message": f"Database error adding subcategory: {e}"}

    def get_all_categories_with_subcategories(self):
        """
        Retrieves all categories and their subcategories, ordered by name.
        """
        categories_list = []
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, name FROM categories ORDER BY name ASC")
                categories_rows = cursor.fetchall()

                for cat_row in categories_rows:
                    category_dict = dict(cat_row)
                    category_dict['subcategories'] = []

                    cursor.execute("""
                        SELECT id, name, category_id
                        FROM subcategories
                        WHERE category_id = ?
                        ORDER BY name ASC
                    """, (cat_row['id'],))
                    subcategories_rows = cursor.fetchall()
                    for sub_row in subcategories_rows:
                        category_dict['subcategories'].append(dict(sub_row))
                    categories_list.append(category_dict)
            return categories_list
        except sqlite3.Error as e:
            print(f"Database error retrieving categories with subcategories: {e}")
            return [] # Return empty list on error

    def get_all_categories_export(self):
        """
        Retrieves all categories (id, name) for data export.
        Orders by name.
        """
        items = []
        query = "SELECT id, name FROM categories ORDER BY name ASC"
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(query)
                rows = cursor.fetchall()
                for row in rows:
                    items.append(dict(row))
        except sqlite3.Error as e:
            print(f"Database error fetching all categories for export: {e}")
        return items

    def get_all_subcategories_export(self):
        """
        Retrieves all subcategories for data export.
        Includes category_name for context.
        Orders by category_name and then by subcategory name.
        """
        items = []
        query = """
            SELECT
                s.id,
                s.name,
                s.category_id,
                c.name AS category_name
            FROM subcategories s
            JOIN categories c ON s.category_id = c.id
            ORDER BY category_name ASC, s.name ASC
        """
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(query)
                rows = cursor.fetchall()
                for row in rows:
                    items.append(dict(row))
        except sqlite3.Error as e:
            print(f"Database error fetching all subcategories for export: {e}")
        return items

    def upload_products_excel(self, file_stream, overwrite_logic_choice):
        """
        Uploads products from an Excel file stream.
        - file_stream: The stream of the uploaded Excel file.
        - overwrite_logic_choice: "skip" or "overwrite" for handling duplicates.
        Returns a dictionary with counts of added, updated, skipped products, and errors.
        """
        import openpyxl # Ensure openpyxl is imported

        results = {"added": 0, "updated": 0, "skipped": 0, "errors": []}

        try:
            workbook = openpyxl.load_workbook(file_stream)
            sheet = workbook.active
        except Exception as e:
            results["errors"].append(f"Error reading Excel file: {str(e)}")
            return results

        header_row_values = [cell.value for cell in sheet[1]]
        header_map = {str(h).strip().lower(): idx for idx, h in enumerate(header_row_values) if h}

        required_headers = ['name', 'category_name', 'unit_of_measure', 'default_expiry_days']
        missing_headers = [req_h for req_h in required_headers if req_h not in header_map]
        if missing_headers:
            results["errors"].append(f"Missing required columns in Excel: {', '.join(missing_headers)}. Please check headers.")
            return results

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_error_prefix = f"Row {row_idx}: "

            # Helper to get cell value by header name, returns None if header missing or cell empty
            def get_cell_val(header_key, default=None):
                col_idx = header_map.get(header_key.lower())
                if col_idx is not None and len(row) > col_idx and row[col_idx] is not None:
                    return str(row[col_idx]).strip()
                return default

            name = get_cell_val('name')
            category_name_str = get_cell_val('category_name')
            subcategory_name_str = get_cell_val('subcategory_name') # Optional
            unit_of_measure = get_cell_val('unit_of_measure')
            default_expiry_days_str = get_cell_val('default_expiry_days')
            par_level_str = get_cell_val('par_level', "0") # Default to "0"
            max_holding_amount_str = get_cell_val('max_holding_amount', "0") # Default to "0"
            purchase_location = get_cell_val('purchase_location') # Optional

            # --- Validation ---
            if not name:
                results["errors"].append(f"{row_error_prefix}Product name is required.")
                continue
            if not category_name_str:
                results["errors"].append(f"{row_error_prefix}Category name is required for product '{name}'.")
                continue
            if not unit_of_measure:
                results["errors"].append(f"{row_error_prefix}Unit of measure is required for product '{name}'.")
                continue
            if not default_expiry_days_str:
                results["errors"].append(f"{row_error_prefix}Default expiry days are required for product '{name}'.")
                continue

            default_expiry_days = None
            try:
                default_expiry_days = int(default_expiry_days_str)
                if default_expiry_days < 0:
                    results["errors"].append(f"{row_error_prefix}Default expiry days for '{name}' must be non-negative.")
                    continue
            except ValueError:
                results["errors"].append(f"{row_error_prefix}Default expiry days for '{name}' must be a whole number.")
                continue

            par_level = 0.0
            try:
                par_level = float(par_level_str) if par_level_str else 0.0
                if par_level < 0:
                    results["errors"].append(f"{row_error_prefix}Par level for '{name}' must be non-negative.")
                    continue
            except ValueError:
                results["errors"].append(f"{row_error_prefix}Par level for '{name}' must be a valid number.")
                continue

            max_holding_amount = 0.0
            try:
                max_holding_amount = float(max_holding_amount_str) if max_holding_amount_str else 0.0
                if max_holding_amount < 0:
                    results["errors"].append(f"{row_error_prefix}Max holding amount for '{name}' must be non-negative.")
                    continue
            except ValueError:
                results["errors"].append(f"{row_error_prefix}Max holding amount for '{name}' must be a valid number.")
                continue

            # --- Category/Subcategory Resolution ---
            category_obj = self.get_category_by_name(category_name_str)
            if not category_obj:
                results["errors"].append(f"{row_error_prefix}Category '{category_name_str}' for product '{name}' not found. Please create it first.")
                continue
            category_id = category_obj['id']

            subcategory_id = None
            if subcategory_name_str:
                subcategory_obj = self.get_subcategory_by_name_and_category_id(subcategory_name_str, category_id)
                if not subcategory_obj:
                    results["errors"].append(f"{row_error_prefix}Subcategory '{subcategory_name_str}' under category '{category_name_str}' for product '{name}' not found. Please create it first.")
                    continue
                subcategory_id = subcategory_obj['id']

            # --- Product Existence Check & Action ---
            existing_product = self.get_product_by_name(name)

            product_data_dict = {
                "name": name,
                "category_id": category_id,
                "subcategory_id": subcategory_id,
                "unit_of_measure": unit_of_measure,
                "default_expiry_days": default_expiry_days,
                "par_level": par_level,
                "max_holding_amount": max_holding_amount,
                "purchase_location": purchase_location if purchase_location else None
            }

            if existing_product:
                if overwrite_logic_choice == "overwrite":
                    update_result = self.update_product(product_id=existing_product['id'], **product_data_dict)
                    if update_result.get("success"):
                        results["updated"] += 1
                    else:
                        results["errors"].append(f"{row_error_prefix}Error updating product '{name}': {update_result.get('message')}")
                else: # skip
                    results["skipped"] += 1
            else: # Product does not exist, create new
                create_result = self.create_product(**product_data_dict)
                if create_result.get("success"):
                    results["added"] += 1
                else:
                    results["errors"].append(f"{row_error_prefix}Error creating product '{name}': {create_result.get('message')}")

        return results

    def upload_historical_inventory_excel(self, file_stream):
        """
        Uploads historical inventory consumption data from an Excel file stream.
        - file_stream: The stream of the uploaded Excel file.
        Returns a dictionary with counts of added records and errors.
        """
        import openpyxl # Ensure openpyxl is imported
        from datetime import datetime # For date parsing and validation

        results = {"added": 0, "errors": []}

        try:
            workbook = openpyxl.load_workbook(file_stream)
            sheet = workbook.active
        except Exception as e:
            results["errors"].append(f"Error reading Excel file: {str(e)}")
            return results

        header_row_values = [cell.value for cell in sheet[1]]
        # Normalize headers: convert to string, strip whitespace, lowercase
        header_map = {str(h).strip().lower(): idx for idx, h in enumerate(header_row_values) if h}

        required_headers = ['product_name', 'quantity_consumed_this_time', 'consumed_date']
        missing_headers = [req_h for req_h in required_headers if req_h.lower() not in header_map]

        if missing_headers:
            results["errors"].append(f"Missing required columns in Excel: {', '.join(missing_headers)}. Please check headers (product_name, quantity_consumed_this_time, consumed_date).")
            return results

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_error_prefix = f"Row {row_idx}: "

            def get_cell_val(header_key, default=None):
                col_idx = header_map.get(header_key.lower()) # Use lowercased key
                if col_idx is not None and len(row) > col_idx and row[col_idx] is not None:
                    # Convert to string and strip for consistency, handle potential non-string values from Excel
                    return str(row[col_idx]).strip()
                return default

            product_name_str = get_cell_val('product_name')
            quantity_consumed_str = get_cell_val('quantity_consumed_this_time')
            consumed_date_str = get_cell_val('consumed_date')
            original_quantity_str = get_cell_val('original_quantity_string') # Optional
            purchase_date_str = get_cell_val('purchase_date') # Optional
            expiry_date_str = get_cell_val('expiry_date') # Optional

            # --- Validation ---
            current_row_errors = []
            if not product_name_str:
                current_row_errors.append("Product name is required.")
            if not quantity_consumed_str:
                current_row_errors.append("Quantity consumed is required.")
            if not consumed_date_str:
                current_row_errors.append("Consumed date is required.")

            if current_row_errors: # Check after gathering all missing required fields
                results["errors"].extend([f"{row_error_prefix}{err}" for err in current_row_errors])
                continue # Skip to next row

            # Validate quantity_consumed_this_time
            quantity_consumed_numeric = None
            try:
                quantity_consumed_numeric = float(quantity_consumed_str)
                if quantity_consumed_numeric <= 0:
                    current_row_errors.append("Quantity consumed must be a positive number.")
            except ValueError:
                current_row_errors.append("Quantity consumed must be a valid number.")

            # Validate consumed_date
            consumed_date_obj = None
            try:
                consumed_date_obj = datetime.strptime(consumed_date_str, '%Y-%m-%d').date()
            except ValueError:
                current_row_errors.append("Consumed date must be in YYYY-MM-DD format.")

            # Validate optional purchase_date
            purchase_date_obj = None
            if purchase_date_str:
                try:
                    purchase_date_obj = datetime.strptime(purchase_date_str, '%Y-%m-%d').date()
                except ValueError:
                    current_row_errors.append("Purchase date, if provided, must be in YYYY-MM-DD format.")

            # Validate optional expiry_date
            expiry_date_obj = None
            if expiry_date_str:
                try:
                    expiry_date_obj = datetime.strptime(expiry_date_str, '%Y-%m-%d').date()
                except ValueError:
                    current_row_errors.append("Expiry date, if provided, must be in YYYY-MM-DD format.")

            if current_row_errors:
                results["errors"].extend([f"{row_error_prefix}{err}" for err in current_row_errors])
                continue

            # --- Product Resolution ---
            product_info = self.get_product_by_name(product_name_str)
            if not product_info:
                results["errors"].append(f"{row_error_prefix}Product '{product_name_str}' not found in the database. Please add it first or check for typos.")
                continue

            product_id_to_use = product_info['id']
            # Use canonical product name from DB for consistency in historical_items
            product_name_canonical = product_info['name']

            # --- Database Insertion ---
            try:
                with self._get_db_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute('''
                        INSERT INTO historical_items
                        (product_id, name, quantity_consumed_this_time, consumed_date,
                         original_quantity_string, purchase_date, expiry_date)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (product_id_to_use, product_name_canonical, quantity_consumed_numeric,
                          consumed_date_obj.isoformat() if consumed_date_obj else None,
                          original_quantity_str if original_quantity_str else None,
                          purchase_date_obj.isoformat() if purchase_date_obj else None,
                          expiry_date_obj.isoformat() if expiry_date_obj else None))
                    conn.commit()
                    results["added"] += 1
            except sqlite3.Error as e:
                results["errors"].append(f"{row_error_prefix}Database error inserting record for '{product_name_canonical}': {e}")
            except Exception as e_general: # Catch any other unexpected error during DB operation
                results["errors"].append(f"{row_error_prefix}Unexpected error inserting record for '{product_name_canonical}': {e_general}")

        return results

    def upload_production_items_excel(self, file_stream):
        """
        Uploads production (garden) items from an Excel file stream.
        - file_stream: The stream of the uploaded Excel file.
        Returns a dictionary with counts of added records and errors.
        """
        import openpyxl
        from datetime import datetime

        results = {"added": 0, "errors": []}

        try:
            workbook = openpyxl.load_workbook(file_stream)
            sheet = workbook.active
        except Exception as e:
            results["errors"].append(f"Error reading Excel file: {str(e)}")
            return results

        header_row_values = [cell.value for cell in sheet[1]]
        header_map = {str(h).strip().lower(): idx for idx, h in enumerate(header_row_values) if h}

        required_headers = ['name', 'plant_date', 'time_to_harvest_days',
                            'expected_harvest_period_days', 'expected_yield_total']
        missing_headers = [req_h for req_h in required_headers if req_h.lower() not in header_map]

        if missing_headers:
            results["errors"].append(f"Missing required columns in Excel: {', '.join(missing_headers)}.")
            return results

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_error_prefix = f"Row {row_idx}: "

            def get_cell_val(header_key, default=None):
                col_idx = header_map.get(header_key.lower())
                if col_idx is not None and len(row) > col_idx and row[col_idx] is not None:
                    return str(row[col_idx]).strip()
                return default

            name_str = get_cell_val('name')
            associated_product_name_str = get_cell_val('associated_product_name') # Optional
            plant_date_str = get_cell_val('plant_date')
            time_to_harvest_days_str = get_cell_val('time_to_harvest_days')
            expected_harvest_period_days_str = get_cell_val('expected_harvest_period_days')
            expected_yield_total_str = get_cell_val('expected_yield_total')
            status_str = get_cell_val('status', 'Growing') # Default to 'Growing'

            current_row_errors = []
            if not name_str:
                current_row_errors.append("Name is required.")
            if not plant_date_str:
                current_row_errors.append("Plant date is required.")
            if not time_to_harvest_days_str:
                current_row_errors.append("Time to harvest days is required.")
            if not expected_harvest_period_days_str:
                current_row_errors.append("Expected harvest period days is required.")
            if not expected_yield_total_str:
                current_row_errors.append("Expected yield total is required.")

            if current_row_errors:
                results["errors"].extend([f"{row_error_prefix}{err}" for err in current_row_errors])
                continue

            # --- Further Validation ---
            plant_date_obj = None
            try:
                plant_date_obj = datetime.strptime(plant_date_str, '%Y-%m-%d').date()
            except ValueError:
                current_row_errors.append("Plant date must be in YYYY-MM-DD format.")

            time_to_harvest_days_int = None
            try:
                time_to_harvest_days_int = int(time_to_harvest_days_str)
                if time_to_harvest_days_int < 0:
                    current_row_errors.append("Time to harvest days must be non-negative.")
            except ValueError:
                current_row_errors.append("Time to harvest days must be a whole number.")

            expected_harvest_period_days_int = None
            try:
                expected_harvest_period_days_int = int(expected_harvest_period_days_str)
                if expected_harvest_period_days_int <= 0: # Typically should be positive
                    current_row_errors.append("Expected harvest period days must be positive.")
            except ValueError:
                current_row_errors.append("Expected harvest period days must be a whole number.")

            expected_yield_total_float = None
            try:
                expected_yield_total_float = float(expected_yield_total_str)
                if expected_yield_total_float < 0:
                     current_row_errors.append("Expected yield total must be non-negative.")
            except ValueError:
                current_row_errors.append("Expected yield total must be a valid number.")

            valid_statuses = ["Growing", "Harvesting", "Finished"]
            if status_str and status_str not in valid_statuses: # status_str defaults to 'Growing' if empty
                current_row_errors.append(f"Status must be one of {', '.join(valid_statuses)}.")

            if current_row_errors:
                results["errors"].extend([f"{row_error_prefix}{err}" for err in current_row_errors])
                continue

            # --- Associated Product Resolution ---
            associated_product_id = None
            if associated_product_name_str:
                product_info = self.get_product_by_name(associated_product_name_str)
                if product_info:
                    associated_product_id = product_info['id']
                else:
                    # Log a warning, but proceed with associated_product_id = None
                    warning_msg = f"{row_error_prefix}Associated product '{associated_product_name_str}' not found. Production item '{name_str}' will be added without product association."
                    if "warnings" not in results: results["warnings"] = [] # Ensure warnings list exists
                    results["warnings"].append(warning_msg)
                    # Not adding to results["errors"] as per instruction "treat associated_product_id as None"

            # --- Call add_production_item ---
            # Ensure plant_date_obj is used for the string representation if valid
            final_plant_date_str = plant_date_obj.isoformat() if plant_date_obj else plant_date_str

            add_result = self.add_production_item(
                name=name_str,
                associated_product_id=associated_product_id,
                plant_date_str=final_plant_date_str,
                time_to_harvest_days=time_to_harvest_days_int,
                expected_harvest_period_days=expected_harvest_period_days_int,
                expected_yield_total=expected_yield_total_float,
                status=status_str if status_str else "Growing" # Ensure default if somehow became empty
            )

            if add_result.get("success"):
                results["added"] += 1
            else:
                results["errors"].append(f"{row_error_prefix}Failed to add production item '{name_str}': {add_result.get('message', 'Unknown error')}")

        return results

    def get_category_by_name(self, name):
        """Retrieves a category by its name (case-insensitive)."""
        if not name or not isinstance(name, str) or not name.strip():
            return None
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, name FROM categories WHERE LOWER(name) = LOWER(?)", (name.strip(),))
                row = cursor.fetchone()
                return dict(row) if row else None
        except sqlite3.Error as e:
            print(f"Database error getting category by name '{name}': {e}")
            return None

    def get_subcategory_by_name_and_category_id(self, subcategory_name, category_id):
        """Retrieves a subcategory by its name and parent category_id (case-insensitive name)."""
        if not subcategory_name or not isinstance(subcategory_name, str) or not subcategory_name.strip() or not isinstance(category_id, int):
            return None
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT id, name, category_id
                    FROM subcategories
                    WHERE LOWER(name) = LOWER(?) AND category_id = ?
                """, (subcategory_name.strip(), category_id))
                row = cursor.fetchone()
                return dict(row) if row else None
        except sqlite3.Error as e:
            print(f"Database error getting subcategory by name '{subcategory_name}' and category_id {category_id}: {e}")
            return None

    def get_category_name_by_id(self, category_id):
        """Retrieves a category name by its ID."""
        if not isinstance(category_id, int):
            return None
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT name FROM categories WHERE id = ?", (category_id,))
                row = cursor.fetchone()
                return row['name'] if row else None
        except sqlite3.Error as e:
            print(f"Database error getting category name by ID {category_id}: {e}")
            return None

    # --- Production Item (Garden & Harvest) Methods ---
    def add_production_item(self, name, associated_product_id, plant_date_str,
                            time_to_harvest_days, expected_harvest_period_days,
                            expected_yield_total, status='Growing'):
        """Adds a new production item to the production_items table."""
        if not all([name, plant_date_str, time_to_harvest_days is not None,
                    expected_harvest_period_days is not None, expected_yield_total is not None]):
            return {"success": False, "message": "Missing required production item fields."}

        try:
            # Validate plant_date_str format
            date.fromisoformat(plant_date_str)
        except ValueError:
            return {"success": False, "message": "Invalid plant_date format. Use YYYY-MM-DD."}

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT INTO production_items
                    (name, associated_product_id, plant_date, time_to_harvest_days,
                    expected_harvest_period_days, expected_yield_total, status)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (name, associated_product_id, plant_date_str, time_to_harvest_days,
                      expected_harvest_period_days, expected_yield_total, status))
                conn.commit()
                item_id = cursor.lastrowid
                return {"success": True, "message": f"Production item '{name}' added successfully.", "item_id": item_id}
        except sqlite3.Error as e:
            return {"success": False, "message": f"Database error adding production item: {e}"}

    def get_production_item(self, item_id):
        """Retrieves a production item by its ID."""
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM production_items WHERE id = ?", (item_id,))
                row = cursor.fetchone()
                if row:
                    item = dict(row)
                    # Dynamically calculate status and yield for single item view as well
                    # This ensures consistency with get_all_production_items
                    # However, the core requirement was for get_all_production_items,
                    # so this is an enhancement. If not desired, remove calculation here.
                    try:
                        plant_dt = date.fromisoformat(item['plant_date'])
                        harvest_start_date = plant_dt + timedelta(days=item['time_to_harvest_days'])
                        harvest_end_date = harvest_start_date + timedelta(days=item['expected_harvest_period_days'])
                        current_dt = date.today()

                        if current_dt < harvest_start_date:
                            item['calculated_status'] = 'Growing'
                        elif harvest_start_date <= current_dt <= harvest_end_date:
                            item['calculated_status'] = 'Harvesting'
                        else:
                            item['calculated_status'] = 'Finished'

                        if item['expected_harvest_period_days'] > 0:
                            item['estimated_daily_yield'] = item['expected_yield_total'] / item['expected_harvest_period_days']
                        else:
                            item['estimated_daily_yield'] = 0
                    except (ValueError, TypeError) as e:
                        # Handle cases where date conversion or calculation might fail
                        item['calculated_status'] = item.get('status', 'Error calculating status') # Fallback to stored status
                        item['estimated_daily_yield'] = 'Error'
                        print(f"Error calculating dynamic fields for production item {item_id}: {e}")
                    return item
                return None
        except sqlite3.Error as e:
            print(f"Database error getting production item by ID {item_id}: {e}")
            return None

    def get_all_production_items(self, filters=None, sort_by='plant_date', sort_order='ASC',
                                 page=1, per_page=10):
        """Retrieves production items with filtering, sorting, pagination, and calculated fields."""
        items = []
        params = []

        query = "SELECT * FROM production_items"

        # Basic filtering (e.g., by status)
        where_clauses = []
        if filters and 'status' in filters:
            # Note: This filters by the *stored* status. Dynamic status is calculated later.
            # If filtering by dynamic status is needed, the query or post-processing becomes more complex.
            where_clauses.append("status = ?")
            params.append(filters['status'])

        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)

        # Sorting
        valid_sort_columns = {
            'name': 'name', 'plant_date': 'plant_date', 'status': 'status',
            'expected_yield_total': 'expected_yield_total'
            # Add more as needed
        }
        sort_column = valid_sort_columns.get(sort_by, 'plant_date')
        sort_order_upper = 'DESC' if sort_order.upper() == 'DESC' else 'ASC'
        query += f" ORDER BY {sort_column} {sort_order_upper}"

        # Pagination
        # If page or per_page is None, or if per_page is not a positive integer,
        # do not apply LIMIT and OFFSET, effectively fetching all products.
        if page is not None and per_page is not None and \
           isinstance(page, int) and isinstance(per_page, int) and \
           page > 0 and per_page > 0:
            offset = (page - 1) * per_page
            query += " LIMIT ? OFFSET ?"
            params.extend([per_page, offset])
        # No 'else' or 'elif' here: if conditions are not met, pagination is skipped.

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(query, tuple(params))
                rows = cursor.fetchall()
                current_dt = date.today()
                for row_data in rows:
                    item = dict(row_data)
                    try:
                        plant_dt = date.fromisoformat(item['plant_date'])
                        # Ensure time_to_harvest_days and expected_harvest_period_days are not None
                        time_to_harvest = item['time_to_harvest_days'] if item['time_to_harvest_days'] is not None else 0
                        expected_period = item['expected_harvest_period_days'] if item['expected_harvest_period_days'] is not None else 0

                        harvest_start_date = plant_dt + timedelta(days=time_to_harvest)
                        harvest_end_date = harvest_start_date + timedelta(days=expected_period)

                        if current_dt < harvest_start_date:
                            item['calculated_status'] = 'Growing'
                        elif harvest_start_date <= current_dt <= harvest_end_date:
                            item['calculated_status'] = 'Harvesting'
                        else:
                            item['calculated_status'] = 'Finished'

                        if expected_period > 0 and item['expected_yield_total'] is not None:
                            item['estimated_daily_yield'] = item['expected_yield_total'] / expected_period
                        else:
                            item['estimated_daily_yield'] = 0

                    except (ValueError, TypeError, KeyError) as e:
                        item['calculated_status'] = item.get('status', 'Error') # Fallback to stored status or 'Error'
                        item['estimated_daily_yield'] = 'Error'
                        print(f"Error calculating dynamic fields for item ID {item.get('id')}: {e}")

                    items.append(item)
        except sqlite3.Error as e:
            print(f"Database error fetching all production items: {e}")
        return items

    def update_production_item(self, item_id, data):
        """Updates an existing production item."""
        if not data:
            return {"success": False, "message": "No data provided for update."}

        fields = []
        params = []
        for key, value in data.items():
            # Add more validation for column names if necessary
            if key in ['name', 'associated_product_id', 'plant_date',
                       'time_to_harvest_days', 'expected_harvest_period_days',
                       'expected_yield_total', 'status']:
                if key == 'plant_date':
                    try:
                        date.fromisoformat(value) # Validate date format
                    except ValueError:
                        return {"success": False, "message": f"Invalid plant_date format for {key}: {value}. Use YYYY-MM-DD."}
                fields.append(f"{key} = ?")
                params.append(value)

        if not fields:
            return {"success": False, "message": "No valid fields to update."}

        params.append(item_id)
        query = f"UPDATE production_items SET {', '.join(fields)} WHERE id = ?"

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(query, tuple(params))
                conn.commit()
                if cursor.rowcount == 0:
                    return {"success": False, "message": f"Production item with ID {item_id} not found."}
                return {"success": True, "message": f"Production item ID {item_id} updated successfully."}
        except sqlite3.Error as e:
            return {"success": False, "message": f"Database error updating production item: {e}"}

    def record_harvest(self, production_item_id, actual_harvest_amount, harvest_date_str):
        """Records a harvest, adding it to inventory stock."""
        if actual_harvest_amount <= 0:
            return {"success": False, "message": "Actual harvest amount must be positive."}

        production_item = self.get_production_item(production_item_id)
        if not production_item:
            return {"success": False, "message": f"Production item with ID {production_item_id} not found."}

        associated_product_id = production_item.get('associated_product_id')
        if associated_product_id is None:
            return {"success": False, "message": f"Production item ID {production_item_id} does not have an associated product ID."}

        # Validate harvest_date_str format
        try:
            date.fromisoformat(harvest_date_str)
        except ValueError:
            return {"success": False, "message": "Invalid harvest_date format. Use YYYY-MM-DD."}

        # Call add_inventory_stock to add the harvested amount to general inventory
        # self.add_inventory_stock handles fetching product details (like default_expiry_days)
        # and creating the inventory_item record.
        stock_result = self.add_inventory_stock(
            product_id=associated_product_id,
            quantity_str=str(actual_harvest_amount), # add_inventory_stock expects a string
            purchase_date_str=harvest_date_str # Harvest date is treated as purchase date for inventory purposes
        )

        # Optionally, update production_item status or remaining yield here if needed in future.
        # For now, status updates are via update_production_item or dynamic calculation.

        return stock_result

    def get_all_production_items_export(self):
        """
        Retrieves all production items for data export.
        Includes associated_product_name for context.
        Orders by plant_date and then by name.
        """
        items = []
        query = """
            SELECT
                pi.*,
                p.name AS associated_product_name
            FROM production_items pi
            LEFT JOIN products p ON pi.associated_product_id = p.id
            ORDER BY pi.plant_date ASC, pi.name ASC
        """
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(query)
                rows = cursor.fetchall()
                for row in rows:
                    items.append(dict(row))
        except sqlite3.Error as e:
            print(f"Database error fetching all production items for export: {e}")
        return items

    # --- Product Management Methods ---
    def create_product(self, name, category_id, subcategory_id, unit_of_measure, default_expiry_days,
                       par_level=0, max_holding_amount=0, purchase_location=None):
        """Inserts a new product into the products table."""
        # category_id is treated as required for now. subcategory_id is optional.
        if not all([name, unit_of_measure, default_expiry_days is not None, category_id is not None]):
            return {"success": False, "message": "Missing required product fields (name, category_id, unit_of_measure, default_expiry_days)."}
        if category_id is not None and not isinstance(category_id, int):
            return {"success": False, "message": "Category ID must be an integer."}
        if subcategory_id is not None and not isinstance(subcategory_id, int):
            return {"success": False, "message": "Subcategory ID must be an integer if provided."}

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT INTO products
                    (name, category_id, subcategory_id, unit_of_measure, default_expiry_days,
                     par_level, max_holding_amount, purchase_location)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (name, category_id, subcategory_id, unit_of_measure, default_expiry_days,
                      par_level, max_holding_amount, purchase_location))
                conn.commit()
                product_id = cursor.lastrowid
                return {"success": True, "message": f"Product '{name}' created successfully.", "product_id": product_id}
        except sqlite3.IntegrityError:
            return {"success": False, "message": f"Product name '{name}' already exists."}
        except sqlite3.Error as e:
            return {"success": False, "message": f"Database error creating product: {e}"}

    def get_product(self, product_id):
        """Retrieves a product by its ID, including category and subcategory names."""
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT
                        p.id, p.name, p.unit_of_measure, p.default_expiry_days,
                        p.par_level, p.max_holding_amount, p.purchase_location,
                        p.consumption_override_rate, p.category_id, p.subcategory_id,
                        c.name AS category_name,
                        s.name AS subcategory_name
                    FROM products p
                    LEFT JOIN categories c ON p.category_id = c.id
                    LEFT JOIN subcategories s ON p.subcategory_id = s.id
                    WHERE p.id = ?
                """, (product_id,))
                row = cursor.fetchone()
                return dict(row) if row else None
        except sqlite3.Error as e:
            print(f"Database error getting product by ID {product_id}: {e}")
            return None # Or raise error

    def get_product_details(self, product_id):
        """
        Retrieves a product by its ID, and enhances it with current on-hand inventory
        and the nearest expiry date from inventory_items.
        """
        product = self.get_product(product_id)
        if not product:
            return None

        # Calculate current_on_hand_inventory
        # self.get_total_item_quantity already handles product_id or name
        current_on_hand_inventory = self.get_total_item_quantity(product_id)
        product['current_on_hand_inventory'] = current_on_hand_inventory

        # Find nearest_expiry_date
        nearest_expiry_date_str = "N/A"
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT expiry_date
                    FROM inventory_items
                    WHERE product_id = ?
                    ORDER BY expiry_date ASC
                    LIMIT 1
                ''', (product_id,))
                row = cursor.fetchone()
                if row and row['expiry_date']:
                    nearest_expiry_date_str = row['expiry_date']
        except sqlite3.Error as e:
            print(f"Database error getting nearest expiry date for product ID {product_id}: {e}")
            # nearest_expiry_date_str remains "N/A" or could be set to an error indicator

        product['nearest_expiry_date'] = nearest_expiry_date_str
        return product

    def get_daily_consumption(self, product_id, days=30):
        """
        Retrieves daily consumption for a product over a specified number of days.
        - product_id: The ID of the product.
        - days: The number of past days to fetch data for.
        """
        consumption_data = []
        today = date.today()
        start_date = today - timedelta(days=days)
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT
                        strftime('%Y-%m-%d', consumed_date) as consumption_date,
                        SUM(quantity_consumed_this_time) as total_quantity_consumed
                    FROM historical_items
                    WHERE product_id = ? AND consumed_date >= ?
                    GROUP BY consumption_date
                    ORDER BY consumption_date ASC
                ''', (product_id, start_date.isoformat()))
                rows = cursor.fetchall()
                for row in rows:
                    consumption_data.append(dict(row))
        except sqlite3.Error as e:
            print(f"Database error getting daily consumption for product ID {product_id}: {e}")
            # Depending on requirements, could return empty list or raise error
        return consumption_data

    def get_monthly_consumption(self, product_id, months=12):
        """
        Retrieves monthly consumption for a product over a specified number of months.
        - product_id: The ID of the product.
        - months: The number of past months to fetch data for.
        """
        consumption_data = []
        # Calculate the first day of the month 'months' ago
        # This is a bit tricky to get precisely the Nth month back.
        # A simpler approach for SQL is to filter for dates >= the first day of that month.
        today = date.today()
        # Approximate start_date: go back (months * average_days_in_month)
        # A more robust way is to iterate back month by month or use date library functions
        # For SQL, we can use date functions if available and portable,
        # or filter for a wider range and then process in Python if needed.
        # Simplest for SQLite: calculate the date string for "N months ago"
        # and group by YYYY-MM.

        # Calculate the first day of the current month
        first_day_current_month = today.replace(day=1)

        # Iterate backwards to find the first day of the month N months ago
        target_month_date = first_day_current_month
        for _ in range(months -1): # Subtract 1 because current month counts as one of the 'months'
            # Move to the last day of the previous month
            target_month_date = (target_month_date - timedelta(days=1)).replace(day=1)

        start_month_iso = target_month_date.isoformat()

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT
                        strftime('%Y-%m', consumed_date) as consumption_month,
                        SUM(quantity_consumed_this_time) as total_quantity_consumed
                    FROM historical_items
                    WHERE product_id = ? AND date(consumed_date) >= date(?)
                    GROUP BY consumption_month
                    ORDER BY consumption_month ASC
                ''', (product_id, start_month_iso))
                rows = cursor.fetchall()
                for row in rows:
                    consumption_data.append(dict(row))
        except sqlite3.Error as e:
            print(f"Database error getting monthly consumption for product ID {product_id}: {e}")
        return consumption_data

    def get_daily_inventory_history(self, product_id, days=30):
        """
        Retrieves the daily inventory quantity for a product over a specified number of past days.
        - product_id: The ID of the product.
        - days: The number of past days to fetch data for (defaults to 30).
        Returns a list of dictionaries: [{'inventory_date': 'YYYY-MM-DD', 'quantity_on_hand': float}]
        """
        if not isinstance(product_id, int):
            raise ValueError("product_id must be an integer.")
        if not isinstance(days, int) or days <= 0:
            raise ValueError("days must be a positive integer.")

        today = date.today()
        report_start_date = today - timedelta(days=days - 1)
        events = []

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()

                # Fetch purchase events
                cursor.execute('''
                    SELECT purchase_date, original_quantity_string
                    FROM inventory_items
                    WHERE product_id = ?
                ''', (product_id,))
                for row in cursor.fetchall():
                    try:
                        event_date = date.fromisoformat(row['purchase_date'])
                        quantity = self._parse_quantity_string(row['original_quantity_string'])
                        if quantity > 0:
                            events.append({'date': event_date, 'change': quantity, 'type': 'purchase'})
                        else:
                            print(f"Warning: Skipping purchase event with non-positive quantity for product_id {product_id}: {row}")
                    except (ValueError, TypeError) as e:
                        print(f"Warning: Skipping purchase event due to invalid data for product_id {product_id}: {row} - Error: {e}")

                # Fetch consumption events
                cursor.execute('''
                    SELECT consumed_date, quantity_consumed_this_time
                    FROM historical_items
                    WHERE product_id = ?
                ''', (product_id,))
                for row in cursor.fetchall():
                    try:
                        event_date = date.fromisoformat(row['consumed_date'])
                        quantity = float(row['quantity_consumed_this_time'])
                        if quantity > 0:
                            events.append({'date': event_date, 'change': -quantity, 'type': 'consumption'})
                        else:
                            print(f"Warning: Skipping consumption event with non-positive quantity for product_id {product_id}: {row}")
                    except (ValueError, TypeError) as e:
                        print(f"Warning: Skipping consumption event due to invalid data for product_id {product_id}: {row} - Error: {e}")

        except sqlite3.Error as e:
            print(f"Database error retrieving inventory history for product ID {product_id}: {e}")
            return []

        # Sort events: by date, then by type (purchases first on same day)
        events.sort(key=lambda x: (x['date'], 0 if x['type'] == 'purchase' else 1))

        daily_inventory_data = []
        current_simulated_stock = 0.0
        event_idx = 0

        # Pre-window stock calculation: Accumulate stock changes before the report_start_date
        while event_idx < len(events) and events[event_idx]['date'] < report_start_date:
            current_simulated_stock += events[event_idx]['change']
            event_idx += 1

        # Simulate and record for the report window
        for i in range(days):
            current_report_day = report_start_date + timedelta(days=i)

            # Process events for the current_report_day
            while event_idx < len(events) and events[event_idx]['date'] == current_report_day:
                current_simulated_stock += events[event_idx]['change']
                event_idx += 1

            daily_inventory_data.append({
                'inventory_date': current_report_day.isoformat(),
                'quantity_on_hand': max(0, round(current_simulated_stock, 3))
            })

        return daily_inventory_data

    def get_past_actual_inventory_summary(self, product_id, days_past):
        """
        Provides a day-by-day summary of actual past inventory levels and consumption
        for a given product over a specified number of past days.
        """
        product = self.get_product(product_id)
        if not product:
            return {"success": False, "message": f"Product with ID {product_id} not found."}

        if not isinstance(days_past, int) or days_past <= 0:
            # Or raise ValueError, depending on desired error handling
            return {"success": False, "message": "days_past must be a positive integer."}

        # Fetch raw data
        # These methods already return data sorted by date ascending
        inventory_history_raw = self.get_daily_inventory_history(product_id, days=days_past)
        consumption_history_raw = self.get_daily_consumption(product_id, days=days_past)

        # Convert raw data to dictionaries for easier lookup by date
        inventory_map = {item['inventory_date']: item['quantity_on_hand'] for item in inventory_history_raw}
        consumption_map = {item['consumption_date']: item['total_quantity_consumed'] for item in consumption_history_raw}

        past_summary_results = []
        today = date.today()

        # Iterate from `days_past - 1` down to `0` to cover the period from oldest to newest
        # The dates generated by get_daily_inventory_history are the actual dates we care about.
        # We can iterate through a date range from (today - days_past) to (today - 1 day)

        # The get_daily_inventory_history method returns a list of dicts for 'days' in the past,
        # starting from (today - days + 1) up to today.
        # Let's use the dates from inventory_history_raw as the canonical list of dates to report on.
        # If inventory_history_raw is empty (e.g. new product), we need a date range.

        start_date_of_summary = today - timedelta(days=days_past) # Inclusive start

        # If no inventory history, generate for all days in range with 0s
        if not inventory_history_raw:
            for i in range(days_past):
                target_date = start_date_of_summary + timedelta(days=i)
                target_date_str = target_date.isoformat()
                past_summary_results.append({
                    'date': target_date_str,
                    'actual_ending_inventory': 0.0,
                    'actual_consumption': consumption_map.get(target_date_str, 0.0),
                    'actual_shrink': 0.0,  # Placeholder
                    'actual_harvest': 0.0, # Placeholder
                })
            return past_summary_results # Already sorted by date


        # If there is inventory history, use its dates primarily
        # This loop ensures we cover the days as per get_daily_inventory_history,
        # which is 'days' number of records ending yesterday (if 'today' is not included in its range).
        # get_daily_inventory_history returns `days` number of records.
        # Example: if days=7, it returns 7 records, from (today-6) to today.
        # The problem statement implies "past N days", so if today is D, and days_past=7,
        # we want D-7, D-6, ..., D-1.
        # get_daily_inventory_history for days=7 would give:
        # (today-7+1) = today-6, ..., today.
        # This is slightly off from "N past days" if it includes today.
        # Let's adjust the loop to be explicit about the N past days.

        # Iterate for each of the 'days_past' looking backwards from yesterday.
        # days_past = 1 means yesterday. days_past = 7 means from 7 days ago up to yesterday.
        for i in range(days_past):
            # target_date goes from (today - days_past) up to (today - 1)
            target_date = (today - timedelta(days=days_past)) + timedelta(days=i)
            target_date_str = target_date.isoformat()

            quantity_on_hand = inventory_map.get(target_date_str, 0.0)
            # If inventory_map doesn't have the date, it means stock was 0 or no event.
            # get_daily_inventory_history should ideally fill dates with 0 if no specific event.
            # Based on current get_daily_inventory_history, it *does* create entries for each day in the window.

            consumption_for_day = consumption_map.get(target_date_str, 0.0)

            past_summary_results.append({
                'date': target_date_str,
                'actual_ending_inventory': round(quantity_on_hand, 2),
                'actual_consumption': round(consumption_for_day, 2),
                'actual_shrink': 0.0,  # Placeholder
                'actual_harvest': 0.0, # Placeholder
            })

        # The list is already sorted by date ascending due to the loop's construction.
        return past_summary_results

    def get_product_by_name(self, name):
        """Retrieves a product by its name."""
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM products WHERE LOWER(name) = LOWER(?)", (name,))
                row = cursor.fetchone()
                return dict(row) if row else None
        except sqlite3.Error as e:
            print(f"Database error getting product by name '{name}': {e}")
            return None

    def get_all_products(self, search_term=None, category=None, purchase_location=None,
                         sort_by='name', sort_order='ASC', page=None, per_page=None):
        """
        Retrieves products from the products table with filtering, sorting, and pagination.
        - search_term: Filters by product name (case-insensitive).
        - category: Filters by category name (case-insensitive).
        - purchase_location: Filters by purchase location (case-insensitive).
        - sort_by: Column to sort by ('name', 'category', 'subcategory', 'purchase_location').
                   Defaults to 'name'. Invalid values also default to 'name'.
        - sort_order: 'ASC' or 'DESC'. Defaults to 'ASC'.
        - page: For pagination, the page number to retrieve. Defaults to 1.
        - per_page: For pagination, items per page. Defaults to 10.
        """
        items = []
        params = []

        # Base query
        query = """
            SELECT
                p.id, p.name, p.unit_of_measure, p.default_expiry_days,
                p.par_level, p.max_holding_amount, p.purchase_location,
                p.consumption_override_rate, p.category_id, p.subcategory_id,
                c.name AS category_name,
                s.name AS subcategory_name
            FROM products p
            LEFT JOIN categories c ON p.category_id = c.id
            LEFT JOIN subcategories s ON p.subcategory_id = s.id
        """

        # Filtering
        where_clauses = []
        if search_term:
            where_clauses.append("LOWER(p.name) LIKE ?")
            params.append(f"%{search_term.lower()}%")
        if category: # This now filters by category NAME from the joined 'categories' table
            where_clauses.append("LOWER(c.name) = ?")
            params.append(category.lower())
        if purchase_location:
            where_clauses.append("LOWER(p.purchase_location) = ?")
            params.append(purchase_location.lower())

        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)

        # Sorting
        valid_sort_columns = {
            'name': 'p.name',
            'category': 'c.name',
            'subcategory': 's.name',
            'purchase_location': 'p.purchase_location'
            # Add other p.columns if needed for sorting, e.g. p.default_expiry_days
        }
        # Default to 'p.name' if sort_by is invalid or not in valid_sort_columns
        sort_column = valid_sort_columns.get(sort_by.lower(), 'p.name')

        sort_order_upper = sort_order.upper()
        if sort_order_upper not in ['ASC', 'DESC']:
            sort_order_upper = 'ASC' # Default to 'ASC'

        query += f" ORDER BY {sort_column} {sort_order_upper}, p.id {sort_order_upper}" # Added p.id for stable sort

        # Pagination
        # If page or per_page is None, or if per_page is not a positive integer,
        # do not apply LIMIT and OFFSET, effectively fetching all products.
        if page is not None and per_page is not None and \
           isinstance(page, int) and isinstance(per_page, int) and \
           page > 0 and per_page > 0:
            offset = (page - 1) * per_page
            query += " LIMIT ? OFFSET ?"
            params.extend([per_page, offset])
        # No 'else' or 'elif' here: if conditions are not met, pagination is skipped.

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(query, tuple(params))
                rows = cursor.fetchall()
                for row in rows:
                    items.append(dict(row))
        except sqlite3.Error as e:
            print(f"Database error fetching all products with filters: {e}")
        return items

    def get_all_products_export(self):
        """
        Retrieves all products from the products table for data export.
        Includes category and subcategory names. No pagination.
        Selects specific fields relevant for export.
        """
        items = []
        query = """
            SELECT
                p.id,
                p.name,
                p.unit_of_measure,
                p.default_expiry_days,
                p.par_level,
                p.max_holding_amount,
                p.purchase_location,
                c.name AS category_name,
                s.name AS subcategory_name
            FROM products p
            LEFT JOIN categories c ON p.category_id = c.id
            LEFT JOIN subcategories s ON p.subcategory_id = s.id
            ORDER BY p.name ASC
        """
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(query)
                rows = cursor.fetchall()
                for row in rows:
                    items.append(dict(row))
        except sqlite3.Error as e:
            print(f"Database error fetching all products for export: {e}")
        return items

    def get_all_inventory_batches_export(self, start_date_str=None, end_date_str=None):
        """
        Retrieves all inventory batches for data export, optionally filtered by purchase date.
        Includes product_name for context.
        Orders by purchase_date and then by id.
        """
        items = []
        params = []

        query = """
            SELECT
                ii.*,
                p.name AS product_name
            FROM inventory_items ii
            JOIN products p ON ii.product_id = p.id
        """

        where_clauses = []
        if start_date_str and start_date_str.strip():
            try:
                date.fromisoformat(start_date_str.strip()) # Validate date format
                where_clauses.append("ii.purchase_date >= ?")
                params.append(start_date_str.strip())
            except ValueError:
                print(f"Warning: Invalid start_date_str format '{start_date_str}'. It will be ignored.")
                # Optionally, raise an error or return an empty list with an error message
                # For now, just ignoring the invalid date string for the query

        if end_date_str and end_date_str.strip():
            try:
                date.fromisoformat(end_date_str.strip()) # Validate date format
                where_clauses.append("ii.purchase_date <= ?")
                params.append(end_date_str.strip())
            except ValueError:
                print(f"Warning: Invalid end_date_str format '{end_date_str}'. It will be ignored.")

        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)

        query += " ORDER BY ii.purchase_date ASC, ii.id ASC"

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(query, tuple(params))
                rows = cursor.fetchall()
                for row in rows:
                    items.append(dict(row))
        except sqlite3.Error as e:
            print(f"Database error fetching all inventory batches for export: {e}")
            # Consider raising the error or returning a specific error indicator
        return items

    def get_product_count(self, search_term=None, category=None, purchase_location=None):
        """
        Gets the total number of products, filtered by search_term, category, and purchase_location.
        """
        params = []
        query = """
            SELECT COUNT(p.id) as count
            FROM products p
            LEFT JOIN categories c ON p.category_id = c.id
        """ # s (subcategories) is not needed for count if not filtering by subcategory name or id

        where_clauses = []
        if search_term:
            where_clauses.append("LOWER(p.name) LIKE ?") # Alias product table as p
            params.append(f"%{search_term.lower()}%")
        if category: # This now filters by category NAME from the joined 'categories' table
            where_clauses.append("LOWER(c.name) = ?") # Alias category table as c
            params.append(category.lower())
        if purchase_location:
            where_clauses.append("LOWER(p.purchase_location) = ?") # Alias product table as p
            params.append(purchase_location.lower())

        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(query, tuple(params))
                result = cursor.fetchone()
                return result['count'] if result else 0
        except sqlite3.Error as e:
            print(f"Database error getting product count: {e}")
            return 0

    def get_all_categories(self):
        """Retrieves all category names from the 'categories' table."""
        # This method now fetches from the dedicated 'categories' table.
        category_names = []
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT name FROM categories ORDER BY name ASC")
                rows = cursor.fetchall()
                for row in rows:
                    category_names.append(row['name'])
        except sqlite3.Error as e:
            print(f"Database error fetching all category names: {e}")
        return category_names

    def get_all_purchase_locations(self):
        """Retrieves all unique purchase location names from the products table."""
        locations = []
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT DISTINCT purchase_location FROM products WHERE purchase_location IS NOT NULL ORDER BY purchase_location ASC")
                rows = cursor.fetchall()
                for row in rows:
                    locations.append(row['purchase_location'])
        except sqlite3.Error as e:
            print(f"Database error fetching all purchase locations: {e}")
        return locations

    def get_products_for_projection_list(self, search_term=None, category=None, purchase_location=None,
                                         sort_by='name', sort_order='ASC', page=1, per_page=10):
        """
        Retrieves a paginated and filtered list of products for the demand projection page.
        This is similar to get_all_products but specifically for listing products
        for which projections will be calculated.
        """
        # This method is essentially a wrapper or identical to get_all_products.
        # Re-using get_all_products logic directly.
        return self.get_all_products(
            search_term=search_term,
            category=category,
            purchase_location=purchase_location,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            per_page=per_page
        )

    def get_products_for_projection_list_count(self, search_term=None, category=None, purchase_location=None):
        """
        Gets the total number of products for the projection list, filtered by product attributes.
        This is similar to get_product_count.
        """
        # This method is essentially a wrapper or identical to get_product_count.
        # Re-using get_product_count logic directly.
        return self.get_product_count(
            search_term=search_term,
            category=category,
            purchase_location=purchase_location
        )

    def update_product(self, product_id, name, category_id, subcategory_id, unit_of_measure,
                       default_expiry_days, par_level, max_holding_amount, purchase_location):
        """Updates an existing product in the products table."""
        # category_id is treated as required for now. subcategory_id is optional.
        if not all([name, unit_of_measure, default_expiry_days is not None, category_id is not None]):
            return {"success": False, "message": "Missing required product fields for update (name, category_id, unit_of_measure, default_expiry_days)."}
        if category_id is not None and not isinstance(category_id, int):
            return {"success": False, "message": "Category ID must be an integer."}
        if subcategory_id is not None and not isinstance(subcategory_id, int): # subcategory_id can be None
            return {"success": False, "message": "Subcategory ID must be an integer if provided."}

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    UPDATE products SET
                    name = ?, category_id = ?, subcategory_id = ?, unit_of_measure = ?,
                    default_expiry_days = ?, par_level = ?, max_holding_amount = ?, purchase_location = ?
                    WHERE id = ?
                ''', (name, category_id, subcategory_id, unit_of_measure, default_expiry_days,
                      par_level, max_holding_amount, purchase_location, product_id))
                conn.commit()
                if cursor.rowcount == 0:
                    return {"success": False, "message": f"Product with ID {product_id} not found."}
                return {"success": True, "message": f"Product ID {product_id} updated successfully."}
        except sqlite3.IntegrityError:
            return {"success": False, "message": f"Product name '{name}' may already exist for another product."}
        except sqlite3.Error as e:
            return {"success": False, "message": f"Database error updating product: {e}"}

    def save_consumption_overrides(self, product_overrides: list):
        """
        Saves consumption override rates for products.
        - product_overrides: A list of dictionaries, where each dictionary is
                             {'product_id': int, 'override_rate': float_or_none_or_empty_str}.
        """
        if not isinstance(product_overrides, list):
            return {"success": False, "message": "Invalid input: product_overrides must be a list."}

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                updates_made = 0
                errors_encountered = 0

                for override_item in product_overrides:
                    if not isinstance(override_item, dict) or \
                       'product_id' not in override_item or \
                       'override_rate' not in override_item:
                        print(f"Skipping invalid override item: {override_item}")
                        errors_encountered += 1
                        continue

                    product_id = override_item['product_id']
                    override_rate_input = override_item['override_rate']

                    final_override_rate = None # Default to NULL

                    if isinstance(override_rate_input, str):
                        if override_rate_input.strip() == "":
                            final_override_rate = None
                        else:
                            try:
                                final_override_rate = float(override_rate_input)
                            except ValueError:
                                print(f"Warning: Could not parse override_rate '{override_rate_input}' for product ID {product_id}. Setting to NULL.")
                                final_override_rate = None # Explicitly NULL if unparsable
                    elif isinstance(override_rate_input, (int, float)):
                        final_override_rate = float(override_rate_input)
                    # If override_rate_input is None, final_override_rate remains None (SQL NULL)

                    try:
                        cursor.execute('''
                            UPDATE products
                            SET consumption_override_rate = ?
                            WHERE id = ?
                        ''', (final_override_rate, product_id))
                        if cursor.rowcount > 0:
                            updates_made += 1
                        # No error if product_id not found, rowcount will be 0
                    except sqlite3.Error as e_item:
                        print(f"Error updating override for product ID {product_id}: {e_item}")
                        errors_encountered += 1

                conn.commit()

                message = f"Consumption overrides saved. {updates_made} products updated."
                if errors_encountered > 0:
                    message += f" {errors_encountered} items had errors."

                return {"success": errors_encountered == 0, "message": message, "updates_made": updates_made, "errors": errors_encountered}

        except sqlite3.Error as e:
            return {"success": False, "message": f"Database error saving consumption overrides: {e}"}

    def _parse_quantity_string(self, quantity_str):
        """Attempts to parse a quantity string (e.g., "2 lbs", "10 units", "1") into a float.
           Returns float if parsable, otherwise None or a specific handling.
           For simplicity, we'll try to extract the first numeric part.
           - If the string is "a loaf", "one piece", etc., it's treated as 1.0.
           - If the string is "2 lbs", "0.5 kg", it extracts 2.0 or 0.5.
           - If the string is just a number "10", it returns 10.0.
           - If unparseable or empty in a numeric context, returns 0.0.
        """
        if isinstance(quantity_str, (int, float)):
            return float(quantity_str)
        
        s_quantity_str = str(quantity_str).strip()
        if not s_quantity_str:
            return 0.0

        parts = s_quantity_str.split()
        try:
            # Try to convert the first part to a float
            return float(parts[0])
        except (ValueError, IndexError):
            # If the first part is not a number (e.g., "a loaf", "one unit")
            # and the string is not empty, consider it as 1.0.
            # This helps in contexts where such string quantities imply a single unit.
            if s_quantity_str: # Check if the original stripped string was non-empty
                return 1.0 
            return 0.0 # Default for completely unparseable or empty strings

    def add_inventory_stock(self, product_id, quantity_str, purchase_date_str):
        """Adds a new inventory item stock linked to a product."""
        product = self.get_product(product_id)
        if not product:
            return {"success": False, "message": f"Product with ID {product_id} not found."}

        try:
            purchase_dt = date.fromisoformat(purchase_date_str)
            expiry_dt = purchase_dt + timedelta(days=int(product['default_expiry_days']))
        except (ValueError, TypeError) as e:
            return {"success": False, "message": f"Invalid date or expiry day format for product {product['name']}: {e}"}

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT INTO inventory_items 
                    (product_id, name, quantity, purchase_date, expiry_date, original_quantity_string)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (product_id, product['name'], str(quantity_str), purchase_dt.isoformat(),
                      expiry_dt.isoformat(), str(quantity_str)))
                conn.commit()
                stock_item_id = cursor.lastrowid # Get the ID of the inserted item
            print(f"Added stock to DB: {product['name']} ({quantity_str}), Expires: {expiry_dt.isoformat()}")
            return {"success": True, "message": f"Stock for '{product['name']}' added successfully.", "stock_item_id": stock_item_id}
        except sqlite3.Error as e:
            print(f"Database error adding inventory stock: {e}")
            return {"success": False, "message": f"Database error: {e}"}

    # This method is called by the Excel upload in app.py
    # It now relies on product_id for linking, and creates product if not found.
    def add_item_to_list(self, name, quantity_str, purchase_date_str, expiry_days,
                         category=None, subcategory=None, par_level=0, max_holding_amount=0,
                         purchase_location=None, unit_of_measure=None,
                         cost_per_unit_str=None, vendor=None, # New parameters for costing
                         # The following are for handling pending actions from app.py
                         confirmed_action=None, temp_category_id=None):
        """
        Adds an item to inventory. Handles product creation, category/subcategory resolution,
        and potential user confirmation steps for new categories/subcategories.
        If cost_per_unit is provided, logs the purchase in PurchaseLog.
        """
        print(f"LOG: add_item_to_list called with: name='{name}', quantity_str='{quantity_str}', purchase_date_str='{purchase_date_str}', expiry_days={expiry_days}, category='{category}', subcategory='{subcategory}', par_level={par_level}, max_holding_amount={max_holding_amount}, purchase_location='{purchase_location}', unit_of_measure='{unit_of_measure}', cost_per_unit_str='{cost_per_unit_str}', vendor='{vendor}', confirmed_action='{confirmed_action}', temp_category_id={temp_category_id}")

        product_data_for_confirmation = {
            "name": name, "quantity_str": quantity_str, "purchase_date_str": purchase_date_str,
            "expiry_days": expiry_days, "category": category, "subcategory": subcategory,
            "par_level": par_level, "max_holding_amount": max_holding_amount,
            "purchase_location": purchase_location, "unit_of_measure": unit_of_measure,
            "cost_per_unit_str": cost_per_unit_str, "vendor": vendor # Include new fields
        }

        cost_per_unit_float = None
        if cost_per_unit_str:
            try:
                cost_per_unit_float = float(cost_per_unit_str)
                if cost_per_unit_float < 0:
                    # Add to warnings or return error, for now let's make it an error that stops processing this item
                    return {"success": False, "message": f"Cost per unit for '{name}' cannot be negative ('{cost_per_unit_str}'). Item not added.", "warnings": []} # Assuming warnings is part of return structure
            except ValueError:
                 return {"success": False, "message": f"Cost per unit for '{name}' is not a valid number ('{cost_per_unit_str}'). Item not added.", "warnings": []}


        product_info = self.get_product_by_name(name)
        product_id_to_use = None
        category_id_to_use = None
        subcategory_id_to_use = None
        action_required = None
        confirmation_details = {}
        warnings = []

        if product_info:
            # Existing product logic (remains largely the same)
            product_id_to_use = product_info['id']
            category_id_to_use = product_info.get('category_id')
            subcategory_id_to_use = product_info.get('subcategory_id')
            print(f"LOG: Product '{name}' found. ID: {product_id_to_use}. CategoryID: {product_info.get('category_id')}, SubcategoryID: {product_info.get('subcategory_id')}")

            if category and category_id_to_use:
                db_category_name = self.get_category_name_by_id(category_id_to_use)
                if db_category_name and db_category_name.lower() != category.strip().lower():
                    warnings.append(f"Category '{category.strip()}' in Excel for existing product '{name}' differs from DB category '{db_category_name}'. DB category retained.")

            if subcategory and subcategory_id_to_use:
                # (warnings for category/UoM mismatch can stay)
                pass # End of existing product logic, will proceed to common inventory addition

            if unit_of_measure and unit_of_measure.strip() and product_info.get('unit_of_measure') != unit_of_measure.strip():
                warnings.append(f"UoM for '{name}' in Excel ('{unit_of_measure.strip()}') differs from DB ('{product_info.get('unit_of_measure')}'). DB UoM retained.")

        else: # New Product
            print(f"LOG: Product '{name}' is new. Excel Category='{category}', Excel Subcategory='{subcategory}'")
            if not category or not category.strip():
                print(f"LOG: Returning due to missing category name for new product '{name}'.")
                return {"success": False, "message": f"Category name is missing in Excel for new product '{name}'. Product not added.", "warnings": warnings}

            excel_category_name = category.strip()
            excel_subcategory_name = subcategory.strip() if subcategory and subcategory.strip() else None

            # Flag to indicate if we should proceed to product creation
            can_create_product = False

            print(f"LOG_DIAGNOSTIC: Checking confirmed_action. Value: '{confirmed_action}', Type: {type(confirmed_action)}") # Existing diagnostic log

            # New more detailed diagnostic logs:
            print(f"LOG_DIAGNOSTIC_REPR: repr(confirmed_action): {repr(confirmed_action)}")
            print(f"LOG_DIAGNOSTIC_STRIP_CMP: (confirmed_action.strip() == 'confirm_new_category'): {confirmed_action.strip() == 'confirm_new_category' if isinstance(confirmed_action, str) else 'confirmed_action is not a string'}")
            print(f"LOG_DIAGNOSTIC_DIRECT_CMP: (confirmed_action == 'confirm_new_category'): {confirmed_action == 'confirm_new_category'}")

            print(f"LOG_PRE_CONFIRM_CHECK: confirmed_action='{confirmed_action}' (type: {type(confirmed_action)}), temp_category_id='{temp_category_id}' (type: {type(temp_category_id)})")
            if isinstance(confirmed_action, str) and confirmed_action.strip() == "confirm_new_category":
                print(f"LOG: Handling confirmed_new_category for '{excel_category_name}'")
                add_cat_result = self.add_category(excel_category_name)
                print(f"LOG: add_category result: {add_cat_result}")
                if not add_cat_result.get("success"):
                    if "already exists" in add_cat_result.get("message", "").lower(): # Check if it failed due to already existing
                        print(f"LOG: Category '{excel_category_name}' already exists. Fetching its ID.")
                        existing_category_obj = self.get_category_by_name(excel_category_name)
                        if existing_category_obj:
                            category_id_to_use = existing_category_obj['id']
                            print(f"LOG: Successfully fetched existing category ID: {category_id_to_use}")
                        else:
                            # This would be an unexpected state: add_category said it exists, but get_category_by_name can't find it.
                            print(f"LOG: ERROR - Category '{excel_category_name}' reported as existing, but could not be fetched.")
                            return {"success": False, "message": f"Error resolving already existing category '{excel_category_name}'.", "warnings": warnings}
                    else:
                        # Original failure reason, not "already exists"
                        print(f"LOG: Returning due to add_category failure (Reason: {add_cat_result.get('message')}).")
                        return {"success": False, "message": f"Failed to create new category '{excel_category_name}': {add_cat_result.get('message')}", "warnings": warnings}
                else: # add_category was successful
                    category_id_to_use = add_cat_result['category_id']
                # category_id_to_use should now be set, either from new creation or fetched existing.
                print(f"LOG: category_id_to_use is now {category_id_to_use} for '{excel_category_name}'.")

                if excel_subcategory_name:
                    print(f"LOG: Handling subcategory '{excel_subcategory_name}' for category ID {category_id_to_use}")
                    add_subcat_result = self.add_subcategory(excel_subcategory_name, category_id_to_use)
                    print(f"LOG: add_subcategory result: {add_subcat_result}")
                    if not add_subcat_result.get("success"):
                        if "already exists" in add_subcat_result.get("message", "").lower():
                            print(f"LOG: Subcategory '{excel_subcategory_name}' already exists for category ID {category_id_to_use}. Fetching its ID.")
                            existing_subcategory_obj = self.get_subcategory_by_name_and_category_id(excel_subcategory_name, category_id_to_use)
                            if existing_subcategory_obj:
                                subcategory_id_to_use = existing_subcategory_obj['id']
                                print(f"LOG: Successfully fetched existing subcategory ID: {subcategory_id_to_use}")
                            else:
                                print(f"LOG: ERROR - Subcategory '{excel_subcategory_name}' reported as existing for cat ID {category_id_to_use}, but could not be fetched.")
                                return {"success": False, "message": f"Error resolving already existing subcategory '{excel_subcategory_name}'.", "warnings": warnings}
                        else:
                            print(f"LOG: Returning due to add_subcategory failure (Reason: {add_subcat_result.get('message')}).")
                            return {"success": False, "message": f"Failed to create new subcategory '{excel_subcategory_name}' for category '{excel_category_name}': {add_subcat_result.get('message')}", "warnings": warnings}
                    else: # add_subcategory was successful
                        subcategory_id_to_use = add_subcat_result['subcategory_id']
                    print(f"LOG: subcategory_id_to_use is now {subcategory_id_to_use} for '{excel_subcategory_name}'.")
                can_create_product = True
                print(f"LOG: can_create_product set to True after confirmed_new_category processing.")

            elif isinstance(confirmed_action, str) and confirmed_action.strip() == "confirm_new_subcategory" and temp_category_id is not None:
                print(f"LOG_ENTERED_CONFIRM_SUBCATEGORY_BLOCK: Processing confirmed new subcategory '{excel_subcategory_name}' for category ID {temp_category_id}.")
                print(f"LOG: Handling confirmed_new_subcategory for '{excel_subcategory_name}' under temp_category_id {temp_category_id}")
                category_id_to_use = temp_category_id
                print(f"LOG: category_id_to_use set to {category_id_to_use} (from temp_category_id).")
                if not excel_subcategory_name: # Should ideally not happen if this action was triggered
                    print(f"LOG: Returning due to missing subcategory name for confirmed_new_subcategory.")
                    return {"success": False, "message": f"Subcategory name missing for confirmed_new_subcategory action for product '{name}'.", "warnings": warnings}

                add_subcat_result = self.add_subcategory(excel_subcategory_name, category_id_to_use)
                print(f"LOG: add_subcategory result (confirmed_new_subcategory path): {add_subcat_result}")
                if not add_subcat_result.get("success"):
                    if "already exists" in add_subcat_result.get("message", "").lower():
                        print(f"LOG: Subcategory '{excel_subcategory_name}' (confirmed path) already exists for category ID {category_id_to_use}. Fetching its ID.")
                        existing_subcategory_obj = self.get_subcategory_by_name_and_category_id(excel_subcategory_name, category_id_to_use)
                        if existing_subcategory_obj:
                            subcategory_id_to_use = existing_subcategory_obj['id']
                            print(f"LOG: Successfully fetched existing subcategory ID (confirmed path): {subcategory_id_to_use}")
                        else:
                            print(f"LOG: ERROR - Subcategory '{excel_subcategory_name}' (confirmed path) reported as existing for cat ID {category_id_to_use}, but could not be fetched.")
                            return {"success": False, "message": f"Error resolving already existing subcategory '{excel_subcategory_name}' (confirmed path).", "warnings": warnings}
                    else:
                        print(f"LOG: Returning due to add_subcategory failure (confirmed_new_subcategory path, Reason: {add_subcat_result.get('message')}).")
                        return {"success": False, "message": f"Failed to create new subcategory '{excel_subcategory_name}': {add_subcat_result.get('message')}", "warnings": warnings}
                else: # add_subcategory was successful
                    subcategory_id_to_use = add_subcat_result['subcategory_id']
                print(f"LOG: subcategory_id_to_use (confirmed_new_subcategory path) is now {subcategory_id_to_use} for '{excel_subcategory_name}'.")
                can_create_product = True
                print(f"LOG: can_create_product set to True after confirmed_new_subcategory processing.")

            else: # No confirmed_action, this is the initial check
                print(f"LOG: No confirmed_action. Initial check for category '{excel_category_name}'.")
                existing_category_obj = self.get_category_by_name(excel_category_name)
                if existing_category_obj:
                    print(f"LOG: Existing category found: ID {existing_category_obj['id']}")
                    category_id_to_use = existing_category_obj['id']
                    if excel_subcategory_name:
                        print(f"LOG: Checking for existing subcategory '{excel_subcategory_name}' under category ID {category_id_to_use}.")
                        existing_subcategory_obj = self.get_subcategory_by_name_and_category_id(excel_subcategory_name, category_id_to_use)
                        if existing_subcategory_obj:
                            print(f"LOG: Existing subcategory found: ID {existing_subcategory_obj['id']}")
                            subcategory_id_to_use = existing_subcategory_obj['id']
                            can_create_product = True
                            print(f"LOG: can_create_product set to True. Category and Subcategory exist.")
                        else: # New subcategory for existing category
                            print(f"LOG: New subcategory '{excel_subcategory_name}' for existing category ID {category_id_to_use}. Requesting confirmation.")
                            action_required = "confirm_new_subcategory"
                            confirmation_details = {
                                "category_id": category_id_to_use,
                                "category_name": existing_category_obj['name'],
                                "new_subcategory_name": excel_subcategory_name
                            }
                            print(f"LOG: Returning action_required='{action_required}'")
                            return {"success": False, "action_required": action_required, "confirmation_details": confirmation_details, "product_data": product_data_for_confirmation, "warnings": warnings}
                    else: # No subcategory provided, existing category is enough
                        can_create_product = True
                        print(f"LOG: can_create_product set to True. Category exists, no subcategory provided.")
                else: # New category
                    print(f"LOG: New category '{excel_category_name}'. Requesting confirmation.")
                    action_required = "confirm_new_category"
                    confirmation_details = {
                        "new_category_name": excel_category_name,
                        "new_subcategory_name": excel_subcategory_name # Pass along subcategory name if provided
                    }
                    print(f"LOG: Returning action_required='{action_required}'")
                    return {"success": False, "action_required": action_required, "confirmation_details": confirmation_details, "product_data": product_data_for_confirmation, "warnings": warnings}

            # Create the product if it's new and category/subcategory are resolved
            print(f"LOG: Before create_product call. can_create_product={can_create_product}. Name='{name}', CategoryID={category_id_to_use}, SubcategoryID={subcategory_id_to_use}, UoM='{unit_of_measure}'")
            if can_create_product:
                if not unit_of_measure:
                    print(f"LOG: Returning due to missing UoM for new product '{name}'.")
                    return {"success": False, "message": f"Unit of Measure is missing in Excel for new product '{name}'. Product not added.", "warnings": warnings}

                create_product_result = self.create_product(
                    name=name,
                    category_id=category_id_to_use,
                    subcategory_id=subcategory_id_to_use,
                    unit_of_measure=unit_of_measure,
                    default_expiry_days=expiry_days, # This comes from Excel row
                    par_level=par_level,
                    max_holding_amount=max_holding_amount,
                    purchase_location=purchase_location
                )
                print(f"LOG: create_product result: {create_product_result}")
                if create_product_result.get("success"):
                    product_id_to_use = create_product_result.get("product_id")
                    print(f"LOG: Product '{name}' created successfully with ID {product_id_to_use}.") # Note: This is user-facing print, distinct from LOG
                else:
                    print(f"LOG: Returning due to create_product failure for '{name}'.")
                    return {"success": False, "message": create_product_result.get("message", f"Failed to create product '{name}'."), "warnings": warnings}
            else:
                # This case should ideally not be reached if logic is correct,
                # as all paths should either return an action_required, an error, or set can_create_product to True.
                # If it is reached, it means a logic path for new products didn't result in product creation or a request for confirmation.
                print(f"LOG: Product not created because can_create_product is False. Product name: '{name}'.")
                return {"success": False, "message": f"Internal logic error: Could not determine action for new product '{name}'. Product not created.", "warnings": warnings}

        # --- Common logic for adding inventory item once product_id_to_use is determined ---
        print(f"LOG: Before adding to inventory_items. ProductIDToUse={product_id_to_use}, Name='{name}', Qty='{quantity_str}'")
        if product_id_to_use is None:
             # This case should ideally be caught by earlier checks (e.g., product not found and action not confirmed)
            print(f"LOG: Returning because product_id_to_use is None before final inventory add for '{name}'.")
            return {"success": False, "message": f"Could not determine product ID for '{name}'. Inventory item not added.", "warnings": warnings}

        # If cost_per_unit is provided, use log_purchase, which handles both PurchaseLog and inventory_items.
        # Otherwise, use the old logic to just add to inventory_items.
        if cost_per_unit_float is not None:
            print(f"LOG: Cost provided for '{name}'. Using log_purchase.")
            # Ensure quantity_str is parsed to float for log_purchase
            try:
                quantity_purchased_float = self._parse_quantity_string(quantity_str)
                if quantity_purchased_float <= 0: # log_purchase also checks this, but good for consistency
                    return {"success": False, "message": f"Quantity for '{name}' must be positive ('{quantity_str}'). Item not added.", "warnings": warnings}
            except ValueError: # Should be caught by _parse_quantity_string if it returns non-numeric or error
                 return {"success": False, "message": f"Invalid quantity format for '{name}' ('{quantity_str}'). Item not added.", "warnings": warnings}


            log_purchase_result = self.log_purchase(
                product_id=product_id_to_use,
                purchase_date_str=purchase_date_str,
                quantity_purchased_float=quantity_purchased_float,
                cost_per_unit_float=cost_per_unit_float,
                vendor_str=vendor
            )
            # Adapt the return message to be consistent with add_item_to_list's expectations
            if log_purchase_result.get("success"):
                return {
                    "success": True,
                    "message": f"Item '{name}' purchase logged and stock added.",
                    "item_id": log_purchase_result.get("stock_item_id"), # Use stock_item_id from log_purchase
                    "product_id": product_id_to_use,
                    "purchase_log_id": log_purchase_result.get("purchase_log_id"),
                    "warnings": warnings # Pass along any warnings accumulated during product/category resolution
                }
            else:
                # If log_purchase failed, append its message to warnings or return directly
                # For now, returning its message as the main message
                return {
                    "success": False,
                    "message": log_purchase_result.get("message", f"Failed to log purchase for '{name}'."),
                    "warnings": warnings
                }
        else: # No cost provided, just add to inventory as before
            print(f"LOG: No cost provided for '{name}'. Adding directly to inventory_items.")
            try:
                purchase_dt = date.fromisoformat(purchase_date_str)
                batch_expiry_dt = purchase_dt + timedelta(days=int(expiry_days))
            except (ValueError, TypeError) as e: # Catches if expiry_days is not int-convertible too
                print(f"LOG: Returning due to invalid date or expiry days for item '{name}'. Error: {e}")
                return {"success": False, "message": f"Invalid purchase date '{purchase_date_str}' or expiry days '{expiry_days}' for item '{name}': {e}", "warnings": warnings}

            try:
                with self._get_db_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute('''
                        INSERT INTO inventory_items
                        (product_id, name, quantity, purchase_date, expiry_date, original_quantity_string)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (product_id_to_use, name, str(quantity_str), purchase_dt.isoformat(),
                          batch_expiry_dt.isoformat(), str(quantity_str)))
                    conn.commit()
                    item_id = cursor.lastrowid
                    print(f"LOG: Successfully added item '{name}' (Batch ID: {item_id}) to inventory_items (no cost). Expires: {batch_expiry_dt.isoformat()}")
                    return {"success": True, "message": f"Item '{name}' added to inventory (no cost info).", "item_id": item_id, "product_id": product_id_to_use, "warnings": warnings}
            except sqlite3.Error as e:
                print(f"LOG: Returning due to database error adding item '{name}' to inventory_items (no cost). Error: {e}")
                return {"success": False, "message": f"Database error adding item '{name}' to inventory (no cost): {e}", "warnings": warnings}


    def get_current_inventory(self, search_term=None, category=None, purchase_location=None,
                              sort_by='p.name', sort_order='ASC',
                              page=1, per_page=10):
        """
        Retrieves aggregated product data from the current inventory.
        - search_term: Filters by product name (p.name).
        - category: Filters by category name (c.name).
        - purchase_location: Filters by product purchase location (p.purchase_location).
        - sort_by: Column to sort by (e.g., 'p.name', 'c.name'). Defaults to 'p.name'.
        - sort_order: 'ASC' or 'DESC'. Defaults to 'ASC'.
        - page: For pagination.
        - per_page: For pagination.
        Returns a list of dictionaries, each representing a product with its aggregated data.
        """
        items = []
        params = []

        base_query = """
            SELECT
                p.id AS product_id,
                p.name AS product_name,
                p.unit_of_measure,
                p.par_level,
                c.name AS category_name,
                sc.name AS subcategory_name,
                SUM(CAST(ii.quantity AS REAL)) AS total_quantity
            FROM products p
            JOIN inventory_items ii ON p.id = ii.product_id
            LEFT JOIN categories c ON p.category_id = c.id
            LEFT JOIN subcategories sc ON p.subcategory_id = sc.id
        """

        where_clauses = []
        if search_term:
            where_clauses.append("LOWER(p.name) LIKE ?")
            params.append(f"%{search_term.lower()}%")
        if category:
            where_clauses.append("LOWER(c.name) = ?")
            params.append(category.lower())
        if purchase_location:
            where_clauses.append("LOWER(p.purchase_location) = ?")
            params.append(purchase_location.lower())

        if where_clauses:
            base_query += " WHERE " + " AND ".join(where_clauses)

        base_query += """
            GROUP BY p.id, p.name, p.unit_of_measure, p.par_level, c.name, sc.name
        """

        # Sorting
        # Valid sort columns should now refer to product or category attributes
        valid_sort_columns = {
            'p.name': 'p.name',
            'c.name': 'c.name',
            'total_quantity': 'total_quantity', # Can sort by aggregated quantity
            'p.par_level': 'p.par_level'
        }
        sort_column = valid_sort_columns.get(sort_by.lower(), 'p.name')

        sort_order_upper = sort_order.upper()
        if sort_order_upper not in ['ASC', 'DESC']:
            sort_order_upper = 'ASC'

        base_query += f" ORDER BY {sort_column} {sort_order_upper}, p.id {sort_order_upper}"

        # Pagination
        if page is not None and per_page is not None and page > 0 and per_page > 0:
            offset = (page - 1) * per_page
            base_query += " LIMIT ? OFFSET ?"
            params.extend([per_page, offset])
        elif per_page is not None and per_page > 0: # Apply limit if only per_page is specified
            base_query += " LIMIT ?"
            params.append(per_page)


        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(base_query, tuple(params))
                rows = cursor.fetchall()
                for row in rows:
                    items.append(dict(row))
        except sqlite3.Error as e:
            print(f"Database error fetching aggregated current inventory: {e}")
        return items

    def get_current_inventory_count(self, search_term=None, category=None, purchase_location=None):
        """
        Gets the total count of distinct products in current inventory based on filters.
        """
        params = []
        # Query to count distinct products after filtering
        query = """
            SELECT COUNT(DISTINCT p.id) as count
            FROM products p
            JOIN inventory_items ii ON p.id = ii.product_id
            LEFT JOIN categories c ON p.category_id = c.id
            LEFT JOIN subcategories sc ON p.subcategory_id = sc.id
        """

        where_clauses = []
        if search_term:
            where_clauses.append("LOWER(p.name) LIKE ?")
            params.append(f"%{search_term.lower()}%")
        if category:
            where_clauses.append("LOWER(c.name) = ?")
            params.append(category.lower())
        if purchase_location:
            where_clauses.append("LOWER(p.purchase_location) = ?")
            params.append(purchase_location.lower())

        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(query, tuple(params))
                result = cursor.fetchone()
                return result['count'] if result else 0
        except sqlite3.Error as e:
            print(f"Database error getting distinct product count from inventory: {e}")
            return 0

    def get_current_inventory_categories(self):
        """Retrieves unique categories from products currently in inventory."""
        categories = []
        query = """
            SELECT DISTINCT c.name
            FROM inventory_items ii
            JOIN products p ON ii.product_id = p.id
            JOIN categories c ON p.category_id = c.id
            WHERE c.name IS NOT NULL
            ORDER BY c.name ASC
        """
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(query)
                rows = cursor.fetchall()
                for row in rows:
                    categories.append(row['name']) # Select c.name
        except sqlite3.Error as e:
            print(f"Database error fetching current inventory categories: {e}")
        return categories

    def get_current_inventory_purchase_locations(self):
        """Retrieves unique purchase locations from products currently in inventory."""
        locations = []
        query = """
            SELECT DISTINCT p.purchase_location
            FROM products p
            JOIN inventory_items ii ON p.id = ii.product_id
            WHERE p.purchase_location IS NOT NULL
            ORDER BY p.purchase_location ASC
        """
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(query)
                rows = cursor.fetchall()
                for row in rows:
                    locations.append(row['purchase_location'])
        except sqlite3.Error as e:
            print(f"Database error fetching current inventory purchase locations: {e}")
        return locations

    def get_inventory_batches_with_cost(self, search_term=None, category=None, purchase_location=None,
                                   sort_by='product_name', sort_order='ASC',
                                   page=1, per_page=20,
                                   start_purchase_date=None, end_purchase_date=None,
                                   start_expiry_date=None, end_expiry_date=None):
        """
        Retrieves current inventory batches with associated product details and, if available,
        the cost_per_unit and vendor from the most relevant PurchaseLog entry.
        """
        items = []
        params = []

        # This query aims to get each inventory_item and join it with product details.
        # Then, for each batch, it attempts to find the *latest* PurchaseLog entry
        # for that product_id that occurred on or before the batch's purchase_date.
        # This is an approximation if direct batch-to-purchase_log linking isn't perfect.
        # Using a subquery or a more complex join might be needed for precise cost matching if
        # multiple purchases happen on the same day for the same product.
        # For simplicity, this initial version might show the weighted average cost or latest direct cost.

        # Let's try to get the cost from the purchase log that matches product_id and purchase_date.
        # If multiple purchases for the same product on the same day, this could be ambiguous.
        # A more robust solution might involve storing purchase_log_id in inventory_items,
        # or using weighted average cost at the time of batch creation.

        # For this version, we'll fetch all batches and then try to enrich with cost.
        # This two-step approach might be less efficient for large datasets but simpler to implement initially.

        # Step 1: Fetch filtered and paginated inventory batches with product details.
        query = """
            SELECT
                ii.id AS batch_id,
                ii.product_id,
                p.name AS product_name,
                c.name AS category_name,
                sc.name AS subcategory_name,
                p.unit_of_measure,
                ii.quantity,
                ii.purchase_date,
                ii.expiry_date,
                p.par_level,
                p.purchase_location AS product_purchase_location
            FROM inventory_items ii
            JOIN products p ON ii.product_id = p.id
            LEFT JOIN categories c ON p.category_id = c.id
            LEFT JOIN subcategories sc ON p.subcategory_id = sc.id
        """
        where_clauses = ["CAST(ii.quantity AS REAL) > 0"] # Only include items with quantity > 0

        if search_term:
            where_clauses.append("LOWER(p.name) LIKE ?")
            params.append(f"%{search_term.lower()}%")
        if category:
            where_clauses.append("LOWER(c.name) = ?")
            params.append(category.lower())
        if purchase_location: # This filters by the product's default purchase location
            where_clauses.append("LOWER(p.purchase_location) = ?")
            params.append(purchase_location.lower())
        if start_purchase_date:
            where_clauses.append("ii.purchase_date >= ?")
            params.append(start_purchase_date)
        if end_purchase_date:
            where_clauses.append("ii.purchase_date <= ?")
            params.append(end_purchase_date)
        if start_expiry_date:
            where_clauses.append("ii.expiry_date >= ?")
            params.append(start_expiry_date)
        if end_expiry_date:
            where_clauses.append("ii.expiry_date <= ?")
            params.append(end_expiry_date)

        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)

        # Sorting - apply to the main query
        # Valid sort columns for batch view
        valid_sort_columns = {
            'product_name': 'p.name',
            'category_name': 'c.name',
            'quantity': 'CAST(ii.quantity AS REAL)', # Ensure numeric sort for quantity
            'purchase_date': 'ii.purchase_date',
            'expiry_date': 'ii.expiry_date',
            # 'cost_per_unit' would require joining PurchaseLog here or sorting in Python later
        }
        sort_column_sql = valid_sort_columns.get(sort_by.lower(), 'p.name')
        sort_order_sql = 'DESC' if sort_order.upper() == 'DESC' else 'ASC'
        query += f" ORDER BY {sort_column_sql} {sort_order_sql}, ii.id {sort_order_sql}"

        # Pagination - apply to the main query
        if page is not None and per_page is not None and page > 0 and per_page > 0:
            offset = (page - 1) * per_page
            query += " LIMIT ? OFFSET ?"
            params.extend([per_page, offset])

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(query, tuple(params))
                raw_batches = cursor.fetchall()

                for row_data in raw_batches:
                    batch_dict = dict(row_data)
                    batch_dict['cost_per_unit'] = None # Default
                    batch_dict['vendor'] = None      # Default

                    # Attempt to find a matching purchase log entry
                    # This is a simplification: it finds the first purchase log for that product on that day.
                    # If multiple purchases on the same day, this might not be the exact cost for *this specific batch*.
                    # A more accurate system would link inventory_items.id to PurchaseLog.id upon creation.
                    cursor.execute("""
                        SELECT cost_per_unit, vendor
                        FROM PurchaseLog
                        WHERE product_id = ? AND purchase_date = ?
                        ORDER BY id DESC  -- Get the latest purchase on that day if multiple
                        LIMIT 1
                    """, (batch_dict['product_id'], batch_dict['purchase_date']))
                    cost_row = cursor.fetchone()

                    if cost_row:
                        batch_dict['cost_per_unit'] = cost_row['cost_per_unit']
                        batch_dict['vendor'] = cost_row['vendor']
                    else:
                        # Fallback: use weighted average cost if no direct match on purchase_date
                        # This might be too broad if purchase_date is the primary key for cost.
                        # Consider if this fallback is desired. For now, if no direct match, cost remains None.
                        # batch_dict['cost_per_unit'] = self.get_weighted_average_cost(batch_dict['product_id'])
                        pass


                    # Convert date strings to date objects for consistent handling in template if needed
                    # However, for display, strings are often fine. Let's keep as string from DB for now.
                    # If date objects are needed:
                    # if batch_dict.get('purchase_date'): batch_dict['purchase_date'] = date.fromisoformat(batch_dict['purchase_date'])
                    # if batch_dict.get('expiry_date'): batch_dict['expiry_date'] = date.fromisoformat(batch_dict['expiry_date'])
                    items.append(batch_dict)
        except sqlite3.Error as e:
            print(f"Database error fetching inventory batches with cost: {e}")
            # Potentially return an error status or empty list
        return items

    def get_inventory_batches_with_cost_count(self, search_term=None, category=None, purchase_location=None,
                                             start_purchase_date=None, end_purchase_date=None,
                                             start_expiry_date=None, end_expiry_date=None):
        """
        Gets the total count of current inventory batches based on filters.
        """
        params = []
        query = """
            SELECT COUNT(ii.id) as count
            FROM inventory_items ii
            JOIN products p ON ii.product_id = p.id
            LEFT JOIN categories c ON p.category_id = c.id
            LEFT JOIN subcategories sc ON p.subcategory_id = sc.id
        """
        where_clauses = ["CAST(ii.quantity AS REAL) > 0"] # Only count items with quantity > 0

        if search_term:
            where_clauses.append("LOWER(p.name) LIKE ?")
            params.append(f"%{search_term.lower()}%")
        if category:
            where_clauses.append("LOWER(c.name) = ?")
            params.append(category.lower())
        if purchase_location:
            where_clauses.append("LOWER(p.purchase_location) = ?")
            params.append(purchase_location.lower())
        if start_purchase_date:
            where_clauses.append("ii.purchase_date >= ?")
            params.append(start_purchase_date)
        if end_purchase_date:
            where_clauses.append("ii.purchase_date <= ?")
            params.append(end_purchase_date)
        if start_expiry_date:
            where_clauses.append("ii.expiry_date >= ?")
            params.append(start_expiry_date)
        if end_expiry_date:
            where_clauses.append("ii.expiry_date <= ?")
            params.append(end_expiry_date)

        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(query, tuple(params))
                result = cursor.fetchone()
                return result['count'] if result else 0
        except sqlite3.Error as e:
            print(f"Database error getting inventory batch count: {e}")
            return 0

    def get_inventory_batches_for_product(self, product_id, limit=None, order_by_purchase_desc=False, order_by_id_desc=False):
        """
        Retrieves specific inventory batches for a given product_id.
        Can be ordered by purchase_date descending or id descending, and limited.
        """
        items = []
        if not product_id: # Ensure product_id is not None or empty before query
            return items

        # Ensure product_id is an integer if it's coming from a form/URL
        try:
            valid_product_id = int(product_id)
        except ValueError:
            print(f"Invalid product_id format: {product_id}")
            return items # Or raise an error

        query = '''
            SELECT
                ii.id, ii.product_id, ii.quantity, ii.purchase_date, ii.expiry_date,
                ii.original_quantity_string,
                p.name AS product_name,
                c.name AS category_name,
                s.name AS subcategory_name,
                p.unit_of_measure
            FROM inventory_items ii
            JOIN products p ON ii.product_id = p.id
            LEFT JOIN categories c ON p.category_id = c.id
            LEFT JOIN subcategories s ON p.subcategory_id = s.id
            WHERE ii.product_id = ?
        '''
        params = [valid_product_id]

        if order_by_id_desc:
            query += " ORDER BY ii.id DESC"
        elif order_by_purchase_desc:
            query += " ORDER BY ii.purchase_date DESC, ii.id DESC"
        else: # Default order (e.g., by expiry date as in get_current_inventory)
            query += " ORDER BY ii.expiry_date ASC, ii.id ASC"


        if limit is not None:
            query += " LIMIT ?"
            params.append(limit)

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(query, tuple(params))
                rows = cursor.fetchall()
                for row in rows:
                    item = dict(row)
                    # Convert date strings to date objects
                    if item.get('purchase_date'):
                        item['purchase_date'] = date.fromisoformat(item['purchase_date'])
                    if item.get('expiry_date'):
                        item['expiry_date'] = date.fromisoformat(item['expiry_date'])
                    items.append(item)
        except sqlite3.Error as e:
            print(f"Database error fetching inventory batches for product ID {valid_product_id}: {e}")
        return items

    def get_shopping_list_items(self, store_filter=None, search_term=None):
        """
        Generates a shopping list based on current inventory, consumption rates,
        par levels, and purchase locations.
        """
        SOBEYS_FREQUENCY_WEEKS = 1
        COSTCO_FREQUENCY_WEEKS = 3
        shopping_list = []

        # get_current_inventory() now returns items with product details joined
        inventory_with_product_details = self.get_current_inventory()

        # We need a unique list of product_ids that are below par to avoid multiple projections for the same product
        # if it has multiple inventory batches.
        # However, par_level is per product, so we should iterate through products that have par levels.

        all_products = self.get_all_products()
        products_to_check = [p for p in all_products if p.get('par_level', 0) is not None and p.get('par_level', 0) > 0]

        for product in products_to_check:
            product_id = product['id']
            product_name = product['name']
            par_level = product.get('par_level', 0.0)
            purchase_location = product.get('purchase_location')
            unit_of_measure = product.get('unit_of_measure', 'units')

            # Skip if par_level is not set or not positive
            if par_level <= 0:
                continue

            projection_days_for_item = 0
            if purchase_location == 'Sobeys':
                projection_days_for_item = SOBEYS_FREQUENCY_WEEKS * 7
            elif purchase_location == 'Costco':
                projection_days_for_item = COSTCO_FREQUENCY_WEEKS * 7
            else:
                # If purchase location doesn't match known cycles, skip or use a default.
                # For now, skipping.
                continue

            # Use project_demand with product_id
            demand_projection = self.project_demand(product_id, lookback_days=30, projection_days=projection_days_for_item)

            if not demand_projection.get("success"):
                print(f"Skipping shopping list item for {product_name} due to projection error: {demand_projection.get('message')}")
                continue

            avg_daily_consumption = demand_projection.get('avg_daily_consumption', 0.0)
            # current_numeric_quantity is the total stock for this product_id
            current_numeric_quantity = self.get_total_item_quantity(product_id)

            target_stock_after_shopping = par_level + (avg_daily_consumption * projection_days_for_item)
            recommended_purchase_amount = target_stock_after_shopping - current_numeric_quantity
            recommended_purchase_amount = max(0, round(recommended_purchase_amount, 2))

            if recommended_purchase_amount > 0:
                earliest_expiry_date_str = "N/A"
                try:
                    with self._get_db_connection() as conn_inner: # Use a new connection or ensure thread safety if reusing
                        cursor_inner = conn_inner.cursor()
                        cursor_inner.execute('''
                            SELECT MIN(expiry_date) as min_expiry
                            FROM inventory_items
                            WHERE product_id = ? AND CAST(quantity AS REAL) > 0
                        ''', (product_id,))
                        expiry_row = cursor_inner.fetchone()
                        if expiry_row and expiry_row['min_expiry']:
                            earliest_expiry_date_str = expiry_row['min_expiry']
                except sqlite3.Error as e_expiry:
                    print(f"Database error getting earliest expiry date for product ID {product_id} in shopping list: {e_expiry}")
                    # earliest_expiry_date_str remains "N/A"

                # Current quantity display string (sum of original strings might be complex, so just use numeric total with unit)
                current_quantity_display = f"{current_numeric_quantity:.2f} {unit_of_measure}"


                shopping_list_item = {
                    'product_id': product_id,
                    'name': product_name, # This is product_name from products table
                    'current_quantity_display': current_quantity_display,
                    'current_numeric_quantity': current_numeric_quantity,
                    'unit_of_measure': unit_of_measure,
                    'purchase_location': purchase_location,
                    'earliest_expiry_date': earliest_expiry_date_str, # Informational
                    'recommended_purchase_amount': recommended_purchase_amount,
                    'par_level': par_level,
                    'days_to_next_shop': projection_days_for_item,
                    'avg_daily_consumption': round(avg_daily_consumption, 2)
                }
                shopping_list.append(shopping_list_item)

        # Filter logic (remains largely the same, but fields might have changed names slightly)
        if store_filter:
            shopping_list = [
                item for item in shopping_list
                if item['purchase_location'] and item['purchase_location'].lower() == store_filter.lower()
            ]

        if search_term:
            shopping_list = [
                item for item in shopping_list
                if search_term.lower() in item['name'].lower()
            ]

        return shopping_list

    def get_historical_inventory(self, search_term=None, category=None,
                                 consumed_start_date=None, consumed_end_date=None,
                                 sort_by='consumed_date', sort_order='DESC',
                                 page=1, per_page=10,
                                 export_all=False, export_start_date_str=None, export_end_date_str=None):
        """
        Retrieves items from historical inventory, joined with product details,
        with filtering, sorting, and pagination.
        - search_term: Filters by product name (p.name or hi.name).
        - category: Filters by product category (p.category).
        - consumed_start_date: Filters for hi.consumed_date >= this date.
        - consumed_end_date: Filters for hi.consumed_date <= this date.
        - sort_by: ('product_name', 'category', 'quantity_consumed', 'consumed_date').
        - sort_order: 'ASC' or 'DESC'.
        - page: For pagination.
        - per_page: For pagination.
        """
        items = []
        params = []

        base_query = """
            SELECT
                hi.id, hi.product_id,
                COALESCE(p.name, hi.name) AS product_display_name,
                hi.quantity_consumed_this_time, hi.original_quantity_string,
                hi.purchase_date, hi.expiry_date, hi.consumed_date, hi.cost_of_goods_used,
                p.unit_of_measure,
                c.name AS category_name,
                sc.name AS subcategory_name
            FROM historical_items hi
            LEFT JOIN products p ON hi.product_id = p.id
            LEFT JOIN categories c ON p.category_id = c.id
            LEFT JOIN subcategories sc ON p.subcategory_id = sc.id
        """

        where_clauses = []
        if search_term:
            where_clauses.append("LOWER(COALESCE(p.name, hi.name)) LIKE ?")
            params.append(f"%{search_term.lower()}%")
        if category: # Filter by category NAME
            # This will only include items that have a linked product and category
            where_clauses.append("c.name IS NOT NULL AND LOWER(c.name) = ?")
            params.append(category.lower())
        # Date filtering: Prioritize export dates if export_all is True
        final_start_date = None
        final_end_date = None

        if export_all:
            if export_start_date_str and export_start_date_str.strip():
                try:
                    date.fromisoformat(export_start_date_str.strip()) # Validate
                    final_start_date = export_start_date_str.strip()
                except ValueError:
                    print(f"Warning: Invalid export_start_date_str format '{export_start_date_str}'. It will be ignored.")
            if export_end_date_str and export_end_date_str.strip():
                try:
                    date.fromisoformat(export_end_date_str.strip()) # Validate
                    final_end_date = export_end_date_str.strip()
                except ValueError:
                    print(f"Warning: Invalid export_end_date_str format '{export_end_date_str}'. It will be ignored.")
        else:
            # Use regular pagination/filter dates if not exporting all
            if consumed_start_date and consumed_start_date.strip():
                final_start_date = consumed_start_date
            if consumed_end_date and consumed_end_date.strip():
                final_end_date = consumed_end_date

        if final_start_date:
            where_clauses.append("hi.consumed_date >= ?")
            params.append(final_start_date)
        if final_end_date:
            where_clauses.append("hi.consumed_date <= ?")
            params.append(final_end_date)

        if where_clauses:
            base_query += " WHERE " + " AND ".join(where_clauses)

        # Sorting
        valid_sort_columns = {
            'product_name': 'product_display_name',
            'category_name': 'c.name', # Updated to category_name
            'quantity_consumed': 'hi.quantity_consumed_this_time',
            'consumed_date': 'hi.consumed_date',
            'purchase_date': 'hi.purchase_date', # Added for completeness
            'expiry_date': 'hi.expiry_date'    # Added for completeness
        }
        sort_column = valid_sort_columns.get(sort_by.lower(), 'hi.consumed_date')

        sort_order_upper = sort_order.upper()
        if sort_order_upper not in ['ASC', 'DESC']:
            sort_order_upper = 'DESC' # Default for historical

        base_query += f" ORDER BY {sort_column} {sort_order_upper}, hi.id {sort_order_upper}"

        # Pagination (skip if export_all is True)
        if not export_all:
            if page is not None and per_page is not None and page > 0 and per_page > 0:
                offset = (page - 1) * per_page
                base_query += " LIMIT ? OFFSET ?"
                params.extend([per_page, offset])
            elif per_page is not None and per_page > 0: # Only per_page is specified
                base_query += " LIMIT ?"
                params.append(per_page)

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(base_query, tuple(params))
                rows = cursor.fetchall()
                for row_data in rows:
                    item = dict(row_data)
                    item['name'] = item['product_display_name'] # Ensure 'name' key is present for consistency
                    if item.get('purchase_date'):
                        item['purchase_date'] = date.fromisoformat(item['purchase_date'])
                    if item.get('expiry_date'):
                        item['expiry_date'] = date.fromisoformat(item['expiry_date'])
                    if item.get('consumed_date'): # Should always be present for historical
                        item['consumed_date'] = date.fromisoformat(item['consumed_date'])
                    items.append(item)
        except sqlite3.Error as e:
            print(f"Database error fetching historical inventory with filters: {e}")
        return items

    def get_historical_inventory_count(self, search_term=None, category=None,
                                       consumed_start_date=None, consumed_end_date=None):
        """
        Gets the total count of historical inventory items based on filters.
        """
        params = []
        query = """
            SELECT COUNT(hi.id) as count
            FROM historical_items hi
            LEFT JOIN products p ON hi.product_id = p.id
            LEFT JOIN categories c ON p.category_id = c.id
        """

        where_clauses = []
        if search_term:
            where_clauses.append("LOWER(COALESCE(p.name, hi.name)) LIKE ?")
            params.append(f"%{search_term.lower()}%")
        if category: # Filter by category NAME
            where_clauses.append("c.name IS NOT NULL AND LOWER(c.name) = ?")
            params.append(category.lower())
        if consumed_start_date:
            where_clauses.append("hi.consumed_date >= ?")
            params.append(consumed_start_date)
        if consumed_end_date:
            where_clauses.append("hi.consumed_date <= ?")
            params.append(consumed_end_date)

        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(query, tuple(params))
                result = cursor.fetchone()
                return result['count'] if result else 0
        except sqlite3.Error as e:
            print(f"Database error getting historical inventory count: {e}")
            return 0

    def get_historical_inventory_categories(self):
        """Retrieves unique categories from products recorded in historical inventory."""
        categories = []
        query = """
            SELECT DISTINCT c.name
            FROM historical_items hi
            JOIN products p ON hi.product_id = p.id
            JOIN categories c ON p.category_id = c.id
            WHERE c.name IS NOT NULL
            ORDER BY c.name ASC
        """
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(query)
                rows = cursor.fetchall()
                for row in rows:
                    categories.append(row['name']) # Select c.name
        except sqlite3.Error as e:
            print(f"Database error fetching historical inventory categories: {e}")
        return categories

    def get_total_item_quantity(self, product_name_or_id):
        """Calculates total quantity of a specific item in current inventory, using product_id or name."""
        total_quantity = 0.0
        product_id = None

        if isinstance(product_name_or_id, int): # Assume it's an ID
            product_id = product_name_or_id
        elif isinstance(product_name_or_id, str): # Assume it's a name
            product = self.get_product_by_name(product_name_or_id)
            if product:
                product_id = product['id']
            else:
                print(f"Product '{product_name_or_id}' not found for quantity check.")
                return 0.0 # Product not found
        else:
            print(f"Invalid identifier type for get_total_item_quantity: {product_name_or_id}")
            return 0.0

        if product_id is None: # Handles case where name was given but not found
             return 0.0

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT quantity FROM inventory_items WHERE product_id = ?", (product_id,))
                rows = cursor.fetchall()
                for row in rows:
                    total_quantity += self._parse_quantity_string(row['quantity'])
        except sqlite3.Error as e:
            print(f"Database error getting total item quantity for product ID {product_id}: {e}")
        return total_quantity

    def consume_item(self, item_name_to_consume, quantity_to_consume_float):
        """
        Consumes a specified quantity of an item from the inventory.
        - Fetches all batches of the item, ordered by expiry date (soonest first).
        - Iterates through batches, consuming the required quantity.
        - Updates item quantities in `inventory_items` table (or deletes if fully consumed).
        - Logs each consumption event (even partial from a batch) to `historical_items`.
        - Returns a dictionary with success status and detailed messages.
        """
        if quantity_to_consume_float <= 0:
            return {"success": False, "message": "Quantity to consume must be positive."}

        log_messages = []
        consumed_amount_total_overall = 0.0
        quantity_remaining_to_consume = float(quantity_to_consume_float)

        product_to_consume = self.get_product_by_name(item_name_to_consume)
        if not product_to_consume:
            return {"success": False, "message": f"Product '{item_name_to_consume}' not found in products table."}

        product_id_to_consume = product_to_consume['id']
        # Ensure product_name is the canonical name from the products table
        product_name_canonical = product_to_consume['name']

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                # Fetch all batches of the item, ordered by expiry_date to consume oldest first.
                cursor.execute('''
                    SELECT id, product_id, name, quantity, purchase_date, expiry_date, original_quantity_string
                    FROM inventory_items 
                    WHERE product_id = ?
                    ORDER BY expiry_date ASC
                ''', (product_id_to_consume,))
                items_in_stock = [dict(row) for row in cursor.fetchall()]

                if not items_in_stock:
                    return {"success": False, "message": f"Item '{product_name_canonical}' not found in inventory."}

                for item_stock_dict in items_in_stock:
                    if quantity_remaining_to_consume <= 0:
                        break
                    
                    item_id = item_stock_dict['id']
                    current_original_qty_str = item_stock_dict['original_quantity_string']
                    current_qty_str_in_db = item_stock_dict['quantity']
                    
                    numeric_qty_in_stock = self._parse_quantity_string(current_qty_str_in_db)
                    consumable_from_this_batch = min(quantity_remaining_to_consume, numeric_qty_in_stock)

                    if consumable_from_this_batch <= 0:
                        continue

                    new_numeric_qty = numeric_qty_in_stock - consumable_from_this_batch
                    
                    new_quantity_db_str = "0" # Default if fully consumed
                    if new_numeric_qty > 0:
                        # Store floating-point numbers with decimals and integers without decimals
                        new_quantity_db_str = str(new_numeric_qty) if new_numeric_qty % 1 else str(int(new_numeric_qty))
                    
                    if new_numeric_qty <= 0:
                        cursor.execute("DELETE FROM inventory_items WHERE id = ?", (item_id,))
                        log_messages.append(f"Fully consumed batch ID {item_id} of '{product_name_canonical}'.")
                    else:
                        cursor.execute("UPDATE inventory_items SET quantity = ? WHERE id = ?", 
                                       (new_quantity_db_str, item_id))
                        log_messages.append(f"Partially consumed batch ID {item_id} of '{product_name_canonical}'. New qty: {new_quantity_db_str}")
                    
                    # Log to historical_items, now including product_id and cost_of_goods_used
                    weighted_avg_cost = self.get_weighted_average_cost(product_id_to_consume)
                    cost_of_goods_this_consumption = 0.0
                    if weighted_avg_cost is not None and weighted_avg_cost > 0: # Ensure cost is valid
                        cost_of_goods_this_consumption = consumable_from_this_batch * weighted_avg_cost

                    cursor.execute('''
                        INSERT INTO historical_items 
                        (product_id, name, quantity_consumed_this_time, original_quantity_string,
                         purchase_date, expiry_date, consumed_date, cost_of_goods_used)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (product_id_to_consume, product_name_canonical, consumable_from_this_batch, current_original_qty_str,
                          item_stock_dict['purchase_date'], item_stock_dict['expiry_date'],
                          date.today().isoformat(), cost_of_goods_this_consumption))
                    
                    consumed_amount_total_overall += consumable_from_this_batch
                    quantity_remaining_to_consume -= consumable_from_this_batch
                
                conn.commit()

        except sqlite3.Error as e:
            print(f"Database error consuming item: {e}")
            return {"success": False, "message": f"Database error: {e}", "details": log_messages}

        final_message = f"Total consumed: {consumed_amount_total_overall:.2f} of {product_name_canonical}."
        if quantity_remaining_to_consume > 0 and consumed_amount_total_overall > 0 :
            final_message += f" Could not consume the full requested amount. {quantity_remaining_to_consume:.2f} still pending."
        elif consumed_amount_total_overall == 0 and quantity_to_consume_float > 0:
             final_message = f"No quantity of '{product_name_canonical}' could be consumed (possibly none in stock or issue with parsing quantity)."
        
        print(final_message)
        return {"success": consumed_amount_total_overall > 0, "message": final_message, "details": log_messages}

    def consume_multiple_items(self, items_to_consume: list):
        """
        Consumes multiple items from the inventory.
        - items_to_consume: A list of dictionaries, each with 'item_name' and 'quantity'.
                            Example: [{'item_name': 'Apples', 'quantity': 2.0}, ...]
        Returns a list of results, one for each item consumption attempt.
        """
        overall_results = []
        if not isinstance(items_to_consume, list):
            return [{"success": False, "item_name": "N/A", "message": "Invalid input: items_to_consume must be a list."}]

        for item_spec in items_to_consume:
            item_name = item_spec.get('item_name')
            quantity_str = item_spec.get('quantity') # Assuming quantity comes as string from form/JSON

            if not item_name or quantity_str is None:
                overall_results.append({"success": False, "item_name": item_name or "Unknown", "message": "Missing item_name or quantity."})
                continue

            try:
                quantity_float = float(quantity_str)
                if quantity_float <= 0:
                    overall_results.append({"success": False, "item_name": item_name, "message": "Quantity must be a positive number."})
                    continue
            except ValueError:
                overall_results.append({"success": False, "item_name": item_name, "message": f"Invalid quantity format: {quantity_str}."})
                continue

            # Call the existing single consume_item method
            # consume_item already returns a dict like {"success": True/False, "message": "...", "details": []}
            single_item_result = self.consume_item(item_name, quantity_float)

            # Add item_name to the result for clarity if not already there or to ensure it's present
            single_item_result['item_name'] = item_name
            overall_results.append(single_item_result)

        return overall_results

    def adjust_inventory_batch(self, batch_id, new_quantity_str, new_purchase_date_str=None, new_expiry_date_str=None, include_in_projections=False):
        """
        Adjusts the quantity, purchase date, or expiry date of a specific inventory batch.
        - batch_id: The ID of the inventory_item to adjust.
        - new_quantity_str: The new quantity for the batch (as a string).
        - new_purchase_date_str: Optional. New purchase date in 'YYYY-MM-DD' format.
        - new_expiry_date_str: Optional. New expiry date in 'YYYY-MM-DD' format.
        - include_in_projections: Boolean. If True, quantity changes affect historical consumption for projections.
        Logs changes to historical_items if quantity is changed AND include_in_projections is True.
        Deletes batch if new quantity is 0.
        Returns a dictionary with success status and message.
        """
        log_message_parts = []
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()

                cursor.execute("SELECT * FROM inventory_items WHERE id = ?", (batch_id,))
                batch_row = cursor.fetchone()

                if not batch_row:
                    return {"success": False, "message": f"Inventory batch with ID {batch_id} not found."}

                current_batch = dict(batch_row)
                item_name = current_batch['name'] # Use name from DB for messages
                product_id = current_batch['product_id']
                current_quantity_numeric = self._parse_quantity_string(current_batch['quantity'])

                new_quantity_float = self._parse_quantity_string(new_quantity_str)
                if new_quantity_float < 0:
                    return {"success": False, "message": "Quantity cannot be negative."}

                quantity_diff = new_quantity_float - current_quantity_numeric
                final_quantity_db_str = str(new_quantity_str) # What will be stored if not deleting

                # Date handling (validate and prepare for update if provided)
                final_purchase_date_str = current_batch['purchase_date']
                final_expiry_date_str = current_batch['expiry_date']
                date_fields_to_update = {}

                if new_purchase_date_str and new_purchase_date_str != current_batch['purchase_date']:
                    try:
                        date.fromisoformat(new_purchase_date_str) # Validate
                        date_fields_to_update["purchase_date"] = new_purchase_date_str
                        final_purchase_date_str = new_purchase_date_str
                    except ValueError:
                        return {"success": False, "message": "Invalid new purchase date format. Use YYYY-MM-DD."}

                if new_expiry_date_str and new_expiry_date_str != current_batch['expiry_date']:
                    try:
                        date.fromisoformat(new_expiry_date_str) # Validate
                        date_fields_to_update["expiry_date"] = new_expiry_date_str
                        final_expiry_date_str = new_expiry_date_str
                    except ValueError:
                        return {"success": False, "message": "Invalid new expiry date format. Use YYYY-MM-DD."}

                # --- Database Operations ---
                if new_quantity_float == 0:
                    cursor.execute("DELETE FROM inventory_items WHERE id = ?", (batch_id,))
                    log_message_parts.append(f"Batch ID {batch_id} ('{item_name}') deleted (quantity set to 0).")
                    # If deleted, quantity_diff is based on current_quantity_numeric becoming 0.
                    # So quantity_diff will be -current_quantity_numeric
                    # This ensures the full amount is logged as "consumed" or "removed" if projections are included.
                    quantity_diff = -current_quantity_numeric
                else:
                    update_fields_sql = ["quantity = ?"]
                    update_values_sql = [final_quantity_db_str]
                    if "purchase_date" in date_fields_to_update:
                        update_fields_sql.append("purchase_date = ?")
                        update_values_sql.append(date_fields_to_update["purchase_date"])
                    if "expiry_date" in date_fields_to_update:
                        update_fields_sql.append("expiry_date = ?")
                        update_values_sql.append(date_fields_to_update["expiry_date"])

                    if update_fields_sql: # Only update if quantity or dates actually changed
                        update_values_sql.append(batch_id)
                        cursor.execute(f"UPDATE inventory_items SET {', '.join(update_fields_sql)} WHERE id = ?", tuple(update_values_sql))
                        log_message_parts.append(f"Batch ID {batch_id} ('{item_name}') updated.")
                    else: # No quantity change, no date change
                         if quantity_diff == 0 and not date_fields_to_update:
                             return {"success": True, "message": f"No changes detected for batch ID {batch_id} ('{item_name}')."}


                # Historical Logging based on include_in_projections flag
                if quantity_diff != 0: # Only log if quantity actually changed
                    if include_in_projections:
                        amount_for_history = -quantity_diff # Positive if decreased, negative if increased
                        cursor.execute('''
                            INSERT INTO historical_items
                            (product_id, name, quantity_consumed_this_time, original_quantity_string,
                             purchase_date, expiry_date, consumed_date)
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                        ''', (product_id, item_name, amount_for_history, current_batch['original_quantity_string'],
                              current_batch['purchase_date'], current_batch['expiry_date'], # Log original batch dates
                              date.today().isoformat()))
                        log_message_parts.append(f"Adjustment of {amount_for_history} units recorded for projections. Note: Batch ID {batch_id} quantity adjusted from {current_quantity_numeric} to {new_quantity_float}.")
                    else:
                        log_message_parts.append(f"Quantity updated without impacting projections history. Note: Batch ID {batch_id} quantity adjusted from {current_quantity_numeric} to {new_quantity_float}.")

                conn.commit()
                return {"success": True, "message": " ".join(log_message_parts)}

        except sqlite3.Error as e:
            print(f"Database error adjusting inventory batch ID {batch_id}: {e}")
            return {"success": False, "message": f"Database error: {e}"}
        except Exception as e: # Catch other potential errors like date parsing
            print(f"Error adjusting inventory batch ID {batch_id}: {e}")
            return {"success": False, "message": f"An unexpected error occurred: {e}"}

    def check_for_expiring_items(self, days_threshold=3):
        """Checks for items expiring within a certain number of days from DB."""
        today = date.today()
        threshold_date = (today + timedelta(days=days_threshold)).isoformat()
        expiring_items_list = []
        
        print(f"\n--- Items Expiring Soon (within {days_threshold} days or already expired) ---")
        found_expiring = False
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                # Select items expiring on or before the threshold date, or already expired (expiry_date < today)
                cursor.execute('''
                    SELECT name, quantity, purchase_date, expiry_date 
                    FROM inventory_items 
                    WHERE expiry_date <= ? 
                    ORDER BY expiry_date ASC
                ''', (threshold_date,))
                rows = cursor.fetchall()
                for row in rows:
                    item = dict(row)
                    item_expiry_date = date.fromisoformat(item['expiry_date'])
                    days_to_expiry = (item_expiry_date - today).days
                    
                    if days_to_expiry < 0:
                        print(f"- {item['name']} EXPIRED on {item['expiry_date']}!")
                    else:
                        print(f"- {item['name']} expires in {days_to_expiry} day(s) on {item['expiry_date']}")
                    
                    # Convert dates for the returned list
                    item['purchase_date'] = date.fromisoformat(item['purchase_date'])
                    item['expiry_date'] = item_expiry_date
                    expiring_items_list.append(item)
                    found_expiring = True
            if not found_expiring:
                print("No items expiring soon or already expired.")
        except sqlite3.Error as e:
            print(f"Database error checking expiring items: {e}")
        
        print("---------------------------------------\n")
        return expiring_items_list

    def get_inventory_concerns(self, product_id):
        """
        Analyzes a product's inventory and consumption to identify potential concerns.
        - product_id: The ID of the product to check.
        Returns a list of concern strings.
        """
        concerns = []
        today = date.today()

        product_details = self.get_product_details(product_id) # This now includes nearest_expiry_date
        if not product_details:
            concerns.append(f"Product with ID {product_id} not found.")
            return concerns

        current_stock = product_details.get('current_on_hand_inventory', 0)
        nearest_expiry_date_str = product_details.get('nearest_expiry_date', "N/A")

        # Get average daily consumption
        # Using projection_days=1 as we only need avg_daily_consumption here.
        # lookback_days can be standard e.g. 30 or configurable if needed.
        demand_projection = self.project_demand(product_id, lookback_days=30, projection_days=1)
        avg_daily_consumption = 0.0
        if demand_projection.get("success"):
            avg_daily_consumption = demand_projection.get('avg_daily_consumption', 0.0)
        else:
            # Add a concern if demand projection failed, as it's needed for some checks
            concerns.append(f"Could not retrieve consumption data for {product_details.get('name', 'this product')}: {demand_projection.get('message', 'Unknown error')}")


        # Concern 1: Low stock
        if avg_daily_consumption > 0:
            days_of_stock_left = current_stock / avg_daily_consumption
            # Using a threshold of 3 days for "low stock"
            if days_of_stock_left < 3:
                concerns.append(f"Will run out in the next {days_of_stock_left:.1f} days based on current consumption.")
        elif current_stock > 0: # Stock exists but no consumption
            # Concern 2: No significant usage
            concerns.append("No significant usage data available, but stock exists.")
        # If avg_daily_consumption is 0 and current_stock is 0, it's less of an immediate "concern" unless par levels are set.

        # Date parsing for expiry concerns
        nearest_expiry_date_obj = None
        if nearest_expiry_date_str != "N/A":
            try:
                nearest_expiry_date_obj = date.fromisoformat(nearest_expiry_date_str)
            except ValueError:
                concerns.append(f"Invalid nearest expiry date format found: {nearest_expiry_date_str}")


        if nearest_expiry_date_obj:
            days_until_expiry = (nearest_expiry_date_obj - today).days

            # Concern 3: Nearing expiry (e.g., within 7 days)
            # Let's use a threshold of 7 days for "expiring soon"
            EXPIRY_SOON_THRESHOLD_DAYS = 7
            if days_until_expiry <= EXPIRY_SOON_THRESHOLD_DAYS:
                if days_until_expiry < 0:
                     concerns.append(f"Item has batches already EXPIRED (Nearest expiry: {nearest_expiry_date_str}).")
                else:
                     concerns.append(f"Item has batches expiring soon (on {nearest_expiry_date_str}, in {days_until_expiry} days).")

            # Concern 4: Stock may expire before being fully used
            if avg_daily_consumption > 0 and current_stock > 0:
                # days_of_stock_left was calculated earlier for low stock check
                if days_of_stock_left > days_until_expiry and days_until_expiry >= 0: # only if not already expired
                    concerns.append(f"Stock ({current_stock} units, lasting {days_of_stock_left:.1f} days) may expire before being fully used (expires in {days_until_expiry} days on {nearest_expiry_date_str}).")
        elif current_stock > 0 and nearest_expiry_date_str == "N/A":
            # This case implies stock exists but no expiry date could be found, which might be a data issue.
            concerns.append("Stock exists but nearest expiry date is not available. Check inventory data.")

        return concerns

    def _get_average_daily_consumption(self, product_id, lookback_days=30):
        """
        Calculates or retrieves the average daily consumption for a product.
        1. Checks for consumption_override_rate.
        2. If not available, calculates from historical_items.
        """
        product = self.get_product(product_id)
        if not product:
            # This case should ideally be handled by the caller, but good to have a fallback.
            print(f"Warning: Product with ID {product_id} not found in _get_average_daily_consumption.")
            return 0.0

        if product.get('consumption_override_rate') is not None:
            try:
                override_rate = float(product['consumption_override_rate'])
                print(f"Using consumption_override_rate: {override_rate} for product ID {product_id}")
                return override_rate
            except ValueError:
                print(f"Warning: Could not parse consumption_override_rate '{product['consumption_override_rate']}' for product ID {product_id}. Calculating from history.")

        # Calculate from historical data
        total_consumed_in_lookback = 0.0
        today = date.today()
        lookback_start_dt = today - timedelta(days=lookback_days)

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT SUM(quantity_consumed_this_time) as total_consumed
                    FROM historical_items
                    WHERE product_id = ? AND consumed_date >= ? AND consumed_date < ?
                ''', (product_id, lookback_start_dt.isoformat(), today.isoformat())) # Use < today to not include today's consumption
                result_row = cursor.fetchone()
                if result_row and result_row['total_consumed'] is not None:
                    total_consumed_in_lookback = float(result_row['total_consumed'])

            if lookback_days > 0:
                return total_consumed_in_lookback / lookback_days
            return 0.0
        except sqlite3.Error as e:
            print(f"Database error calculating historical consumption for product ID {product_id}: {e}")
            return 0.0 # Fallback to 0 on error

    def project_demand(self, product_name_or_id, lookback_days=30, projection_days=7):
        """
        Analyzes historical consumption and current stock to project future demand using DB.
        - Calculates average daily consumption based on historical data within a lookback period.
        - Checks current total stock of the item.
        - Estimates how long current stock will last and projects future need.
        """
        today = date.today()
        lookback_start_dt = today - timedelta(days=lookback_days)
        total_consumed_in_lookback = 0.0
        product_id = None
        product_name = None
        product_unit = "units" # Default unit

        if isinstance(product_name_or_id, int): # Assume it's an ID
            product_id = product_name_or_id
            product = self.get_product(product_id)
            if not product:
                return {"success": False, "message": f"Product with ID {product_id} not found."}
            product_name = product['name']
            product_unit = product.get('unit_of_measure', product_unit)
        elif isinstance(product_name_or_id, str): # Assume it's a name
            product = self.get_product_by_name(product_name_or_id)
            if not product:
                return {"success": False, "message": f"Product '{product_name_or_id}' not found."}
            product_id = product['id']
            product_name = product['name']
            product_unit = product.get('unit_of_measure', product_unit)
        else:
            return {"success": False, "message": "Invalid product identifier for projection."}

        # Utilize the new helper method for average daily consumption
        avg_daily_consumption = self._get_average_daily_consumption(product_id, lookback_days)
        # total_consumed_in_lookback is not directly available from the helper,
        # but avg_daily_consumption is the key metric.
        # For display purposes, if needed, it could be avg_daily_consumption * lookback_days,
        # but this assumes the lookback period for the helper matches.

        current_quantity_sum = self.get_total_item_quantity(product_id)
        # Calculation of total_consumed_in_lookback for print output needs adjustment if we rely solely on _get_average_daily_consumption
        # For now, let's assume the print output can be simplified or the value can be derived.
        # To maintain the original print output accurately, project_demand would still need to perform
        # its own historical sum query or _get_average_daily_consumption would need to return more data.
        # For this refactoring, we prioritize using the helper for avg_daily_consumption calculation.
        # The print statement below will use avg_daily_consumption * lookback_days as an estimate.
        total_consumed_in_lookback = avg_daily_consumption * lookback_days # Estimate for print
        
        days_to_depletion_str = "N/A"
        if avg_daily_consumption > 0:
            if current_quantity_sum > 0 :
                days_to_depletion = current_quantity_sum / avg_daily_consumption
                days_to_depletion_str = f"{days_to_depletion:.1f} days"
            else:
                days_to_depletion_str = "0 days (already out of stock)"
        elif current_quantity_sum > 0:
            days_to_depletion_str = "Stock will not deplete based on recent consumption."
        else:
            days_to_depletion_str = "N/A (out of stock, no consumption history)"
        
        projected_need = avg_daily_consumption * projection_days

        # Calculate cost over the last 30 days
        cost_last_30_days = 0.0
        thirty_days_ago_dt = today - timedelta(days=30)
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT SUM(cost_of_goods_used) as total_cost
                    FROM historical_items
                    WHERE product_id = ? AND consumed_date >= ? AND consumed_date < ?
                ''', (product_id, thirty_days_ago_dt.isoformat(), today.isoformat()))
                cost_row = cursor.fetchone()
                if cost_row and cost_row['total_cost'] is not None:
                    cost_last_30_days = float(cost_row['total_cost'])
        except sqlite3.Error as e:
            print(f"Database error calculating cost_last_30_days for product ID {product_id}: {e}")
            # Keep cost_last_30_days as 0.0 on error

        # Get weighted average cost for projected cost calculation
        weighted_average_cost = self.get_weighted_average_cost(product_id)
        projected_cost_per_day = 0.0
        if weighted_average_cost is not None: # Ensure it's not None before multiplication
            projected_cost_per_day = avg_daily_consumption * weighted_average_cost
        else: # Handle case where cost might be None (e.g. no purchase history)
            weighted_average_cost = 0.0 # Default to 0 for display if None

        result = {
            "product_id": product_id, "product_name": product_name, "unit_of_measure": product_unit,
            "current_stock": current_quantity_sum,
            "avg_daily_consumption": avg_daily_consumption, "days_to_depletion": days_to_depletion_str,
            "projected_need": projected_need, "lookback_days": lookback_days, "projection_days": projection_days,
            "consumption_override_rate": product.get('consumption_override_rate'), # Ensure it's in the result
            "cost_last_30_days": cost_last_30_days,
            "weighted_average_cost": weighted_average_cost, # Added for reference
            "projected_cost_per_day": projected_cost_per_day, # New projected cost field
            "success": True
        }

        print(f"\n--- Demand Projection for '{product_name}' (ID: {product_id}) ---")
        print(f"Unit of Measure: {product_unit}")
        print(f"Lookback: {lookback_days} days, Projection: {projection_days} days")
        # Note: total_consumed_in_lookback is now an estimate if using the helper that only returns avg.
        print(f"Total consumed (lookback, estimated): {total_consumed_in_lookback:.2f} {product_unit}")
        print(f"Current stock: {current_quantity_sum:.2f} {product_unit}")
        print(f"Avg daily consumption: {avg_daily_consumption:.2f} {product_unit}/day")
        print(f"Est. days to depletion: {days_to_depletion_str}")
        print(f"Projected need (next {projection_days} days): {projected_need:.2f} {product_unit}")
        print("-----------------------------------------\n")
        return result

    def get_future_inventory_projection(self, product_id, projection_days):
        """
        Projects future inventory levels for a product on a daily basis.

        Method Signature: get_future_inventory_projection(self, product_id, projection_days)

        Initialization:
            Fetch current inventory_items for the product_id, sorted by expiry_date.
            Fetch all production_items (garden harvests) for the product_id that are relevant.
            Get avg_daily_consumption:
                Use _get_average_daily_consumption helper.

        Daily Iteration Loop (for d from 0 to projection_days - 1):
            current_projection_date = date.today() + timedelta(days=d)
            projected_inventory_today = previous_day_ending_inventory (or actual current total stock if d == 0).
            daily_harvest = 0, daily_shrink = 0, daily_consumption = 0.

            Harvests:
                Iterate through relevant production_items.
                If current_projection_date is within a production_item's harvest period,
                add its estimated_daily_yield to projected_inventory_today and daily_harvest.

            Spoilage (Shrink) & Consumption:
                Identify quantity expiring today (expiring_qty_today) from a copy of inventory_items.
                inventory_available_for_consumption = projected_inventory_today (which includes expiring_qty_today).
                consumed_amount_for_day = min(inventory_available_for_consumption, avg_daily_consumption).
                projected_inventory_today -= consumed_amount_for_day.
                daily_consumption = consumed_amount_for_day.
                amount_of_expiring_items_consumed = min(expiring_qty_today, consumed_amount_for_day).
                shrink_for_the_day = expiring_qty_today - amount_of_expiring_items_consumed.
                projected_inventory_today -= shrink_for_the_day.
                daily_shrink = shrink_for_the_day.
                Update remaining quantities in inventory batches or remove them if fully depleted.
                (This part is simplified: we operate on total projected_inventory_today for now,
                 actual batch depletion for *next day's* spoilage is tricky and not fully modeled here yet).

            Record: Store daily data.
            If projected_inventory_today <= 0 and no depletion date recorded, mark this date.

        Return Value: List of daily record dictionaries.
        """
        projection_results = []
        today = date.today()

        # Fetch product details (needed for name, unit for records)
        product = self.get_product(product_id)
        if not product:
            # Optionally, return an error structure or raise an exception
            return {"success": False, "message": f"Product with ID {product_id} not found."}
        product_name = product.get('name', f"Product {product_id}")
        # unit_of_measure = product.get('unit_of_measure', 'units') # Not directly used in loop logic, but good for records

        # Fetch current inventory batches, sorted by expiry date
        # These are dictionaries with date objects if get_inventory_batches_for_product is used
        inventory_batches = self.get_inventory_batches_for_product(product_id, order_by_id_desc=False) # FEFO

        # Create a mutable copy for daily simulation of batch depletion (quantity update or removal)
        # Each item in simulated_batches should store its current quantity for the simulation
        simulated_batches = []
        for batch in inventory_batches:
            try:
                # Ensure quantity is float, expiry_date is date object
                simulated_batches.append({
                    'id': batch['id'],
                    'expiry_date': batch['expiry_date'] if isinstance(batch['expiry_date'], date) else date.fromisoformat(batch['expiry_date']),
                    'current_quantity': self._parse_quantity_string(batch['quantity'])
                })
            except (ValueError, TypeError) as e:
                print(f"Warning: Skipping batch {batch.get('id')} due to invalid data: {e}")
                continue

        # Sort again just in case (get_inventory_batches_for_product should already sort by expiry)
        simulated_batches.sort(key=lambda b: b['expiry_date'])


        # Fetch relevant production_items (garden harvests)
        all_production_items = self.get_all_production_items() # Gets all, then filter
        relevant_production_items = []
        for item in all_production_items:
            if item.get('associated_product_id') == product_id:
                # Check if relevant (status or harvest period overlap)
                # This logic can be refined based on how 'status' interacts with date calculations
                try:
                    plant_dt = date.fromisoformat(item['plant_date'])
                    time_to_harvest = item.get('time_to_harvest_days', 0)
                    period_days = item.get('expected_harvest_period_days', 0)

                    item_harvest_start_date = plant_dt + timedelta(days=time_to_harvest)
                    item_harvest_end_date = item_harvest_start_date + timedelta(days=period_days)

                    # Relevant if its harvest period overlaps with the projection window
                    # Or if status is Growing/Harvesting (more complex to perfectly align with dates)
                    # Simplified: if harvest can occur during projection
                    projection_end_date = today + timedelta(days=projection_days)
                    if not (item_harvest_end_date < today or item_harvest_start_date > projection_end_date):
                         # Ensure estimated_daily_yield is calculated and valid
                        if item.get('estimated_daily_yield') == 'Error' or item.get('estimated_daily_yield') is None:
                            daily_yield = 0
                            if period_days > 0 and item.get('expected_yield_total') is not None:
                                daily_yield = item['expected_yield_total'] / period_days
                            else: # Default to 0 if cannot calculate
                                daily_yield = 0
                            item['estimated_daily_yield'] = daily_yield # Store calculated value

                        relevant_production_items.append({
                            'plant_date': plant_dt,
                            'time_to_harvest_days': time_to_harvest,
                            'expected_harvest_period_days': period_days,
                            'estimated_daily_yield': item['estimated_daily_yield'],
                            'harvest_start_date': item_harvest_start_date, # Store calculated dates
                            'harvest_end_date': item_harvest_end_date
                        })
                except (ValueError, TypeError, KeyError) as e:
                    print(f"Warning: Could not process production item {item.get('id')} for projection: {e}")
                    continue


        # Get average daily consumption
        avg_daily_consumption = self._get_average_daily_consumption(product_id, lookback_days=30)

        # Initial inventory state
        current_total_stock = sum(b['current_quantity'] for b in simulated_batches)
        previous_day_ending_inventory = current_total_stock
        depletion_date_recorded = False

        for d in range(projection_days):
            current_projection_date = today + timedelta(days=d)

            opening_inventory_today = previous_day_ending_inventory
            projected_inventory_today = opening_inventory_today # Start with previous day's end

            daily_harvest = 0.0
            daily_shrink = 0.0
            daily_consumption = 0.0

            # 1. Add Harvests for the day
            for prod_item in relevant_production_items:
                if prod_item['harvest_start_date'] <= current_projection_date <= prod_item['harvest_end_date']:
                    harvest_this_day = float(prod_item.get('estimated_daily_yield', 0))
                    daily_harvest += harvest_this_day
                    projected_inventory_today += harvest_this_day

            # Inventory after harvest, before consumption and spoilage
            inventory_before_consumption_spoilage = projected_inventory_today

            # 2. Determine potential spoilage for *today* from existing batches
            expiring_qty_today = 0.0
            # Iterate over a copy of simulated_batches for inspection, actual modification happens later
            for batch in list(simulated_batches): # Iterate copy for safe removal
                if batch['expiry_date'] == current_projection_date:
                    expiring_qty_today += batch['current_quantity']

            # 3. Account for Consumption for the day
            # Consumption happens from the total available pool (including items that might expire today)
            consumed_amount_for_day = min(projected_inventory_today, avg_daily_consumption)
            daily_consumption = consumed_amount_for_day
            projected_inventory_today -= consumed_amount_for_day # Reduce total inventory

            # 4. Attribute consumption to batches (FEFO) and calculate shrink
            # This section determines how much of the expiring quantity was consumed vs. spoiled

            # Of the quantity that was set to expire today, how much was part of what got consumed?
            amount_of_expiring_items_consumed_today = min(expiring_qty_today, consumed_amount_for_day)

            # Shrink is the portion of "expiring_qty_today" that was NOT consumed
            shrink_for_the_day = expiring_qty_today - amount_of_expiring_items_consumed_today
            daily_shrink = shrink_for_the_day

            # The projected_inventory_today already had consumed_amount_for_day removed.
            # Now, subtract the part of shrink_for_the_day that effectively removes inventory
            # that wasn't consumed and has now spoiled.
            # If consumed_amount_for_day >= expiring_qty_today, all expiring items were consumed, so shrink is 0.
            # If consumed_amount_for_day < expiring_qty_today, then (expiring_qty_today - consumed_amount_for_day) spoiled.
            # This amount needs to be subtracted from inventory if it wasn't already effectively removed
            # by the consumption logic.

            # Let's re-evaluate projected_inventory_today based on clearer steps:
            # Start of day: opening_inventory_today
            # Add harvest: opening_inventory_today + daily_harvest
            # Potential consumption: avg_daily_consumption
            # Potential spoilage: expiring_qty_today (from batches expiring *today*)

            # Logic revised:
            # projected_inventory_after_harvest = opening_inventory_today + daily_harvest
            # consumed_today = min(projected_inventory_after_harvest, avg_daily_consumption)
            # inventory_after_consumption = projected_inventory_after_harvest - consumed_today
            # daily_consumption = consumed_today

            # Now, from inventory_after_consumption, what spoils?
            # Spoilage is from items expiring *today*.
            # We need to track remaining quantities in batches to do this perfectly.

            # Simpler spoilage for now (as per plan step 7):
            # The `shrink_for_the_day` is calculated above.
            # The `projected_inventory_today` was reduced by `consumed_amount_for_day`.
            # Now, it also needs to be reduced by `shrink_for_the_day`.
            projected_inventory_today -= shrink_for_the_day # This was the crucial part.


            # 5. Update simulated_batches for the *next* day's calculation
            # This part is essential for multi-day accuracy of spoilage.
            # Deplete consumed amounts from batches (FEFO)
            temp_consumed_to_allocate = daily_consumption
            for batch in sorted(simulated_batches, key=lambda b: b['expiry_date']): # FEFO
                if temp_consumed_to_allocate <= 0: break
                amount_from_this_batch = min(batch['current_quantity'], temp_consumed_to_allocate)
                batch['current_quantity'] -= amount_from_this_batch
                temp_consumed_to_allocate -= amount_from_this_batch

            # Remove spoiled batches (those that expired today and whose remaining quantity contributed to daily_shrink)
            temp_shrink_to_allocate = daily_shrink
            for batch in list(simulated_batches): # Iterate copy for removal
                if batch['expiry_date'] == current_projection_date:
                    if temp_shrink_to_allocate <=0: break # All shrink accounted for
                    amount_spoiled_this_batch = min(batch['current_quantity'], temp_shrink_to_allocate)
                    batch['current_quantity'] -= amount_spoiled_this_batch
                    temp_shrink_to_allocate -= amount_spoiled_this_batch

            # Clean up empty batches from simulated_batches
            simulated_batches = [b for b in simulated_batches if b['current_quantity'] > 0.001] # Keep if some qty remains

            # Ensure projected_inventory_today is not negative
            projected_inventory_today = max(0, projected_inventory_today)

            # Depletion date
            depleted_this_day = False
            if not depletion_date_recorded and projected_inventory_today <= 0:
                depleted_this_day = True
                depletion_date_recorded = True # Mark that we've found the first depletion date

            projection_results.append({
                'date': current_projection_date.isoformat(),
                'product_id': product_id, # For reference
                'product_name': product_name, # For reference
                'opening_inventory': round(opening_inventory_today, 2),
                'harvest': round(daily_harvest, 2),
                'consumption': round(daily_consumption, 2),
                'shrink': round(daily_shrink, 2),
                'projected_ending_inventory': round(projected_inventory_today, 2),
                'depletion_date_reached': depleted_this_day
            })

            previous_day_ending_inventory = projected_inventory_today

        return projection_results


    def export_data_to_csv(self, filename_prefix="inventory_export_db"):
        """Exports current and historical inventory, and projections to CSV files from DB."""
        def write_to_csv_internal(filename, data_dicts, fieldnames):
            try:
                with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    for row_dict in data_dicts:
                        # Convert date objects to ISO strings for CSV if they are date objects
                        row_to_write = {k: (v.isoformat() if isinstance(v, date) else v) for k, v in row_dict.items()}
                        writer.writerow(row_to_write)
                print(f"Data successfully exported to '{filename}'.")
            except IOError as e: print(f"Error exporting data to '{filename}': {e}")
            except Exception as e: print(f"An unexpected error occurred during CSV export: {e}")

        current_inv_data = self.get_current_inventory() # This method needs to be checked if it returns full item details or aggregated
        # For current_fields, if get_current_inventory() is aggregated, we might need a different source for raw batch data.
        # Assuming get_inventory_batches_for_product(product_id) would be used if exporting specific product batches.
        # The original current_fields were ["id", "name", "quantity", "purchase_date", "expiry_date", "original_quantity_string"]
        # which implies batch-level data. If get_current_inventory provides aggregated data, this export part might need adjustment
        # or a different method call to get all batches of all products.
        # For now, let's assume get_current_inventory() can provide some form of list of dicts suitable for export.
        # A more robust export might iterate all products, then all batches for each.

        # Let's use get_all_products() and then get_inventory_batches_for_product() for a more complete current inventory export.
        all_prods_for_export = self.get_all_products()
        all_current_batches_export = []
        for prod in all_prods_for_export:
            batches = self.get_inventory_batches_for_product(prod['id'])
            all_current_batches_export.extend(batches)

        if all_current_batches_export:
            current_fields = ["id", "product_id", "product_name", "quantity", "purchase_date", "expiry_date", "original_quantity_string", "unit_of_measure"]
            # Filter data to ensure only existing keys are written (safer for DictWriter)
            filtered_current_export = []
            for batch_dict in all_current_batches_export:
                filtered_dict = {k: batch_dict.get(k) for k in current_fields if k in batch_dict}
                filtered_current_export.append(filtered_dict)
            write_to_csv_internal(f"{filename_prefix}_current_batches.csv", filtered_current_export, current_fields)
        else: print("Current inventory (batches) empty. Skipping export.")


        historical_inv_data = self.get_historical_inventory() # This should be fine.
        if historical_inv_data:
            hist_fields = ["id", "product_id", "product_display_name", "quantity_consumed_this_time", "original_quantity_string",
                           "purchase_date", "expiry_date", "consumed_date", "unit_of_measure", "category_name", "subcategory_name"]
            # Filter data for safety
            filtered_hist_export = []
            for hist_dict in historical_inv_data:
                # Use product_display_name as 'name' for consistency if needed by CSV consumer
                hist_dict['name'] = hist_dict.get('product_display_name', hist_dict.get('name'))
                filtered_dict = {k: hist_dict.get(k) for k in hist_fields if k in hist_dict}
                filtered_hist_export.append(filtered_dict)

            write_to_csv_internal(f"{filename_prefix}_historical.csv", filtered_hist_export, hist_fields)
        else: print("Historical inventory empty. Skipping export.")

    # --- Purchase Logging and Costing Methods ---
    def log_purchase(self, product_id, purchase_date_str, quantity_purchased_float, cost_per_unit_float, vendor_str=None):
        """Logs a purchase into the PurchaseLog and updates inventory stock."""
        if not all([product_id, purchase_date_str, quantity_purchased_float is not None, cost_per_unit_float is not None]):
            return {"success": False, "message": "Missing required fields for logging purchase (product_id, purchase_date, quantity, cost_per_unit)."}
        if quantity_purchased_float <= 0:
            return {"success": False, "message": "Quantity purchased must be positive."}
        if cost_per_unit_float < 0: # Allow 0 cost, but not negative
            return {"success": False, "message": "Cost per unit cannot be negative."}

        try:
            # Validate purchase_date_str format
            date.fromisoformat(purchase_date_str)
        except ValueError:
            return {"success": False, "message": "Invalid purchase_date format. Use YYYY-MM-DD."}

        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT INTO PurchaseLog
                    (product_id, purchase_date, quantity_purchased, cost_per_unit, vendor)
                    VALUES (?, ?, ?, ?, ?)
                ''', (product_id, purchase_date_str, quantity_purchased_float, cost_per_unit_float, vendor_str))
                purchase_log_id = cursor.lastrowid
                conn.commit() # Commit PurchaseLog entry first

                # Now add to inventory stock
                # The quantity_str for add_inventory_stock should be the total quantity purchased.
                # add_inventory_stock will use the product's default_expiry_days.
                stock_result = self.add_inventory_stock(
                    product_id=product_id,
                    quantity_str=str(quantity_purchased_float), # add_inventory_stock expects a string
                    purchase_date_str=purchase_date_str
                )

                if stock_result.get("success"):
                    return {
                        "success": True,
                        "message": f"Purchase logged (ID: {purchase_log_id}) and stock added successfully for product ID {product_id}.",
                        "purchase_log_id": purchase_log_id,
                        "stock_item_id": stock_result.get("stock_item_id")
                    }
                else:
                    # Attempt to rollback PurchaseLog entry if stock addition failed?
                    # For simplicity, we'll report the error. Manual correction might be needed if only stock fails.
                    # A more robust solution would use transactions spanning both operations.
                    return {
                        "success": False,
                        "message": f"Purchase logged (ID: {purchase_log_id}), but failed to add stock: {stock_result.get('message')}",
                        "purchase_log_id": purchase_log_id
                    }
        except sqlite3.IntegrityError as e: # e.g. product_id doesn't exist
            return {"success": False, "message": f"Database integrity error logging purchase: {e}. Ensure product ID exists."}
        except sqlite3.Error as e:
            return {"success": False, "message": f"Database error logging purchase: {e}"}

    def get_weighted_average_cost(self, product_id):
        """Calculates the weighted average cost for a product from PurchaseLog."""
        try:
            with self._get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT SUM(quantity_purchased * cost_per_unit) as total_cost,
                           SUM(quantity_purchased) as total_quantity
                    FROM PurchaseLog
                    WHERE product_id = ?
                ''', (product_id,))
                row = cursor.fetchone()

                if row and row['total_quantity'] is not None and row['total_quantity'] > 0:
                    return float(row['total_cost']) / float(row['total_quantity'])
                else:
                    # No purchases logged or total quantity is zero
                    return 0.0 # Or None, depending on how you want to handle "no cost data"
        except sqlite3.Error as e:
            print(f"Database error calculating weighted average cost for product ID {product_id}: {e}")
            return 0.0 # Or raise error
        except TypeError: # Handles if row is None or keys are missing unexpectedly
            print(f"Type error calculating weighted average cost for product ID {product_id}. Likely no purchase data.")
            return 0.0

    def log_multiple_purchases(self, purchases_data_list):
        """
        Logs multiple purchases from a list of purchase data.
        Each item in purchases_data_list is a dict:
        {
            "product_id": int,
            "purchase_date_str": "YYYY-MM-DD",
            "quantity_purchased_float": float,
            "cost_per_unit_float": float,
            "vendor_str": str_or_none
        }
        Returns a summary of successes and failures.
        """
        success_count = 0
        failure_count = 0
        results_details = [] # List of dicts: {"product_id": X, "success": True/False, "message": "..."}

        for purchase_data in purchases_data_list:
            product_id = purchase_data.get("product_id")
            purchase_date_str = purchase_data.get("purchase_date_str")
            quantity_purchased_float = purchase_data.get("quantity_purchased_float")
            cost_per_unit_float = purchase_data.get("cost_per_unit_float")
            vendor_str = purchase_data.get("vendor_str")

            # Basic check for required fields in each item
            if not all([product_id is not None, purchase_date_str,
                        quantity_purchased_float is not None, cost_per_unit_float is not None]):
                failure_count += 1
                results_details.append({
                    "product_id": product_id, # Might be None if missing
                    "success": False,
                    "message": "Missing essential data for purchase."
                })
                continue

            # Call the existing log_purchase method for each item
            result = self.log_purchase(
                product_id=product_id,
                purchase_date_str=purchase_date_str,
                quantity_purchased_float=quantity_purchased_float,
                cost_per_unit_float=cost_per_unit_float,
                vendor_str=vendor_str
            )

            if result.get("success"):
                success_count += 1
            else:
                failure_count += 1

            results_details.append({
                "product_id": product_id,
                "success": result.get("success"),
                "message": result.get("message", "Unknown error during individual purchase logging.")
            })

        return {
            "overall_success": failure_count == 0, # True if all items succeeded
            "success_count": success_count,
            "failure_count": failure_count,
            "results_details": results_details
        }


        # Projections export:
        # The original code projected demand for items in current/historical.
        # For future inventory projection, this part might change or be an additional export.
        # For now, keeping the demand projection export as it was.
        unique_product_ids_for_demand_projection = set()
        if all_current_batches_export: # Use product_ids from current batches
            for item in all_current_batches_export: unique_product_ids_for_demand_projection.add(item['product_id'])
        if historical_inv_data: # Use product_ids from historical items
            # Ensure product_id exists in hist_dict, it might be None if historical item wasn't linked
            for hist_dict in historical_inv_data:
                if hist_dict.get('product_id'):
                    unique_product_ids_for_demand_projection.add(hist_dict['product_id'])

        if unique_product_ids_for_demand_projection:
            demand_projections = [self.project_demand(pid) for pid in sorted(list(unique_product_ids_for_demand_projection))]
            successful_demand_projections = [p for p in demand_projections if p.get("success", True)] # Original logic
            if successful_demand_projections:
                # Make sure all dicts have the same keys for CSV. Get all possible keys from the successful projections.
                all_proj_keys = set()
                for p in successful_demand_projections: all_proj_keys.update(p.keys())
                proj_fields = sorted([key for key in list(all_proj_keys) if key != 'success'])

                write_to_csv_internal(f"{filename_prefix}_demand_projections.csv", successful_demand_projections, proj_fields)
            else:
                print("No successful demand projection data generated. Skipping export.")
        else:
            print("No items found to generate demand projections. Skipping demand projection export.")

        # Example for exporting future inventory projection for a specific product (can be expanded)
        # This is just an example, you'd likely iterate or select products for this export.
        # For demonstration, let's pick the first product_id if available.
        if unique_product_ids_for_demand_projection:
            # Ensure the set is not empty before trying to get an iterator
            if unique_product_ids_for_demand_projection:
                first_product_id_for_future_proj = next(iter(unique_product_ids_for_demand_projection), None)
                if first_product_id_for_future_proj:
                    future_proj_data = self.get_future_inventory_projection(first_product_id_for_future_proj, 30) # Project 30 days
                    if future_proj_data and isinstance(future_proj_data, list) and len(future_proj_data) > 0 : # Check if it's a list of records
                        future_proj_fields = future_proj_data[0].keys() # Assuming all dicts have same keys
                        write_to_csv_internal(f"{filename_prefix}_future_inventory_proj_prod_{first_product_id_for_future_proj}.csv", future_proj_data, list(future_proj_fields))
                    elif isinstance(future_proj_data, dict) and not future_proj_data.get("success", True): # Handle error dict
                        print(f"Could not generate future inventory projection for product {first_product_id_for_future_proj}: {future_proj_data.get('message')}")
                    else:
                        print(f"No future inventory projection data generated for product {first_product_id_for_future_proj}.")
            else:
                print("No product IDs available to generate future inventory projection example.")


# --- Standalone Garden Produce Functions (Not part of InventoryManager class) ---
# These functions remain as they were, operating on a global list,
# as they are not part of the core SQLite-backed InventoryManager.
my_garden_produce_list = [] 

def add_garden_produce(name, harvest_date_str, typical_shelf_life_days):
    """Adds produce harvested from the garden to a global list."""
    global my_garden_produce_list
    try:
        harvest_dt = date.fromisoformat(harvest_date_str)
        produce = {
            "name": name,
            "harvest_date": harvest_dt,
            "estimated_expiry": harvest_dt + timedelta(days=typical_shelf_life_days),
            "source": "garden"
        }
        my_garden_produce_list.append(produce)
        print(f"Logged garden produce: {produce['name']}")
    except ValueError:
        print(f"Error: Invalid date format for garden produce '{name}'. Please use YYYY-MM-DD.")

def display_garden_produce():
    """Prints all items in the garden produce list."""
    if not my_garden_produce_list:
        print("Your garden produce list is empty!")
        return
    print("\n--- Your Garden Produce ---")
    for item in my_garden_produce_list:
        print(f"- {item['name']}, Harvested: {item['harvest_date'].isoformat()}, Est. Expires: {item['estimated_expiry'].isoformat()}")
    print("---------------------------\n")


if __name__ == "__main__":
    DB_FILE = "food_manager_dev.db" # Use a dev-specific DB for demo
    print(f"Welcome to your SQLite Food Manager! Using database: {DB_FILE}")
    print("Note: For a fresh demo, delete the database file before running.")
    
    manager = InventoryManager(db_filepath=DB_FILE)

    # --- Section 1: Initial State ---
    print("\n--- Section 1: Initial Inventory State ---")
    current_items_main = manager.get_current_inventory()
    if not current_items_main: print("Current inventory is empty.")
    else: 
        for item in current_items_main: print(f"- {item['name']} ({item['quantity']})")
    
    historical_items_main = manager.get_historical_inventory()
    if not historical_items_main: print("Historical inventory is empty.")
    else: 
        for item in historical_items_main: print(f"- HIST: {item['name']} ({item['quantity_consumed_this_time']}) on {item['consumed_date']}")


    # --- Section 2: Adding Items ---
    print("\n--- Section 2: Adding New Grocery Items ---")
    today_str = date.today().isoformat()
    
    manager.add_item_to_list(name="DB Apples", quantity_str="6 units", purchase_date_str=today_str, expiry_days=14,
                             category="Produce", subcategory="Fruit", par_level=5, max_holding_amount=20)
    manager.add_item_to_list(name="DB Bananas", quantity_str="12", purchase_date_str=today_str, expiry_days=5,
                             category="Produce", subcategory="Fruit") # Using defaults for par/max
    manager.add_item_to_list(name="DB Milk", quantity_str="1 gallon", purchase_date_str=today_str, expiry_days=7,
                             category="Dairy", par_level=1, max_holding_amount=2)
    
    print("\n--- Current inventory after additions: ---")
    for item in manager.get_current_inventory(): print(f"- {item['name']} ({item['quantity']})")

    # --- Section 3: Checking for Expiring Items ---
    print("\n--- Section 3: Checking for Expiring Items ---")
    manager.check_for_expiring_items(days_threshold=6)

    # --- Section 4: Consuming Items ---
    print("\n--- Section 4: Consuming Items ---")
    print("Consuming 2 DB Apples...")
    manager.consume_item("DB Apples", 2.0)
    print("Consuming 12 DB Bananas...")
    manager.consume_item("DB Bananas", 12.0)
    
    print("\n--- Current inventory after consumption: ---")
    for item in manager.get_current_inventory(): print(f"- {item['name']} ({item['quantity']})")
    print("\n--- Historical inventory after consumption: ---")
    for item in manager.get_historical_inventory(): print(f"- HIST: {item['name']} ({item['quantity_consumed_this_time']}) on {item['consumed_date']}")

    # --- Section 5: Demand Projection ---
    print("\n--- Section 5: Demand Projection ---")
    manager.project_demand("DB Apples")
    manager.project_demand("DB Milk")

    # --- Section 6: Exporting Data ---
    print("\n--- Section 6: Exporting All Data to CSV ---")
    manager.export_data_to_csv(filename_prefix="food_manager_db_demo_export")
    print(f"Data export process finished for DB demo.")

    print("\n\n--- Food Manager (SQLite DB) Demonstration Complete ---")
    print(f"Database file '{DB_FILE}' contains the data.")

    # Run migration if needed (idempotent parts are handled inside the method)
    print("\n--- Running Data Migration (if applicable) ---")
    migration_result = manager.migrate_text_categories_to_ids()
    if migration_result: # Check if not None
        print(migration_result.get("message", "Migration status unknown."))
    print("--- Migration Attempt Finished ---")

    # Note: To truly reset, delete DB_FILE before running again.
