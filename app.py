from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from Food_manager import InventoryManager
from RecipeManager import RecipeManager
from datetime import date, datetime, timedelta # Added timedelta
import openpyxl # For reading Excel files
from io import BytesIO # For handling file streams in memory
import os # For accessing environment variables
import shutil # For file operations (backup/restore)

app = Flask(__name__)
# Configure secret key: Use an environment variable for production, with a fallback for development.
# IMPORTANT: For production, set the FLASK_SECRET_KEY environment variable to a strong, random value.
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'a_default_fallback_secret_key')

# Define allowed extensions for file upload
ALLOWED_EXTENSIONS = {'xlsx'}

# Define constants for application
FUTURE_PROJECTION_HORIZON = 60 # Days for future inventory projection
PAST_ACTUALS_HORIZON = 7       # Days for past actuals summary

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Define the shared database file path
# Use an environment variable for the database path, with a default for development.
# IMPORTANT: For production or specific setups, set DATABASE_FILE_PATH environment variable.
DATABASE_FILE = os.environ.get('DATABASE_FILE_PATH', 'food_app.db')

# Instantiate Managers globally, using the shared database file
# The manager classes themselves will handle DB initialization (table creation IF NOT EXISTS)
manager = InventoryManager(db_filepath=DATABASE_FILE)
recipe_mngr = RecipeManager(db_filepath=DATABASE_FILE)

# --- Helper Functions ---
def _get_unique_item_names(include_historical=False):
    """
    Helper to get sorted unique item names from the products table.
    The `include_historical` parameter is less relevant now as we fetch from a definitive product list.
    """
    # manager.get_all_products() returns a list of product dictionaries
    all_products = manager.get_all_products()
    # Assuming each product dict has a 'name' key
    unique_names = set(product['name'] for product in all_products if 'name' in product)
    return sorted(list(unique_names))

# --- Flask Routes ---
@app.route('/')
def home():
    return render_template('home.html')

@app.route('/inventory-usage-links/')
def inventory_usage_links_view():
    return render_template('inventory_usage_links.html', title="Inventory & Usage Links")

@app.route('/inventory/current')
def current_inventory_view():
    # Retrieve filter and sort parameters from request.args
    search_term = request.args.get('search_term', '').strip()
    selected_category = request.args.get('category', '').strip()
    selected_purchase_location = request.args.get('purchase_location', '').strip()

    filter_is_below_par_str = request.args.get('filter_is_below_par', 'false').strip().lower()
    filter_is_below_par = filter_is_below_par_str == 'true'

    # Sort_by now defaults to product name, as expiry_date is not directly available for aggregated products
    sort_by = request.args.get('sort_by', 'p.name').strip()
    sort_order = request.args.get('sort_order', 'ASC').strip().upper()
    if sort_order not in ['ASC', 'DESC']:
        sort_order = 'ASC'

    try:
        page = int(request.args.get('page', 1))
        if page < 1: page = 1
    except ValueError:
        page = 1

    try:
        per_page = int(request.args.get('per_page', 20)) # Default 20 items per page
        if per_page < 1: per_page = 1
    except ValueError:
        per_page = 20

    # Fetch data from manager using SQL-filterable parameters
    # Renamed inventory_items_raw to products_raw
    products_raw = manager.get_current_inventory(
        search_term=search_term if search_term else None,
        category=selected_category if selected_category else None,
        purchase_location=selected_purchase_location if selected_purchase_location else None,
        sort_by=sort_by,
        sort_order=sort_order,
        page=page,
        per_page=per_page
    )

    # total_items is now the count of distinct products
    total_items = manager.get_current_inventory_count(
        search_term=search_term if search_term else None,
        category=selected_category if selected_category else None,
        purchase_location=selected_purchase_location if selected_purchase_location else None
    )

    # Process products and apply is_below_par filter if requested
    # Renamed processed_inventory_items to processed_products and item_dict to product_dict
    processed_products = []
    for product_dict in products_raw:
        product_processed = dict(product_dict)
        # numeric_quantity is now total_quantity from the aggregated data
        numeric_quantity = product_processed.get('total_quantity', 0.0)
        par_level = product_processed.get('par_level', 0.0)
        if par_level is None: par_level = 0.0
        # Ensure par_level is float for comparison
        else: par_level = float(par_level)
        
        product_processed['is_below_par'] = (numeric_quantity < par_level and par_level > 0)

        if filter_is_below_par:
            if product_processed['is_below_par']:
                processed_products.append(product_processed)
        else:
            processed_products.append(product_processed)

    # Pagination calculation (uses total_items from DB before Python-level 'is_below_par' filtering)
    # If filter_is_below_par is active, the actual number of items on the page might be less than per_page.
    # And total_pages might be misleading if the filter significantly reduces item count.
    # A more accurate pagination for client-side filtering would require adjusting total_items here.
    # For now, pagination reflects the count *before* the Python-side "is_below_par" filter.
    total_pages = (total_items + per_page - 1) // per_page if per_page > 0 else 1
    if page > total_pages and total_pages > 0:
        page = total_pages
        # Optionally re-fetch if strict item count per page is needed after Python filter.
        # products_raw = manager.get_current_inventory(...) with new page
        # then re-process for processed_products.
        # For now, we accept that the page might show fewer items if filtered in Python.
        # The user is on the 'last available page' according to DB count.

    # Fetch filter options
    categories_options = manager.get_current_inventory_categories() # Still relevant
    purchase_locations_options = manager.get_current_inventory_purchase_locations() # Still relevant

    today = date.today() # Still relevant for general display, not for expiry
    return render_template(
        'current_inventory.html',
        items=processed_products, # Pass processed_products
        today=today,
        timedelta=timedelta, # Keep if template uses it for other date logic
        current_page=page,
        total_pages=total_pages,
        per_page=per_page,
        search_term=search_term,
        selected_category=selected_category,
        selected_purchase_location=selected_purchase_location,
        # expiry_start_date and expiry_end_date removed from context
        filter_is_below_par=filter_is_below_par,
        sort_by=sort_by,
        sort_order=sort_order,
        categories=categories_options,
        purchase_locations=purchase_locations_options
    )

@app.route('/inventory/historical')
def historical_inventory_view():
    # Retrieve filter and sort parameters
    search_term = request.args.get('search_term', '').strip()
    selected_category = request.args.get('category', '').strip()
    consumed_start_date = request.args.get('consumed_start_date', '').strip()
    consumed_end_date = request.args.get('consumed_end_date', '').strip()

    sort_by = request.args.get('sort_by', 'consumed_date').strip()
    sort_order = request.args.get('sort_order', 'DESC').strip().upper()
    if sort_order not in ['ASC', 'DESC']:
        sort_order = 'DESC'

    try:
        page = int(request.args.get('page', 1))
        if page < 1: page = 1
    except ValueError:
        page = 1

    try:
        per_page = int(request.args.get('per_page', 20))
        if per_page < 1: per_page = 1
    except ValueError:
        per_page = 20

    # Fetch data from manager
    historical_items_data = manager.get_historical_inventory( # Renamed to avoid confusion in template if 'items' is ambiguous
        search_term=search_term if search_term else None,
        category=selected_category if selected_category else None,
        consumed_start_date=consumed_start_date if consumed_start_date else None,
        consumed_end_date=consumed_end_date if consumed_end_date else None,
        sort_by=sort_by,
        sort_order=sort_order,
        page=page,
        per_page=per_page
    )

    # The historical_items_data already contains all necessary fields including cost_of_goods_used
    # No additional processing needed here for that field if the manager method includes it.

    total_items = manager.get_historical_inventory_count(
        search_term=search_term if search_term else None,
        category=selected_category if selected_category else None,
        consumed_start_date=consumed_start_date if consumed_start_date else None,
        consumed_end_date=consumed_end_date if consumed_end_date else None
    )

    total_pages = (total_items + per_page - 1) // per_page if per_page > 0 else 1

    # Adjust page if out of bounds (e.g., after filters change total item count)
    if page > total_pages and total_pages > 0:
        page = total_pages
        # Re-fetch for the new valid page
        historical_items = manager.get_historical_inventory( # This should be historical_items_data
            search_term=search_term if search_term else None,
            category=selected_category if selected_category else None,
            consumed_start_date=consumed_start_date if consumed_start_date else None,
            consumed_end_date=consumed_end_date if consumed_end_date else None,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            per_page=per_page
        )
        historical_items_data = historical_items # Corrected assignment

    # Fetch filter options
    categories_options = manager.get_historical_inventory_categories()

    return render_template(
        'historical_inventory.html',
        items=historical_items_data, # Pass the fetched data as 'items' to the template
        current_page=page,
        total_pages=total_pages,
        per_page=per_page,
        search_term=search_term,
        selected_category=selected_category,
        consumed_start_date=consumed_start_date,
        consumed_end_date=consumed_end_date,
        sort_by=sort_by,
        sort_order=sort_order,
        categories=categories_options
    )

@app.route('/inventory/batches')
def inventory_batches_view():
    search_term = request.args.get('search_term', '').strip()
    selected_category = request.args.get('category', '').strip()
    sort_by = request.args.get('sort_by', 'product_name').strip()
    sort_order = request.args.get('sort_order', 'ASC').strip().upper()
    if sort_order not in ['ASC', 'DESC']:
        sort_order = 'ASC'

    try:
        page = int(request.args.get('page', 1))
        if page < 1: page = 1
    except ValueError:
        page = 1

    try:
        per_page = int(request.args.get('per_page', 20))
        if per_page < 1: per_page = 1
    except ValueError:
        per_page = 20

    start_purchase_date = request.args.get('start_purchase_date', '').strip()
    end_purchase_date = request.args.get('end_purchase_date', '').strip()
    start_expiry_date = request.args.get('start_expiry_date', '').strip()
    end_expiry_date = request.args.get('end_expiry_date', '').strip()

    inventory_batches_data = manager.get_inventory_batches_with_cost(
        search_term=search_term if search_term else None,
        category=selected_category if selected_category else None,
        sort_by=sort_by,
        sort_order=sort_order,
        page=page,
        per_page=per_page,
        start_purchase_date=start_purchase_date if start_purchase_date else None,
        end_purchase_date=end_purchase_date if end_purchase_date else None,
        start_expiry_date=start_expiry_date if start_expiry_date else None,
        end_expiry_date=end_expiry_date if end_expiry_date else None
    )
    total_items = manager.get_inventory_batches_with_cost_count(
        search_term=search_term if search_term else None,
        category=selected_category if selected_category else None,
        start_purchase_date=start_purchase_date if start_purchase_date else None,
        end_purchase_date=end_purchase_date if end_purchase_date else None,
        start_expiry_date=start_expiry_date if start_expiry_date else None,
        end_expiry_date=end_expiry_date if end_expiry_date else None
    )

    total_pages = (total_items + per_page - 1) // per_page if per_page > 0 else 1
    if page > total_pages and total_pages > 0:
        page = total_pages
        inventory_batches_data = manager.get_inventory_batches_with_cost(
            search_term=search_term if search_term else None,
            category=selected_category if selected_category else None,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            per_page=per_page,
            start_purchase_date=start_purchase_date if start_purchase_date else None,
            end_purchase_date=end_purchase_date if end_purchase_date else None,
            start_expiry_date=start_expiry_date if start_expiry_date else None,
            end_expiry_date=end_expiry_date if end_expiry_date else None
        )

    categories_options = manager.get_all_categories()

    return render_template(
        'inventory_batches.html',
        items=inventory_batches_data,
        current_page=page,
        total_pages=total_pages,
        per_page=per_page,
        search_term=search_term,
        selected_category=selected_category,
        sort_by=sort_by,
        sort_order=sort_order,
        categories=categories_options,
        title="Inventory Batches"
    )

@app.route('/inventory/consume', methods=['GET', 'POST'])
def consume_item_view():
    item_names = _get_unique_item_names()
    all_recipes = recipe_mngr.get_all_recipes()

    if request.method == 'POST':
        consumption_type = request.form.get('consumption_type', 'item')

        if consumption_type == 'recipe':
            recipe_name_to_consume = request.form.get('recipe_name_to_consume')
            num_batches_str = request.form.get('num_batches', '1')

            if not recipe_name_to_consume:
                flash("Please select a recipe to consume.", 'error')
                return render_template('consume_item.html', item_names=item_names, recipes=all_recipes, form_data=request.form, today=date.today().isoformat())

            try:
                num_batches = int(num_batches_str)
                if num_batches <= 0:
                    flash("Number of batches must be a positive integer.", 'error')
                    return render_template('consume_item.html', item_names=item_names, recipes=all_recipes, form_data=request.form, today=date.today().isoformat())
            except ValueError:
                flash("Invalid number of batches specified.", 'error')
                return render_template('consume_item.html', item_names=item_names, recipes=all_recipes, form_data=request.form, today=date.today().isoformat())

            return redirect(url_for('make_recipe_view',
                                    recipe_name=recipe_name_to_consume,
                                    num_batches=num_batches,
                                    origin_page='consume_item_page'))
        
        else:
            if request.is_json:
                data = request.get_json()
                items_to_consume = data.get('items')
                consumption_date_str = data.get('consumption_date', date.today().isoformat())

                if not items_to_consume or not isinstance(items_to_consume, list):
                    return jsonify({"success": False, "message": "Invalid data: 'items' array is required."}), 400

                if consumption_date_str:
                    try:
                        date.fromisoformat(consumption_date_str)
                    except ValueError:
                        return jsonify({"success": False, "message": f"Invalid consumption_date format: {consumption_date_str}. Use YYYY-MM-DD."}), 400

                results = manager.consume_multiple_items(items_to_consume, consumption_date_str=consumption_date_str)
                success_count = sum(1 for r in results if r.get("success"))
                failure_count = len(results) - success_count

                if failure_count == 0 and success_count > 0:
                    return jsonify({"success": True, "message": f"Successfully consumed {success_count} item(s).", "details": results})
                elif success_count > 0 and failure_count > 0:
                    return jsonify({"success": False, "message": f"Consumed {success_count} item(s), but failed for {failure_count} item(s).", "details": results}), 207
                elif failure_count > 0 and success_count == 0:
                     return jsonify({"success": False, "message": f"Failed to consume {failure_count} item(s).", "details": results}), 400
                else:
                    return jsonify({"success": False, "message": "No items were processed.", "details": results}), 400

            else:
                item_name = request.form.get('item_name')
                quantity_consumed_str = request.form.get('quantity_consumed')
                errors = []
                if not item_name:
                    errors.append("Item name is required.")
                numeric_quantity_consumed = None
                if not quantity_consumed_str:
                    errors.append("Quantity to consume is required.")
                else:
                    try:
                        numeric_quantity_consumed = float(quantity_consumed_str)
                        if numeric_quantity_consumed <= 0:
                            errors.append("Quantity to consume must be a positive number.")
                    except ValueError:
                        errors.append("Quantity to consume must be a valid number.")
                if errors:
                    for error in errors:
                        flash(error, 'error')
                    return render_template('consume_item.html', item_names=item_names, recipes=all_recipes, form_data=request.form)
                try:
                    result = manager.consume_item(item_name, numeric_quantity_consumed)
                    if result.get("success"):
                        flash(result.get("message", f"Consumption of '{item_name}' processed."), 'success')
                    else:
                        flash(result.get("message", f"Could not consume '{item_name}'."), 'error')
                    return redirect(url_for('current_inventory_view'))
                except Exception as e:
                    flash(f"An unexpected error occurred while consuming item: {e}", 'error')
                    return render_template('consume_item.html', item_names=item_names, recipes=all_recipes, form_data=request.form)
    return render_template('consume_item.html', item_names=item_names, recipes=all_recipes, form_data={}, today=date.today().isoformat())

@app.route('/inventory/upload_excel', methods=['GET', 'POST'])
def upload_excel_view():
    if request.method == 'POST':
        if 'excel_file' not in request.files:
            flash('No file part in the request.', 'error')
            return redirect(request.url)
        file = request.files['excel_file']
        if file.filename == '':
            flash('No selected file.', 'error')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            try:
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                header_row_values = [cell.value for cell in sheet[1]]
                header_map = {str(h).strip().lower(): idx for idx, h in enumerate(header_row_values) if h}
                required_headers = ['name', 'quantity', 'purchase date', 'expiry days']
                missing_headers = [req_h for req_h in required_headers if req_h not in header_map]
                if missing_headers:
                    flash(f"Missing required columns in Excel: {', '.join(missing_headers)}. Please check headers.", 'error')
                    return redirect(request.url)
                name_col = header_map['name']
                qty_col = header_map['quantity']
                pdate_col = header_map['purchase date']
                expdays_col = header_map['expiry days']
                category_col = header_map.get('category') 
                subcategory_col = header_map.get('subcategory')
                par_level_col = header_map.get('par level')
                max_holding_col = header_map.get('max holding amount')
                purchase_location_col = header_map.get('purchase location')
                unit_of_measure_col = header_map.get('unit of measure')
                cost_per_unit_col = header_map.get('cost_per_unit')
                vendor_col = header_map.get('vendor')
                items_added_count = 0
                error_messages = []
                items_pending_confirmation = []
                all_upload_warnings = []
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if len(row) <= max(name_col, qty_col, pdate_col, expdays_col):
                        error_messages.append(f"Row {row_idx}: Skipped due to insufficient columns.")
                        continue
                    name = str(row[name_col]).strip() if row[name_col] is not None else None
                    quantity = str(row[qty_col]).strip() if row[qty_col] is not None else None
                    purchase_date_val = row[pdate_col]
                    expiry_days_val = row[expdays_col]
                    category = str(row[category_col]).strip() if category_col is not None and row[category_col] is not None else None
                    subcategory = str(row[subcategory_col]).strip() if subcategory_col is not None and row[subcategory_col] is not None else None
                    par_level_val = row[par_level_col] if par_level_col is not None and row[par_level_col] is not None else "0"
                    max_holding_val = row[max_holding_col] if max_holding_col is not None and row[max_holding_col] is not None else "0"
                    purchase_location_val = str(row[purchase_location_col]).strip() if purchase_location_col is not None and row[purchase_location_col] is not None else None
                    unit_of_measure_val = str(row[unit_of_measure_col]).strip() if unit_of_measure_col is not None and row[unit_of_measure_col] is not None else None
                    cost_per_unit_val_str = str(row[cost_per_unit_col]).strip() if cost_per_unit_col is not None and row[cost_per_unit_col] is not None else None
                    vendor_val_str = str(row[vendor_col]).strip() if vendor_col is not None and row[vendor_col] is not None else None
                    if not name:
                        other_cells_have_data = any(
                            row[col_idx] for col_idx in [qty_col, pdate_col, expdays_col,
                                                         category_col, subcategory_col, par_level_col, max_holding_col, purchase_location_col, unit_of_measure_col]
                            if col_idx is not None and len(row) > col_idx and row[col_idx] is not None
                        )
                        if other_cells_have_data:
                            error_messages.append(f"Row {row_idx}: Name is missing but other data present. Skipped.")
                        elif not any(row[col_idx] for col_idx in [name_col, qty_col, pdate_col, expdays_col,
                                                                category_col, subcategory_col, par_level_col, max_holding_col, purchase_location_col, unit_of_measure_col]
                                     if col_idx is not None and len(row) > col_idx and row[col_idx] is not None):
                            continue
                        else:
                             error_messages.append(f"Row {row_idx}: Name is missing. Skipped.")
                        continue
                    row_errors = []
                    if quantity is None or quantity == "": row_errors.append("Quantity is missing.")
                    if purchase_date_val is None: row_errors.append("Purchase Date is missing.")
                    if expiry_days_val is None: row_errors.append("Expiry Days is missing.")
                    purchase_date_str = None
                    if isinstance(purchase_date_val, datetime):
                        purchase_date_str = purchase_date_val.strftime('%Y-%m-%d')
                    elif isinstance(purchase_date_val, str):
                        try:
                            date.fromisoformat(purchase_date_val)
                            purchase_date_str = purchase_date_val
                        except ValueError:
                            row_errors.append(f"Invalid Purchase Date format '{purchase_date_val}'. Use YYYY-MM-DD.")
                    elif purchase_date_val is not None: 
                         row_errors.append(f"Purchase Date '{purchase_date_val}' is not in YYYY-MM-DD text or Excel date format.")
                    expiry_days_int = None
                    if isinstance(expiry_days_val, (int, float)):
                        expiry_days_int = int(expiry_days_val)
                        if expiry_days_int < 0:
                            row_errors.append("Expiry Days must be a non-negative number.")
                            expiry_days_int = None
                    elif isinstance(expiry_days_val, str) and expiry_days_val.strip().lstrip('-').isdigit():
                        expiry_days_int = int(expiry_days_val.strip())
                        if expiry_days_int < 0:
                             row_errors.append("Expiry Days must be a non-negative number.")
                             expiry_days_int = None
                    elif expiry_days_val is not None:
                        row_errors.append(f"Expiry Days '{expiry_days_val}' must be a valid whole number.")
                    par_level_float = None
                    try:
                        par_level_val_cleaned = str(par_level_val).strip() if par_level_val is not None else "0"
                        par_level_float = float(par_level_val_cleaned)
                        if par_level_float < 0:
                            row_errors.append("Par Level must be a non-negative number.")
                            par_level_float = None
                    except (ValueError, TypeError):
                        row_errors.append(f"Invalid Par Level '{par_level_val}'. Must be a valid number.")
                    max_holding_float = None
                    try:
                        max_holding_val_cleaned = str(max_holding_val).strip() if max_holding_val is not None else "0"
                        max_holding_float = float(max_holding_val_cleaned)
                        if max_holding_float < 0:
                            row_errors.append("Max Holding Amount must be a non-negative number.")
                            max_holding_float = None
                    except (ValueError, TypeError):
                        row_errors.append(f"Invalid Max Holding Amount '{max_holding_val}'. Must be a valid number.")
                    purchase_location_to_pass = None
                    if purchase_location_val:
                        allowed_locations = ['Sobeys', 'Costco']
                        if purchase_location_val in allowed_locations:
                            purchase_location_to_pass = purchase_location_val
                        else:
                            row_errors.append(f"Invalid Purchase Location '{purchase_location_val}'. If provided, must be one of: {', '.join(allowed_locations)}.")
                    product_exists = manager.get_product_by_name(name)
                    if not product_exists and not unit_of_measure_val:
                        row_errors.append("Unit of Measure is required for new products.")
                    if row_errors:
                        error_messages.append(f"Row {row_idx} ('{name}'): " + "; ".join(row_errors))
                        continue
                    try:
                        result = manager.add_item_to_list(
                            name=str(name),
                            quantity_str=str(quantity),
                            purchase_date_str=purchase_date_str,
                            expiry_days=expiry_days_int,
                            category=category if category else None,
                            subcategory=subcategory if subcategory else None,
                            par_level=par_level_float,
                            max_holding_amount=max_holding_float,
                            purchase_location=purchase_location_to_pass,
                            unit_of_measure=unit_of_measure_val,
                            cost_per_unit_str=cost_per_unit_val_str,
                            vendor=vendor_val_str
                        )
                        if result.get("action_required"):
                            items_pending_confirmation.append({
                                "product_data": result["product_data"],
                                "confirmation_details": result["confirmation_details"],
                                "action_required": result["action_required"],
                                "row_idx": row_idx
                            })
                        elif result.get("success"):
                            items_added_count += 1
                        else:
                            error_messages.append(f"Row {row_idx} ('{name}'): {result.get('message', 'Unknown error')}")
                        if result.get("warnings"):
                            for warning in result.get("warnings"):
                                all_upload_warnings.append(f"Row {row_idx} ('{name}'): {warning}")
                    except ValueError as ve:
                         error_messages.append(f"Row {row_idx} ('{name}'): Validation Error - {str(ve)}")
                    except Exception as e:
                        error_messages.append(f"Row {row_idx} ('{name}'): Error adding to inventory - {str(e)}")
                session['items_pending_confirmation'] = items_pending_confirmation
                session['upload_warnings'] = all_upload_warnings
                if items_pending_confirmation:
                    flash(f"{len(items_pending_confirmation)} items require confirmation for new categories/subcategories.", 'info')
                    if items_added_count > 0: flash(f"Additionally, {items_added_count} items were added directly.", 'success')
                    if error_messages:
                        flash(f"{len(error_messages)} other rows had errors. Details below:", 'error')
                        for err_msg in error_messages[:5]: flash(err_msg, 'error_detail')
                    return redirect(url_for('upload_excel_view'))
                if items_added_count > 0:
                    flash(f"Successfully processed and added {items_added_count} items from the Excel file.", 'success')
                if all_upload_warnings:
                    flash(f"{len(all_upload_warnings)} warnings encountered during processing:", 'warning')
                    for warn_msg in all_upload_warnings[:10]: flash(warn_msg, 'warning_detail')
                if error_messages:
                    flash(f"{len(error_messages)} rows had errors and were not processed. Details below:", 'error')
                    for err_msg in error_messages[:5]: flash(err_msg, 'error_detail')
                if items_added_count == 0 and not error_messages and not all_upload_warnings and not items_pending_confirmation:
                     flash("No items were found or added from the file. The file might be empty or data starts after row 2.", 'info')
                session.pop('upload_warnings', None)
                return redirect(url_for('current_inventory_view'))
            except Exception as e:
                flash(f"An error occurred while processing the Excel file: {e}", 'error')
                return redirect(request.url)
        else:
            flash('Invalid file type. Please upload an .xlsx file.', 'error')
            return redirect(request.url)
    pending_items = session.get('items_pending_confirmation', [])
    upload_warnings = session.get('upload_warnings', [])
    if pending_items:
        return render_template('upload_excel.html', items_pending_confirmation=pending_items, upload_warnings=upload_warnings)
    return render_template('upload_excel.html', upload_warnings=upload_warnings)

@app.route('/upload_recipes_excel', methods=['GET', 'POST'])
# ... (rest of the file is unchanged) ...

# The specific modification for edit_inventory_view is below this line.
# Note: The SEARCH block for the full function replacement must be the *current* state.
# The following is the *intended new state* of the function.

@app.route('/inventory/edit', methods=['GET', 'POST'])
def edit_inventory_view():
    products_for_dropdown = manager.get_all_products()
    selected_product_id = request.args.get('product_id')
    active_inventory_batches = []
    recently_consumed_batches = []
    selected_product_name = None

    if selected_product_id:
        try:
            valid_product_id = int(selected_product_id)
            # Fetch active batches - REMOVING LIMIT and using default sort (FEFO) for diagnostics
            active_inventory_batches = manager.get_inventory_batches_for_product(
                product_id=valid_product_id,
                limit=None, # Fetch all active batches for this product
                order_by_id_desc=False, # Use default sort (expiry_date ASC)
                order_by_purchase_desc=False # Ensure this is also false for default sort
            )

            recently_consumed_batches = manager.get_recently_consumed_batch_info(
                product_id=valid_product_id,
                limit=10
            )

            selected_product = manager.get_product(valid_product_id)
            if selected_product:
                selected_product_name = selected_product['name']
        except ValueError:
            flash("Invalid product ID format provided.", "error")
            selected_product_id = None
            active_inventory_batches = []
            recently_consumed_batches = []

    if request.method == 'POST':
        include_in_projections = request.form.get('include_in_projections') == 'true'
        page_product_id_for_redirect = request.form.get('product_id_for_redirect', selected_product_id)

        active_batch_idx = 0
        while True:
            batch_id_str = request.form.get(f'active_batch_id_{active_batch_idx}')
            if batch_id_str is None:
                break
            new_quantity_str = request.form.get(f'active_quantity_{active_batch_idx}')
            new_purchase_date_str = request.form.get(f'active_purchase_date_{active_batch_idx}')
            new_expiry_date_str = request.form.get(f'active_expiry_date_{active_batch_idx}')

            if not batch_id_str:
                active_batch_idx += 1
                continue

            try:
                batch_id = int(batch_id_str)
                result = manager.adjust_inventory_batch(
                    batch_id=batch_id,
                    new_quantity_str=new_quantity_str,
                    new_purchase_date_str=new_purchase_date_str,
                    new_expiry_date_str=new_expiry_date_str,
                    include_in_projections=include_in_projections
                )
                if result.get("success"):
                    flash(result["message"], 'success')
                else:
                    flash(f"Error adjusting active batch ID {batch_id}: {result.get('message', 'Unknown error.')}", 'error')
            except ValueError:
                flash(f"Invalid Batch ID format for active batch: {batch_id_str}", "error")
            except Exception as e:
                flash(f"Unexpected error processing active batch ID {batch_id_str}: {e}", "error")
            active_batch_idx += 1

        consumed_batch_idx = 0
        while True:
            quantity_to_return_str = request.form.get(f'consumed_return_quantity_{consumed_batch_idx}')

            if quantity_to_return_str is None or quantity_to_return_str.strip() == "":
                if request.form.get(f'consumed_product_id_{consumed_batch_idx}') is None:
                    break
                consumed_batch_idx += 1
                continue

            consumed_product_id_str = request.form.get(f'consumed_product_id_{consumed_batch_idx}')
            consumed_product_name = request.form.get(f'consumed_product_name_{consumed_batch_idx}')
            consumed_purchase_date = request.form.get(f'consumed_purchase_date_{consumed_batch_idx}')
            consumed_expiry_date = request.form.get(f'consumed_expiry_date_{consumed_batch_idx}')
            consumed_original_qty_str = request.form.get(f'consumed_original_qty_str_{consumed_batch_idx}')

            if not all([consumed_product_id_str, consumed_product_name, consumed_purchase_date, consumed_expiry_date, quantity_to_return_str]):
                if quantity_to_return_str.strip() != "":
                     flash(f"Error: Missing hidden data for returning consumed batch '{consumed_product_name or 'Unknown'}'. Action skipped.", 'error')
                consumed_batch_idx += 1
                continue

            try:
                if float(quantity_to_return_str) <= 0:
                    consumed_batch_idx += 1
                    continue

                result = manager.return_consumed_batch_to_stock(
                    product_id=int(consumed_product_id_str),
                    product_name=consumed_product_name,
                    purchase_date_str=consumed_purchase_date,
                    expiry_date_str=consumed_expiry_date,
                    original_quantity_at_creation_str=consumed_original_qty_str,
                    quantity_to_return_str=quantity_to_return_str,
                    include_in_projections=include_in_projections
                )
                if result.get("success"):
                    flash(result["message"], 'success')
                else:
                    flash(f"Error returning consumed batch of '{consumed_product_name}': {result.get('message', 'Unknown error.')}", 'error')
            except ValueError:
                flash(f"Invalid data format for returning consumed batch of '{consumed_product_name}'.", "error")
            except Exception as e:
                flash(f"Unexpected error returning consumed batch of '{consumed_product_name}': {e}", "error")

            consumed_batch_idx += 1

        if page_product_id_for_redirect:
            return redirect(url_for('edit_inventory_view', product_id=page_product_id_for_redirect))
        else:
            flash("No product context for redirect, returning to general edit page.", "warning")
            return redirect(url_for('edit_inventory_view'))

    return render_template('edit_inventory.html',
                           products=products_for_dropdown,
                           selected_product_id=selected_product_id,
                           selected_product_name=selected_product_name,
                           active_inventory_batches=active_inventory_batches,
                           recently_consumed_batches=recently_consumed_batches)

# --- Category and Subcategory Management Route ---
# ... (rest of app.py, which is unchanged)
@app.route('/manage_categories', methods=['GET', 'POST'])
def manage_categories_view():
    if request.method == 'POST':
        action_type = request.form.get('action_type')
        if action_type == 'add_category':
            category_name = request.form.get('category_name', '').strip()
            if category_name:
                result = manager.add_category(category_name)
                if result.get('success'):
                    flash(result['message'], 'success')
                else:
                    flash(result['message'], 'error')
            else:
                flash("Category name cannot be empty.", 'error')

        elif action_type == 'add_subcategory':
            subcategory_name = request.form.get('subcategory_name', '').strip()
            category_id_str = request.form.get('category_id')
            if subcategory_name and category_id_str:
                try:
                    category_id = int(category_id_str)
                    result = manager.add_subcategory(subcategory_name, category_id)
                    if result.get('success'):
                        flash(result['message'], 'success')
                    else:
                        flash(result['message'], 'error')
                except ValueError:
                    flash("Invalid Category ID.", 'error')
            else:
                flash("Subcategory name and Category ID are required.", 'error')
        else:
            flash("Invalid action.", 'error')
        return redirect(url_for('manage_categories_view'))

    # GET request
    categories_data = manager.get_all_categories_with_subcategories()
    return render_template('manage_categories.html', categories_data=categories_data)

@app.route('/product_modal_details/<int:product_id>', methods=['GET'])
def product_modal_details(product_id):
    product_details = manager.get_product_details(product_id)

    if not product_details:
        return jsonify({"error": "Product not found"}), 404

    daily_consumption = manager.get_daily_consumption(product_id)
    monthly_consumption = manager.get_monthly_consumption(product_id)
    daily_inventory_history = manager.get_daily_inventory_history(product_id)
    recipes_containing_product = []
    if product_details.get('name'):
        recipes_containing_product = recipe_mngr.get_recipes_for_product(product_details['name'])
    inventory_concerns = manager.get_inventory_concerns(product_id)
    shopping_list_amount_today = 0.0
    SOBEYS_FREQUENCY_WEEKS = 1
    COSTCO_FREQUENCY_WEEKS = 3
    par_level = product_details.get('par_level', 0.0)
    if par_level is None: par_level = 0.0
    else: par_level = float(par_level)
    purchase_location = product_details.get('purchase_location')
    current_on_hand_inventory = product_details.get('current_on_hand_inventory', 0.0)
    projection_days_for_item = 0
    if purchase_location == 'Sobeys':
        projection_days_for_item = SOBEYS_FREQUENCY_WEEKS * 7
    elif purchase_location == 'Costco':
        projection_days_for_item = COSTCO_FREQUENCY_WEEKS * 7
    if par_level > 0 and projection_days_for_item > 0:
        demand_projection_for_shopping = manager.project_demand(
            product_id,
            lookback_days=30,
            projection_days=projection_days_for_item
        )
        avg_daily_consumption_for_shopping = 0.0
        if demand_projection_for_shopping.get("success"):
            avg_daily_consumption_for_shopping = demand_projection_for_shopping.get('avg_daily_consumption', 0.0)
        target_stock_after_shopping = par_level + (avg_daily_consumption_for_shopping * projection_days_for_item)
        recommended_purchase_amount = target_stock_after_shopping - current_on_hand_inventory
        shopping_list_amount_today = max(0, round(recommended_purchase_amount, 2))
    future_projection_result = manager.get_future_inventory_projection(product_id, projection_days=FUTURE_PROJECTION_HORIZON)
    final_future_projection_data = []
    if isinstance(future_projection_result, dict) and future_projection_result.get("success") is False:
        app.logger.error(f"Failed to get future inventory projection for product {product_id}: {future_projection_result.get('message')}")
    elif isinstance(future_projection_result, list):
        final_future_projection_data = future_projection_result
    past_actuals_result = manager.get_past_actual_inventory_summary(product_id, days_past=PAST_ACTUALS_HORIZON)
    final_past_actual_data = []
    if isinstance(past_actuals_result, dict) and past_actuals_result.get("success") is False:
        app.logger.error(f"Failed to get past actuals summary for product {product_id}: {past_actuals_result.get('message')}")
    elif isinstance(past_actuals_result, list):
        final_past_actual_data = past_actuals_result
    data_to_return = {
        "product_details": product_details,
        "daily_consumption": daily_consumption,
        "monthly_consumption": monthly_consumption,
        "daily_inventory_history": daily_inventory_history,
        "recipes_containing_product": recipes_containing_product,
        "inventory_concerns": inventory_concerns,
        "shopping_list_amount_today": shopping_list_amount_today,
        "future_projection_data": final_future_projection_data,
        "past_actual_data": final_past_actual_data
    }
    app.logger.debug(f"Data for modal product ID {product_id}: {data_to_return}")
    return jsonify(data_to_return)

@app.route('/upload_products_excel', methods=['GET', 'POST'])
def upload_products_excel_view():
    if request.method == 'POST':
        if 'excel_file' not in request.files:
            flash('No file part in the request.', 'error')
            return redirect(request.url)
        file = request.files['excel_file']
        if file.filename == '':
            flash('No selected file.', 'error')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            overwrite_logic = request.form.get('overwrite_logic_choice', 'skip')
            try:
                result = manager.upload_products_excel(file.stream, overwrite_logic)
                if result.get("errors"):
                    for error_msg in result["errors"][:10]:
                        flash(error_msg, 'error_detail')
                    if len(result["errors"]) > 10:
                        flash(f"...and {len(result['errors']) - 10} more errors.", 'error_detail')
                if result.get("added") > 0:
                    flash(f"Successfully added {result['added']} new products.", 'success')
                if result.get("updated") > 0:
                    flash(f"Successfully updated {result['updated']} existing products.", 'success')
                if result.get("skipped") > 0:
                    flash(f"{result['skipped']} products were skipped (duplicates, 'skip' logic chosen).", 'info')
                if not result.get("errors") and result.get("added") == 0 and result.get("updated") == 0 and result.get("skipped") == 0:
                     flash("No products were added, updated, or skipped. File might be empty or data already matches existing entries (and skip logic was chosen).", 'info')
                if result.get("errors"):
                     flash("Product upload completed with errors. Please check messages above.", 'warning')
                     return redirect(url_for('upload_products_excel_view'))
                else:
                    return redirect(url_for('list_products_view'))
            except Exception as e:
                app.logger.error(f"Error processing product Excel upload: {e}", exc_info=True)
                flash(f"An unexpected error occurred during product upload: {str(e)}", 'error')
                return redirect(request.url)
        else:
            flash('Invalid file type. Please upload an .xlsx file.', 'error')
            return redirect(request.url)
    return render_template('upload_products_excel.html', title="Upload Products from Excel")

@app.route('/upload_historical_excel', methods=['GET', 'POST'])
def upload_historical_excel_view():
    if request.method == 'POST':
        if 'excel_file' not in request.files:
            flash('No file part in the request.', 'error')
            return redirect(request.url)
        file = request.files['excel_file']
        if file.filename == '':
            flash('No selected file.', 'error')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            try:
                result = manager.upload_historical_inventory_excel(file.stream)
                errors = result.get("errors", [])
                added_count = result.get("added", 0)
                if errors:
                    flash(f"Historical data upload completed with {len(errors)} errors:", 'error')
                    for error_msg in errors[:10]:
                        flash(error_msg, 'error_detail')
                    if len(errors) > 10:
                        flash(f"...and {len(errors) - 10} more errors.", 'error_detail')
                if added_count > 0:
                    flash(f"Successfully added {added_count} historical consumption records.", 'success')
                if not errors and added_count == 0:
                     flash("No new historical records were added. File might be empty, data might be invalid, or headers incorrect.", 'info')
                if errors:
                     return redirect(url_for('upload_historical_excel_view'))
                else:
                    return redirect(url_for('historical_inventory_view'))
            except Exception as e:
                app.logger.error(f"Error processing historical inventory Excel upload: {e}", exc_info=True)
                flash(f"An unexpected error occurred during historical data upload: {str(e)}", 'error')
                return redirect(request.url)
        else:
            flash('Invalid file type. Please upload an .xlsx file.', 'error')
            return redirect(request.url)
    return render_template('upload_historical_excel.html', title="Upload Historical Inventory from Excel")

@app.route('/upload_production_items_excel', methods=['GET', 'POST'])
def upload_production_items_excel_view():
    if request.method == 'POST':
        if 'excel_file' not in request.files:
            flash('No file part in the request.', 'error')
            return redirect(request.url)
        file = request.files['excel_file']
        if file.filename == '':
            flash('No selected file.', 'error')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            try:
                result = manager.upload_production_items_excel(file.stream)
                errors = result.get("errors", [])
                added_count = result.get("added", 0)
                warnings = result.get("warnings", [])
                if warnings:
                    flash(f"Production item upload completed with {len(warnings)} warnings:", 'warning')
                    for warn_msg in warnings[:5]:
                        flash(warn_msg, 'warning_detail')
                if errors:
                    flash(f"Production item upload completed with {len(errors)} errors:", 'error')
                    for error_msg in errors[:10]:
                        flash(error_msg, 'error_detail')
                    if len(errors) > 10:
                        flash(f"...and {len(errors) - 10} more errors.", 'error_detail')
                if added_count > 0:
                    flash(f"Successfully added {added_count} production items.", 'success')
                if not errors and not warnings and added_count == 0 :
                     flash("No new production items were added. File might be empty or data invalid.", 'info')
                if errors:
                     return redirect(url_for('upload_production_items_excel_view'))
                else:
                    return redirect(url_for('garden_list_view'))
            except Exception as e:
                app.logger.error(f"Error processing production items Excel upload: {e}", exc_info=True)
                flash(f"An unexpected error occurred during production items upload: {str(e)}", 'error')
                return redirect(request.url)
        else:
            flash('Invalid file type. Please upload an .xlsx file.', 'error')
            return redirect(request.url)
    return render_template('upload_production_items_excel.html', title="Upload Production Items from Excel")

@app.route('/export_data', methods=['GET', 'POST'])
def export_data_view():
    if request.method == 'POST':
        selected_table = request.form.get('selected_table')
        if selected_table == "products":
            try:
                products_data = manager.get_all_products_export()
                if not products_data:
                    flash("No product data found to export.", "info")
                    return redirect(url_for('export_data_view'))
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Products Export"
                headers = list(products_data[0].keys())
                ws.append(headers)
                for product_row in products_data:
                    row_values = [product_row.get(header) for header in headers]
                    ws.append(row_values)
                excel_stream = BytesIO()
                wb.save(excel_stream)
                excel_stream.seek(0)
                return send_file(
                    excel_stream,
                    as_attachment=True,
                    download_name='products_export.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            except Exception as e:
                app.logger.error(f"Error exporting product data: {e}", exc_info=True)
                flash(f"An error occurred while exporting product data: {str(e)}", "error")
                return redirect(url_for('export_data_view'))
        elif selected_table == "inventory_batches":
            export_start_date = request.form.get('export_start_date')
            export_end_date = request.form.get('export_end_date')
            try:
                batches_data = manager.get_all_inventory_batches_export(
                    start_date_str=export_start_date if export_start_date else None,
                    end_date_str=export_end_date if export_end_date else None
                )
                if not batches_data:
                    flash("No inventory batch data found for the selected criteria.", "info")
                    return redirect(url_for('export_data_view'))
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Inventory Batches Export"
                headers = list(batches_data[0].keys())
                ws.append(headers)
                for batch_row in batches_data:
                    row_values = [batch_row.get(header) for header in headers]
                    ws.append(row_values)
                excel_stream = BytesIO()
                wb.save(excel_stream)
                excel_stream.seek(0)
                return send_file(
                    excel_stream,
                    as_attachment=True,
                    download_name='inventory_batches_export.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            except Exception as e:
                app.logger.error(f"Error exporting inventory batch data: {e}", exc_info=True)
                flash(f"An error occurred while exporting inventory batch data: {str(e)}", "error")
                return redirect(url_for('export_data_view'))
        elif selected_table == "historical_inventory":
            export_start_date = request.form.get('export_start_date')
            export_end_date = request.form.get('export_end_date')
            try:
                historical_data = manager.get_historical_inventory(
                    export_all=True,
                    export_start_date_str=export_start_date if export_start_date else None,
                    export_end_date_str=export_end_date if export_end_date else None
                )
                if not historical_data:
                    flash("No historical inventory data found for the selected criteria.", "info")
                    return redirect(url_for('export_data_view'))
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Historical Inventory Export"
                headers = list(historical_data[0].keys())
                ws.append(headers)
                for data_row in historical_data:
                    row_values = [data_row.get(header) for header in headers]
                    ws.append(row_values)
                excel_stream = BytesIO()
                wb.save(excel_stream)
                excel_stream.seek(0)
                return send_file(
                    excel_stream,
                    as_attachment=True,
                    download_name='historical_inventory_export.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            except Exception as e:
                app.logger.error(f"Error exporting historical inventory data: {e}", exc_info=True)
                flash(f"An error occurred while exporting historical inventory data: {str(e)}", "error")
                return redirect(url_for('export_data_view'))
        elif selected_table == "recipes":
            try:
                recipes_data = recipe_mngr.get_all_recipes(export_all=True)
                if not recipes_data:
                    flash("No recipes found to export.", "info")
                    return redirect(url_for('export_data_view'))
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Recipes Export"
                headers = list(recipes_data[0].keys())
                ws.append(headers)
                for recipe_row in recipes_data:
                    row_values = [recipe_row.get(header) for header in headers]
                    ws.append(row_values)
                excel_stream = BytesIO()
                wb.save(excel_stream)
                excel_stream.seek(0)
                return send_file(
                    excel_stream,
                    as_attachment=True,
                    download_name='recipes_export.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            except Exception as e:
                app.logger.error(f"Error exporting recipes data: {e}", exc_info=True)
                flash(f"An error occurred while exporting recipes data: {str(e)}", "error")
                return redirect(url_for('export_data_view'))
        elif selected_table == "recipe_ingredients":
            try:
                ingredients_data = recipe_mngr.get_all_recipe_ingredients_export()
                if not ingredients_data:
                    flash("No recipe ingredients found to export.", "info")
                    return redirect(url_for('export_data_view'))
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Recipe Ingredients Export"
                headers = list(ingredients_data[0].keys())
                ws.append(headers)
                for ingredient_row in ingredients_data:
                    row_values = [ingredient_row.get(header) for header in headers]
                    ws.append(row_values)
                excel_stream = BytesIO()
                wb.save(excel_stream)
                excel_stream.seek(0)
                return send_file(
                    excel_stream,
                    as_attachment=True,
                    download_name='recipe_ingredients_export.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            except Exception as e:
                app.logger.error(f"Error exporting recipe ingredients data: {e}", exc_info=True)
                flash(f"An error occurred while exporting recipe ingredients data: {str(e)}", "error")
                return redirect(url_for('export_data_view'))
        elif selected_table == "production_items":
            try:
                production_data = manager.get_all_production_items_export()
                if not production_data:
                    flash("No production items (garden) found to export.", "info")
                    return redirect(url_for('export_data_view'))
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Production Items Export"
                headers = list(production_data[0].keys())
                ws.append(headers)
                for item_row in production_data:
                    row_values = [item_row.get(header) for header in headers]
                    ws.append(row_values)
                excel_stream = BytesIO()
                wb.save(excel_stream)
                excel_stream.seek(0)
                return send_file(
                    excel_stream,
                    as_attachment=True,
                    download_name='production_items_export.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            except Exception as e:
                app.logger.error(f"Error exporting production items data: {e}", exc_info=True)
                flash(f"An error occurred while exporting production items data: {str(e)}", "error")
                return redirect(url_for('export_data_view'))
        elif selected_table == "categories":
            try:
                categories_data = manager.get_all_categories_export()
                if not categories_data:
                    flash("No categories found to export.", "info")
                    return redirect(url_for('export_data_view'))
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Categories Export"
                headers = list(categories_data[0].keys())
                ws.append(headers)
                for category_row in categories_data:
                    row_values = [category_row.get(header) for header in headers]
                    ws.append(row_values)
                excel_stream = BytesIO()
                wb.save(excel_stream)
                excel_stream.seek(0)
                return send_file(
                    excel_stream,
                    as_attachment=True,
                    download_name='categories_export.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            except Exception as e:
                app.logger.error(f"Error exporting categories data: {e}", exc_info=True)
                flash(f"An error occurred while exporting categories data: {str(e)}", "error")
                return redirect(url_for('export_data_view'))
        elif selected_table == "subcategories": # Duplicated block, ensure this is intended or fixed if one was meant for something else.
            try:
                subcategories_data = manager.get_all_subcategories_export()
                if not subcategories_data:
                    flash("No subcategories found to export.", "info")
                    return redirect(url_for('export_data_view'))
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Subcategories Export"
                headers = list(subcategories_data[0].keys())
                ws.append(headers)
                for subcategory_row in subcategories_data:
                    row_values = [subcategory_row.get(header) for header in headers]
                    ws.append(row_values)
                excel_stream = BytesIO()
                wb.save(excel_stream)
                excel_stream.seek(0)
                return send_file(
                    excel_stream,
                    as_attachment=True,
                    download_name='subcategories_export.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            except Exception as e:
                app.logger.error(f"Error exporting subcategories data: {e}", exc_info=True)
                flash(f"An error occurred while exporting subcategories data: {str(e)}", "error")
                return redirect(url_for('export_data_view'))
        else:
            flash(f"Export functionality for '{selected_table}' is not yet implemented.", "warning")
            return redirect(url_for('export_data_view'))
    return render_template('export_data.html', title="Export Data")

from functools import wraps

def admin_route_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if os.environ.get('FLASK_ENABLE_ADMIN_ROUTES', 'false').lower() != 'true':
            return "Admin functionality is disabled. Set FLASK_ENABLE_ADMIN_ROUTES=true to enable.", 403
        return f(*args, **kwargs)
    return decorated_function

@app.route('/admin/backup_restore', methods=['GET'])
@admin_route_required
def backup_restore_view():
    return render_template('backup_restore.html', title="Backup and Restore")

@app.route('/admin/create_backup', methods=['GET'])
@admin_route_required
def create_backup_view():
    try:
        source_db_path = manager.db_filepath
        if source_db_path == ":memory:":
            flash("Cannot backup an in-memory database.", "error")
            return redirect(url_for('backup_restore_view'))
        if not os.path.exists(source_db_path):
            flash(f"Database file not found at {source_db_path}.", "error")
            return redirect(url_for('backup_restore_view'))
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_dir = os.path.join(os.path.dirname(app.root_path), 'instance', 'backups')
        os.makedirs(backup_dir, exist_ok=True)
        backup_filename = f"food_app_backup_{timestamp}.db"
        backup_filepath = os.path.join(backup_dir, backup_filename)
        shutil.copy2(source_db_path, backup_filepath)
        return send_file(backup_filepath, as_attachment=True, download_name=backup_filename)
    except Exception as e:
        app.logger.error(f"Error creating database backup: {e}", exc_info=True)
        flash(f"An error occurred while creating the backup: {str(e)}", "error")
        return redirect(url_for('backup_restore_view'))

@app.route('/admin/restore_from_backup', methods=['POST'])
@admin_route_required
def restore_from_backup_view():
    if 'backup_file' not in request.files:
        flash('No backup file part in the request.', 'error')
        return redirect(url_for('backup_restore_view'))
    file = request.files['backup_file']
    if file.filename == '':
        flash('No backup file selected.', 'error')
        return redirect(url_for('backup_restore_view'))
    if file and file.filename.endswith('.db'):
        try:
            current_db_path = manager.db_filepath
            if current_db_path == ":memory:":
                flash("Cannot restore an in-memory database.", "error")
                return redirect(url_for('backup_restore_view'))
            backup_dir = os.path.join(os.path.dirname(app.root_path), 'instance', 'backups')
            os.makedirs(backup_dir, exist_ok=True)
            pre_restore_backup_filename = f"food_app_pre_restore_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
            pre_restore_backup_filepath = os.path.join(backup_dir, pre_restore_backup_filename)
            if os.path.exists(current_db_path):
                shutil.copy2(current_db_path, pre_restore_backup_filepath)
                flash(f"Existing database backed up to {pre_restore_backup_filename} before restore.", "info")
            file.save(current_db_path)
            flash(f"Database successfully restored from '{file.filename}'. Please restart the application for changes to take effect.", "success")
        except Exception as e:
            app.logger.error(f"Error restoring database: {e}", exc_info=True)
            flash(f"An error occurred during database restore: {str(e)}", "error")
    else:
        flash('Invalid file type. Please upload a .db file.', 'error')
    return redirect(url_for('backup_restore_view'))

if __name__ == '__main__':
    instance_path = os.path.join(os.path.dirname(app.root_path), 'instance')
    if not os.path.exists(instance_path):
        os.makedirs(instance_path, exist_ok=True)
        print(f"Created instance folder at: {instance_path}")
    app.run(host='0.0.0.0', port=8080, debug=False, use_reloader=False)
